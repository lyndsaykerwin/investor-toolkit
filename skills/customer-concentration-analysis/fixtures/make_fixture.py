#!/usr/bin/env python3
"""Generate the synthetic self-test fixture — fully fictional data.

example_saas_arr.xlsx: a wide customer×year grid on a sheet named
"Annual Recurring Revenue" with an explicit ARR title. Per-customer values are
deliberately small ($8k–$95k) so a magnitude-only heuristic would wrongly guess
MRR — the fixture exists to prove the detector trusts the LABEL, not the scale.
"""
import os
from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))

# (customer, 2023, 2024, 2025-current) — fictional, small-dollar on purpose
ROWS = [
    ("Northwind Trading", 60000, 78000, 95000),
    ("Globex Systems", 50000, 55000, 60000),
    ("Initech Software", 40000, 44000, 48000),
    ("Umbrella Analytics", 30000, 35000, 42000),
    ("Soylent Foods", 28000, 30000, 33000),
    ("Hooli Cloud", 25000, 28000, 31000),
    ("Vehement Capital", 22000, 26000, 30000),
    ("Stark Industries", 20000, 24000, 28000),
    ("Wonka Labs", 18000, 22000, 25000),
    ("Gekko Partners", 15000, 18000, 22000),
    ("Pied Piper", 12000, 14000, 16000),
    ("Wayne Enterprises", 10000, 11000, 12000),
    ("Acme Corp", 9000, 9500, 10000),
    ("Dunder Mifflin", 8000, 8500, 9000),
    ("Bluth Company", 7000, 7500, 8000),
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Annual Recurring Revenue"
    ws["A1"] = "Example SaaS Co. — Annual Recurring Revenue by Customer (USD)"
    ws["A2"] = "Values are annual recurring revenue. Fiscal year ending Dec."
    ws.append([])
    ws.append(["Customer", "FY2023", "FY2024", "FY2025"])  # header row 4
    for r in ROWS:
        ws.append(list(r))
    out = os.path.join(HERE, "example_saas_arr.xlsx")
    wb.save(out)
    total = sum(r[3] for r in ROWS)
    top10 = sum(sorted((r[3] for r in ROWS), reverse=True)[:10])
    print(f"Wrote {out}")
    print(f"  {len(ROWS)} customers, current (FY2025) total ARR = {total:,}")
    print(f"  Top 10 = {top10:,} ({top10/total:.1%}), remaining = {total-top10:,} "
          f"({(total-top10)/total:.1%})")


if __name__ == "__main__":
    main()
