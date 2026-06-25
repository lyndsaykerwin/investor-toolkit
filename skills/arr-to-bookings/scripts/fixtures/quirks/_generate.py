"""Generate 11 synthetic customer-revenue Excel fixtures, each with ONE layout quirk."""
import datetime as dt
from openpyxl import Workbook

OUT = "/Users/lyndsay/.claude/skills/ARR-to-bookings/fixtures/quirks"

# ---------------------------------------------------------------------------
# Shared synthetic data: 10 customers x 36 monthly periods (Jan-2021 .. Dec-2023)
# Each customer has a start month (some are late = new logos) and a monthly MRR
# that may step up over time (= upsell). Values are realistic MRR (500-15000).
# ---------------------------------------------------------------------------
CUSTOMERS = [
    "Acme 01", "Beacon 02", "Cobalt 03", "Delta 04", "Ember 05",
    "Falcon 06", "Granite 07", "Harbor 08", "Ionix 09", "Juniper 10",
]

# (start_month_index 0..35, base_mrr, upsell_step_month or None, upsell_new_mrr)
PROFILES = {
    "Acme 01":    (0,  5000,  12, 7500),    # steady, upsell after a year
    "Beacon 02":  (0,  2200,  None, None),  # flat
    "Cobalt 03":  (3,  9900,  18, 12400),   # late start, big upsell
    "Delta 04":   (0,   800,  6,  1500),    # small, early upsell
    "Ember 05":   (0, 14000,  None, None),  # large flat anchor
    "Falcon 06":  (6,  1542,  24, 3100),    # new logo mid-2021
    "Granite 07": (0,  3300,  30, 4800),    # late upsell
    "Harbor 08":  (12, 6700,  None, None),  # new logo 2022
    "Ionix 09":   (0,  1100,  10, 2050),    # upsell
    "Juniper 10": (18,  500,  None, None),  # late, small new logo 2022 H2
}

N_MONTHS = 36
MONTHS = [dt.date(2021, 1, 1)]
for _ in range(N_MONTHS - 1):
    y, m = MONTHS[-1].year, MONTHS[-1].month
    MONTHS.append(dt.date(y + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1))
# month-end dates for date-typed headers
def eom(d):
    y, m = d.year, d.month
    nxt = dt.date(y + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1)
    return nxt - dt.timedelta(days=1)
MONTH_ENDS = [eom(d) for d in MONTHS]


def mrr(cust, month_idx):
    start, base, up_m, up_new = PROFILES[cust]
    if month_idx < start:
        return 0
    if up_m is not None and month_idx >= up_m:
        return up_new
    return base


def matrix():
    """Return rows: each row is list of MRR values across 36 months for a customer."""
    return {c: [mrr(c, i) for i in range(N_MONTHS)] for c in CUSTOMERS}


M = matrix()

# ---------------------------------------------------------------------------
# q01: text dates "Mon YY" (Jan 21, Feb 21)
# ---------------------------------------------------------------------------
def q01():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 1, "Customer")
    for j, d in enumerate(MONTHS):
        ws.cell(1, 2 + j, d.strftime("%b %y"))   # "Jan 21"
    for i, c in enumerate(CUSTOMERS):
        ws.cell(2 + i, 1, c)
        for j, v in enumerate(M[c]):
            ws.cell(2 + i, 2 + j, v)
    wb.save(f"{OUT}/q01_text_dates_mon_yy.xlsx")


# ---------------------------------------------------------------------------
# q02: text dates ISO "2021-01"
# ---------------------------------------------------------------------------
def q02():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 1, "Customer")
    for j, d in enumerate(MONTHS):
        ws.cell(1, 2 + j, d.strftime("%Y-%m"))   # "2021-01"
    for i, c in enumerate(CUSTOMERS):
        ws.cell(2 + i, 1, c)
        for j, v in enumerate(M[c]):
            ws.cell(2 + i, 2 + j, v)
    wb.save(f"{OUT}/q02_text_dates_iso.xlsx")


# ---------------------------------------------------------------------------
# q03: first header a real date, rest EOMONTH formulas
# ---------------------------------------------------------------------------
def q03():
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 1, "Customer")
    first = ws.cell(1, 2, dt.date(2021, 1, 31))
    first.number_format = "yyyy-mm-dd"
    for j in range(1, N_MONTHS):
        col = 2 + j
        prev_col_letter = get_column_letter(col - 1)
        cell = ws.cell(1, col, f"=EOMONTH({prev_col_letter}1,1)")
        cell.number_format = "yyyy-mm-dd"
    for i, c in enumerate(CUSTOMERS):
        ws.cell(2 + i, 1, c)
        for j, v in enumerate(M[c]):
            ws.cell(2 + i, 2 + j, v)
    wb.save(f"{OUT}/q03_formula_dates_eomonth.xlsx")


# ---------------------------------------------------------------------------
# q04: real dates, REVERSE chronological order (newest in col B)
# ---------------------------------------------------------------------------
def q04():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 1, "Customer")
    rev_idx = list(range(N_MONTHS - 1, -1, -1))   # 35..0
    for j, idx in enumerate(rev_idx):
        c = ws.cell(1, 2 + j, MONTH_ENDS[idx]); c.number_format = "yyyy-mm-dd"
    for i, cust in enumerate(CUSTOMERS):
        ws.cell(2 + i, 1, cust)
        for j, idx in enumerate(rev_idx):
            ws.cell(2 + i, 2 + j, M[cust][idx])
    wb.save(f"{OUT}/q04_reversed_dates.xlsx")


# ---------------------------------------------------------------------------
# q05: real dates with several months MISSING (skip Apr-21 idx3 and Sep-21 idx8)
# ---------------------------------------------------------------------------
def q05():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    skip = {3, 8}   # Apr-2021, Sep-2021
    keep = [i for i in range(N_MONTHS) if i not in skip]
    ws.cell(1, 1, "Customer")
    for j, idx in enumerate(keep):
        c = ws.cell(1, 2 + j, MONTH_ENDS[idx]); c.number_format = "yyyy-mm-dd"
    for i, cust in enumerate(CUSTOMERS):
        ws.cell(2 + i, 1, cust)
        for j, idx in enumerate(keep):
            ws.cell(2 + i, 2 + j, M[cust][idx])
    wb.save(f"{OUT}/q05_gaps_missing_months.xlsx")


# ---------------------------------------------------------------------------
# q06: A=Type, B=Region junk; customer in col C; months from col D
# ---------------------------------------------------------------------------
def q06():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 1, "Type"); ws.cell(1, 2, "Region"); ws.cell(1, 3, "Customer")
    for j, d in enumerate(MONTH_ENDS):
        c = ws.cell(1, 4 + j, d); c.number_format = "yyyy-mm-dd"
    regions = ["US", "EMEA", "APAC"]
    for i, cust in enumerate(CUSTOMERS):
        ws.cell(2 + i, 1, "Recurring")
        ws.cell(2 + i, 2, regions[i % 3])
        ws.cell(2 + i, 3, cust)
        for j, v in enumerate(M[cust]):
            ws.cell(2 + i, 4 + j, v)
    wb.save(f"{OUT}/q06_customer_not_col_a.xlsx")


# ---------------------------------------------------------------------------
# q07: rows 1-4 title/blurb/blank; header row 5; customers from row 6
# ---------------------------------------------------------------------------
def q07():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 1, "Confidential - Customer MRR Detail")
    ws.cell(2, 1, "Prepared for internal diligence use only")
    ws.cell(3, 1, "Figures in USD, monthly recurring revenue")
    # row 4 intentionally blank
    ws.cell(5, 1, "Customer")
    for j, d in enumerate(MONTH_ENDS):
        c = ws.cell(5, 2 + j, d); c.number_format = "yyyy-mm-dd"
    for i, cust in enumerate(CUSTOMERS):
        ws.cell(6 + i, 1, cust)
        for j, v in enumerate(M[cust]):
            ws.cell(6 + i, 2 + j, v)
    wb.save(f"{OUT}/q07_data_below_title_rows.xlsx")


# ---------------------------------------------------------------------------
# q08: multi-row per customer (Product in col B); customer A; months from col C
#      per-customer value = SUM of product rows
# ---------------------------------------------------------------------------
def q08():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 1, "Customer"); ws.cell(1, 2, "Product")
    for j, d in enumerate(MONTH_ENDS):
        c = ws.cell(1, 3 + j, d); c.number_format = "yyyy-mm-dd"
    r = 2
    for cust in CUSTOMERS:
        core = M[cust]
        # Core row = 70% of total, Add-on = 30% of total (split the shared numbers)
        core_vals = [round(v * 0.7) for v in core]
        addon_vals = [v - cv for v, cv in zip(core, core_vals)]
        ws.cell(r, 1, cust); ws.cell(r, 2, "Core")
        for j, v in enumerate(core_vals):
            ws.cell(r, 3 + j, v)
        r += 1
        ws.cell(r, 1, cust); ws.cell(r, 2, "Add-on")
        for j, v in enumerate(addon_vals):
            ws.cell(r, 3 + j, v)
        r += 1
    wb.save(f"{OUT}/q08_multi_row_per_customer.xlsx")


# ---------------------------------------------------------------------------
# q09: real-date headers, messy text/blank/NA/parens value cells
# ---------------------------------------------------------------------------
def q09():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 1, "Customer")
    for j, d in enumerate(MONTH_ENDS):
        c = ws.cell(1, 2 + j, d); c.number_format = "yyyy-mm-dd"

    def messify(v, i, j):
        # Deterministic mess so the file is reproducible
        if v == 0:
            # blanks and NA for pre-start zeros
            return "" if (i + j) % 2 == 0 else "NA"
        bucket = (i * 7 + j) % 11
        if bucket == 0:
            return f"{v:,}"            # "9,900"  thousands separator text
        if bucket == 1:
            return f"${v:,}"           # "$1,542" currency text
        if bucket == 2:
            return f"({min(v, 500)})"  # accounting negative, e.g. "(500)"
        if bucket == 3:
            return ""                  # stray blank
        return v                       # plain number

    for i, cust in enumerate(CUSTOMERS):
        ws.cell(2 + i, 1, cust)
        for j, v in enumerate(M[cust]):
            ws.cell(2 + i, 2 + j, messify(v, i, j))
    wb.save(f"{OUT}/q09_text_numbers_and_parens.xlsx")


# ---------------------------------------------------------------------------
# q10: LONG / tidy format: Customer, Date, MRR (24 months, all 10 customers)
# ---------------------------------------------------------------------------
def q10():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.cell(1, 1, "Customer"); ws.cell(1, 2, "Date"); ws.cell(1, 3, "MRR")
    r = 2
    for cust in CUSTOMERS:
        for j in range(24):                      # ~24 months Jan-2021..Dec-2022
            dcell = ws.cell(r, 2, MONTH_ENDS[j]); dcell.number_format = "yyyy-mm-dd"
            ws.cell(r, 1, cust)
            ws.cell(r, 3, M[cust][j])
            r += 1
    wb.save(f"{OUT}/q10_long_tidy.xlsx")


# ---------------------------------------------------------------------------
# q11: quarterly native. Real quarter-end dates as headers.
# ---------------------------------------------------------------------------
def q11():
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    # quarter-end indices within MONTHS: Mar(2),Jun(5),Sep(8),Dec(11)... every 3rd
    q_idx = [i for i in range(N_MONTHS) if (MONTHS[i].month in (3, 6, 9, 12))]
    ws.cell(1, 1, "Customer")
    for j, idx in enumerate(q_idx):
        c = ws.cell(1, 2 + j, MONTH_ENDS[idx]); c.number_format = "yyyy-mm-dd"
    for i, cust in enumerate(CUSTOMERS):
        ws.cell(2 + i, 1, cust)
        for j, idx in enumerate(q_idx):
            # quarterly recurring value = the MRR in that quarter-end month
            ws.cell(2 + i, 2 + j, M[cust][idx])
    wb.save(f"{OUT}/q11_quarterly_native.xlsx")


for fn in (q01, q02, q03, q04, q05, q06, q07, q08, q09, q10, q11):
    fn()
print("All 11 fixtures written to", OUT)
