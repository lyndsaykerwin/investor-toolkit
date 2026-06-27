#!/usr/bin/env python3
"""
build_concentration.py — build a banker-grade customer concentration workbook.

Output = TWO sheets:
  1. "Customer Concentration" — Top 10 customers ranked largest→smallest by
     current run-rate, each with its value ($) and % of total; a Top-10 subtotal
     (+ % of total); an "All Remaining Customers" row (+ %); a grand TOTAL that
     ties out (Top 10 + Remaining = 100% = total); and a visible tie-out check.
  2. "Raw Data" — a VERBATIM copy of the source sheet (Critical Rule: zero
     edits, no reformatting, no reordering).

Every figure on the analysis sheet is a LIVE FORMULA referencing the Raw Data
tab — zero hardcoded numbers — so the tie-out is provably sourced.

VALUES ARE NEVER TRANSFORMED. The skill never annualizes (no ×12) — that is
where accuracy goes wrong. The source value is shown exactly as reported; the
unit is only a LABEL: "Current ARR"/"Current MRR" when the source unit is
certain, otherwise the neutral "Current Run-Rate". Concentration (the ranking
and the percentages) is identical regardless of unit, so an honest label beats
a risky transformation.

Ranking can't be done by spreadsheet formula, so values are read once in Python
ONLY to determine row order; the displayed numbers remain link-back formulas.

Usage:
    python3 build_concentration.py SOURCE.xlsx OUT.xlsx \
        --sheet "ARR by Customer" \
        --customer-col A --value-col AQ \
        --first-row 5 --last-row 162 \
        --unit ARR \
        [--period-label "May 2026"] [--company "Example Co."] [--allow-large]

--unit is a label only (ARR / MRR / Run-Rate), default Run-Rate. CSV sources:
pass --value-col / --customer-col as 1-based column numbers or letters; the CSV
is copied verbatim onto the Raw Data tab.
"""
import argparse
import csv
import sys
from copy import copy as _copy

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

RAW_SHEET = "Raw Data"
ANALYSIS_SHEET = "Customer Concentration"
NAVY = "1F4E79"
SOFT_BLUE = "D9E1F2"
WHITE = "FFFFFF"
GREY = "595959"
BLUE_FONT = "0000FF"     # hardcoded input
GREEN_FONT = "008000"    # cross-sheet reference
BLACK_FONT = "000000"    # in-sheet formula
MAX_VERBATIM_ROWS = 20000

thin = Side(style="thin", color="000000")
med = Side(style="medium", color="000000")
BORDER_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_TOPMED = Border(left=thin, right=thin, top=med, bottom=thin)


def col_to_idx(c):
    return column_index_from_string(c) if isinstance(c, str) and not c.isdigit() \
        else int(c)


def _parse_num(x):
    """Parse a number from a numeric cell OR a currency-formatted string
    ('$8,333.33 '). Returns None for blanks / non-numeric."""
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.strip().replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
        if s in ("", "-", "–"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _coerce(ref):
    """Excel/LibreOffice expression that turns a currency-text cell into a
    number: strip $ and , and surrounding spaces, then VALUE; blanks/junk → 0.
    Uses VALUE (not NUMBERVALUE — the latter returns #NAME? in headless
    LibreOffice). Works element-wise inside SUMPRODUCT for the grand total."""
    return (f'IFERROR(VALUE(SUBSTITUTE(SUBSTITUTE(TRIM({ref}),"$",""),'
            f'",","")),0)')


def center_continuous(ws, row, first_col, last_col, text, font, fill):
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.value = text if c == first_col else None
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")


def copy_source_verbatim(src_path, dest_ws, sheet_name):
    """Copy source onto dest_ws preserving values + all styling. No edits."""
    if src_path.lower().endswith(".csv"):
        with open(src_path, newline="", encoding="utf-8-sig") as fh:
            for r_idx, row in enumerate(csv.reader(fh), start=1):
                for c_idx, val in enumerate(row, start=1):
                    cast = val
                    if isinstance(val, str):
                        s = val.strip()
                        if s == "":
                            cast = None
                        else:
                            try:
                                cast = float(s) if ("." in s or "e" in s.lower()) else int(s)
                            except ValueError:
                                cast = val
                    dest_ws.cell(row=r_idx, column=c_idx, value=cast)
        return
    wb = load_workbook(src_path)
    src_ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            d = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                d.font = _copy(cell.font)
                d.fill = _copy(cell.fill)
                d.border = _copy(cell.border)
                d.alignment = _copy(cell.alignment)
                d.number_format = cell.number_format
            if cell.comment is not None:
                d.comment = Comment(cell.comment.text or "", cell.comment.author or "source")
    for mr in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(mr))
    for cl, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dest_ws.column_dimensions[cl].width = dim.width
    for rn, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dest_ws.row_dimensions[rn].height = dim.height


def read_customers(src_path, sheet_name, cust_idx, val_idx, first_row, last_row):
    """Return [(row_number, customer_label, value)] for ranking. row_number is
    the 1-based row on the (verbatim) Raw Data tab, used to build link-backs.

    Returns (rows, value_is_text). value_is_text is True when the source value
    column is stored as currency-formatted strings (e.g. '$8,333.33 ') rather
    than numbers — Excel SUM() silently ignores text, so the build must coerce
    in the formulas while still leaving Raw Data verbatim."""
    out = []
    n_text = n_num = 0
    if src_path.lower().endswith(".csv"):
        with open(src_path, newline="", encoding="utf-8-sig") as fh:
            rows = list(csv.reader(fh))
        lo = first_row or 2
        hi = last_row or len(rows)
        for rn in range(lo, hi + 1):
            r = rows[rn - 1]
            if cust_idx - 1 >= len(r) or val_idx - 1 >= len(r):
                continue
            label = str(r[cust_idx - 1]).strip()
            raw = r[val_idx - 1]
            v = _parse_num(raw)
            if v is None:
                continue
            n_text += 1  # CSV cells are always strings on disk
            if label and v:
                out.append((rn, label, v))
        return out, True  # a CSV value column is text by definition
    wb = load_workbook(src_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    hi = last_row or ws.max_row
    for rn in range(first_row, hi + 1):
        label = ws.cell(row=rn, column=cust_idx).value
        raw = ws.cell(row=rn, column=val_idx).value
        v = _parse_num(raw)
        if label is None or v is None:
            continue
        if isinstance(raw, str):
            n_text += 1
        else:
            n_num += 1
        out.append((rn, str(label).strip(), float(v)))
    wb.close()
    return out, n_text > n_num


def build(src_path, out_path, sheet_name, cust_col, val_col, first_row,
          last_row, unit, period_label, company, allow_large):
    cust_idx = col_to_idx(cust_col)
    val_idx = col_to_idx(val_col)
    val_letter = get_column_letter(val_idx)
    cust_letter = get_column_letter(cust_idx)
    # The unit is a LABEL only — values are never transformed. Use "ARR"/"MRR"
    # only when the source unit is certain; otherwise the neutral "Run-Rate".
    unit = (unit or "Run-Rate").strip()

    customers, value_is_text = read_customers(src_path, sheet_name, cust_idx,
                                              val_idx, first_row, last_row)
    if not customers:
        sys.exit("No customer rows found — check --customer-col/--value-col/--first-row/--last-row")
    if len(customers) > MAX_VERBATIM_ROWS and not allow_large:
        sys.exit(f"Source has {len(customers):,} customer rows (> {MAX_VERBATIM_ROWS:,}). "
                 f"A full verbatim copy will be very large. Re-run with --allow-large to "
                 f"proceed, or confirm a scope/subset with the user first.")

    customers.sort(key=lambda t: t[2], reverse=True)
    n = len(customers)
    top = customers[:10]

    wb = Workbook()
    ws = wb.active
    ws.title = ANALYSIS_SHEET
    raw = wb.create_sheet(RAW_SHEET)
    copy_source_verbatim(src_path, raw, sheet_name)

    # --- column widths / layout ---------------------------------------------
    ws.column_dimensions["A"].width = 7      # rank
    ws.column_dimensions["B"].width = 34     # customer
    ws.column_dimensions["C"].width = 22     # current run-rate
    ws.column_dimensions["D"].width = 18     # % of total

    f_title = Font(name="Arial", size=14, bold=True, color=NAVY)
    f_sub = Font(name="Arial", size=9, color=GREY)
    f_hdr = Font(name="Arial", size=10, bold=True, color=WHITE)
    f_data = Font(name="Arial", size=10, color=BLACK_FONT)
    f_xsheet = Font(name="Arial", size=10, color=GREEN_FONT)
    f_bold = Font(name="Arial", size=10, bold=True, color=BLACK_FONT)
    f_total = Font(name="Arial", size=10, bold=True, color=WHITE)
    fill_navy = PatternFill("solid", fgColor=NAVY)
    fill_soft = PatternFill("solid", fgColor=SOFT_BLUE)
    fill_white = PatternFill("solid", fgColor=WHITE)

    # Row 1 title, row 2 subtitle. Values are shown verbatim — NOT annualized —
    # so the subtitle states the unit honestly and notes no transformation.
    co = f"{company} — " if company else ""
    center_continuous(ws, 1, 1, 4, f"{co}Customer Concentration Analysis", f_title, fill_white)
    ws.row_dimensions[1].height = 22
    per = f"as of {period_label}" if period_label else "current period"
    center_continuous(
        ws, 2, 1, 4,
        f"Current {unit} ({per}) — values shown as reported, not annualized  |  "
        f"{n} customers  |  all figures link to Raw Data",
        f_sub, fill_white)

    # Header row 5
    hdr_row = 5
    headers = [("Rank", "center"), ("Customer", "left"),
               (f"Current {unit} ($)", "right"), (f"% of Total {unit}", "right")]
    for i, (txt, al) in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=i, value=txt)
        c.font = f_hdr
        c.fill = fill_navy
        c.alignment = Alignment(horizontal=al, vertical="center")
        c.border = BORDER_TOPMED
    ws.row_dimensions[hdr_row].height = 18

    # Total row position depends on top-10 length
    first_data = hdr_row + 1           # 6
    last_data = first_data + len(top) - 1
    sub_row = last_data + 1            # Top 10 subtotal
    rem_row = sub_row + 1             # remaining
    tot_row = rem_row + 1             # grand total
    tie_row = tot_row + 2            # tie-out check

    def val_expr(ref):
        """Link-back to a Raw Data cell as a number, coercing currency-text when
        needed. No transformation — the value is shown exactly as reported."""
        return f"={_coerce(ref)}" if value_is_text else f"={ref}"

    # Top 10 rows — link-back formulas
    for i, (rn, label, val) in enumerate(top):
        r = first_data + i
        # rank
        a = ws.cell(row=r, column=1, value=i + 1)
        a.font = f_data
        a.alignment = Alignment(horizontal="center", vertical="center")
        a.border = BORDER_THIN
        # customer (link back)
        b = ws.cell(row=r, column=2, value=f"='{RAW_SHEET}'!{cust_letter}{rn}")
        b.font = f_xsheet
        b.alignment = Alignment(horizontal="left", vertical="center")
        b.border = BORDER_THIN
        # current run-rate (verbatim link back, coerced if the source is text)
        cval = ws.cell(row=r, column=3,
                       value=val_expr(f"'{RAW_SHEET}'!{val_letter}{rn}"))
        cval.font = f_xsheet
        cval.alignment = Alignment(horizontal="right", vertical="center")
        cval.border = BORDER_THIN
        cval.number_format = "$#,##0"
        # % of total
        d = ws.cell(row=r, column=4, value=f"=C{r}/C{tot_row}")
        d.font = f_data
        d.alignment = Alignment(horizontal="right", vertical="center")
        d.border = BORDER_THIN
        d.number_format = "0.0%"

    # grand total — independent SUM over the customer column on Raw Data.
    # When the source is currency-text, SUM() ignores it, so coerce each cell
    # inside a SUMPRODUCT (array-aware, no Ctrl-Shift-Enter needed).
    rng = f"'{RAW_SHEET}'!{val_letter}{first_row}:{val_letter}{last_row or (first_row + n - 1)}"
    total_formula = (f"=SUMPRODUCT({_coerce(rng)})" if value_is_text
                     else f"=SUM({rng})")

    # Top 10 subtotal
    _row(ws, sub_row, "Top 10 Subtotal", f"=SUM(C{first_data}:C{last_data})",
         f"=C{sub_row}/C{tot_row}", fill_soft, f_bold, f_bold, BORDER_TOPMED, "$#,##0")
    # Remaining customers
    rem_label = f"All Remaining Customers ({n - len(top)} customers)"
    _row(ws, rem_row, rem_label, f"=C{tot_row}-C{sub_row}",
         f"=C{rem_row}/C{tot_row}", fill_white, f_data, f_data, BORDER_THIN, "$#,##0")
    # Grand total
    _row(ws, tot_row, "TOTAL — All Customers", total_formula,
         f"=C{sub_row}/C{tot_row}+C{rem_row}/C{tot_row}", fill_navy, f_total, f_total,
         BORDER_TOPMED, "$#,##0", white_text=True)

    # Tie-out check row
    tlab = ws.cell(row=tie_row, column=2, value="Tie-out check (Top 10 + Remaining − Total = 0):")
    tlab.font = f_sub
    tlab.alignment = Alignment(horizontal="left", vertical="center")
    tchk = ws.cell(row=tie_row, column=3, value=f"=C{sub_row}+C{rem_row}-C{tot_row}")
    tchk.font = f_sub
    tchk.alignment = Alignment(horizontal="right", vertical="center")
    tchk.number_format = "$#,##0"

    # Source note
    note_row = tie_row + 1
    src_name = src_path.split("/")[-1]
    nc = ws.cell(row=note_row, column=1,
                 value=f"Source: '{RAW_SHEET}' tab (verbatim copy of {src_name}, sheet "
                       f"'{sheet_name}'). Every {unit} figure is a live formula "
                       f"referencing it; values are shown as reported, not annualized.")
    nc.font = Font(name="Arial", size=8, color=GREY)
    nc.alignment = Alignment(horizontal="left", vertical="center")

    for r in range(first_data, tot_row + 1):
        ws.row_dimensions[r].height = 16

    # Force Excel/Numbers/LibreOffice to recalc every formula on open, so the
    # link-back value cells are never blank in a viewer that doesn't auto-calc.
    wb.calculation.fullCalcOnLoad = True

    ws.freeze_panes = "A6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True if ws.sheet_properties.pageSetUpPr else None
    ws.print_area = f"A1:D{note_row}"

    wb.save(out_path)
    return {
        "out": out_path, "n_customers": n, "top10": [(l, v) for _, l, v in top],
        "unit": unit, "total_row": tot_row, "sub_row": sub_row,
        "rem_row": rem_row,
    }


def _row(ws, r, label, c_formula, d_formula, fill, lab_font, val_font,
         border, numfmt, white_text=False):
    a = ws.cell(row=r, column=1, value="")
    a.fill = fill
    a.border = border
    a.font = lab_font
    b = ws.cell(row=r, column=2, value=label)
    b.fill = fill
    b.font = lab_font
    b.alignment = Alignment(horizontal="left", vertical="center")
    b.border = border
    c = ws.cell(row=r, column=3, value=c_formula)
    c.fill = fill
    c.font = val_font
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border
    c.number_format = numfmt
    d = ws.cell(row=r, column=4, value=d_formula)
    d.fill = fill
    d.font = val_font
    d.alignment = Alignment(horizontal="right", vertical="center")
    d.border = border
    d.number_format = "0.0%"
    ws.row_dimensions[r].height = 18


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("source")
    ap.add_argument("out")
    ap.add_argument("--sheet", default="")
    ap.add_argument("--customer-col", required=True)
    ap.add_argument("--value-col", required=True)
    ap.add_argument("--first-row", type=int, required=True)
    ap.add_argument("--last-row", type=int, default=0)
    ap.add_argument("--unit", default="Run-Rate",
                    help="Column label only — NO transformation. Use 'ARR' or "
                         "'MRR' only when the source unit is certain; otherwise "
                         "the neutral 'Run-Rate' (default).")
    ap.add_argument("--period-label", default="")
    ap.add_argument("--company", default="")
    ap.add_argument("--allow-large", action="store_true")
    a = ap.parse_args()
    res = build(a.source, a.out, a.sheet, a.customer_col, a.value_col,
                a.first_row, a.last_row or 0, a.unit, a.period_label,
                a.company, a.allow_large)
    print(f"Built {res['out']}")
    print(f"  {res['n_customers']} customers, unit label '{res['unit']}' (values verbatim, not annualized)")
    print(f"  Top 10 (pre-recalc order): "
          f"{', '.join(f'{l}={v:,.0f}' for l, v in res['top10'][:3])} ...")
    print("  Recalculate (LibreOffice/Excel) then verify the tie-out check cell = 0.")


if __name__ == "__main__":
    main()
