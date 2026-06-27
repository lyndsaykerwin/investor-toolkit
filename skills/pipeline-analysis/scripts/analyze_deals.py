#!/usr/bin/env python3
"""
Pipeline Analysis engine.

Reads a deal-level export (.xlsx / .csv), maps headers to canonical fields,
classifies each deal's outcome, assigns a tier, and writes a formula-driven
Excel workbook (Raw / Control Panel / Analysis / Findings). Every Analysis
number is a SUMIFS/COUNTIFS formula referencing the Raw tab, so it traces back.

Also prints a JSON summary to stdout for the calling agent to report.

Usage:
  python3 analyze_deals.py --input deals.xlsx --output out.xlsx \
      [--target 2000000] [--sheet "Pipeline"] [--header-row N] [--map map.json]

No real client data belongs in this skill folder. Write outputs elsewhere.
"""
import argparse, json, re, sys, statistics, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

# ---------- canonical field synonyms ----------
SYNONYMS = {
    "deal_id":     ["deal id","opportunity id","opp id","record id","id","deal","opportunity name","opportunity","opp name"],
    "account":     ["account","customer","company","account name","logo","client","customer name","company name"],
    "arr":         ["arr","booking arr","acv","annual contract value","deal value","amount","$arr","value","tcv","mrr","new arr","total arr","contract value"],
    "outcome":     ["stage","deal stage","status","outcome","stage name","forecast category","opportunity stage","sales stage","pipeline stage","deal status"],
    "close_date":  ["close date","closed date","close period","won date","expected close","expected close date","close","date closed","est close date","forecast close"],
    "create_date": ["create date","created","created date","open date","start date","created on","date created"],
    "deal_type":   ["deal type","type","opportunity type","business type","deal category","opp type"],
    "deal_source": ["deal source","lead source","source","channel","origin","lead origin"],
    "product":     ["product","product family","product line","sku","module","offering","product name"],
    "owner":       ["owner","deal owner","opportunity owner","rep","ae","sales rep","account executive","opportunity owner name","sales owner"],
    "probability": ["probability","% probability","win %","likelihood","confidence","prob","probability %","win probability","stage probability"],
    "age":         ["age","age (days)","days open","deal age","days in pipeline","aging"],
}

def norm(s):
    return re.sub(r"[^a-z0-9]+"," ",str(s).strip().lower()).strip()

TOTALS_TOKENS={"total","totals","sum","subtotal","sub total","grand total",
    "count","average","avg","mean","totals row","running total"}
def _is_totals_row(rec):
    """Drop subtotal/total/count rows common at the bottom of CRM exports."""
    for k in ("deal_id","account","outcome","product","deal_type"):
        v=rec.get(k)
        if v is not None and norm(v) in TOTALS_TOKENS:
            return True
    return False

def match_header(h):
    n = norm(h)
    if not n: return None
    for canon, syns in SYNONYMS.items():
        if n in [norm(x) for x in syns]:
            return canon
    # loose match: synonym is a WHOLE WORD in the header, or (multi-word / >=5 char)
    # a substring. Whole-word matching lets short tokens like "arr"/"id"/"age"
    # match "total contract arr" while preventing junk like "c" matching "acv".
    tokens=set(n.split())
    best=None; best_len=0
    for canon, syns in SYNONYMS.items():
        for x in syns:
            xn=norm(x)
            if not xn: continue
            hit = (xn in tokens) or ((" " in xn or len(xn)>=5) and xn in n)
            if hit and len(xn)>best_len:
                best=canon; best_len=len(xn)
    return best

# ---------- input loading ----------
def col_to_idx(spec):
    """'B' or 'b' -> 1; '7' -> 6 (1-based number)."""
    spec=str(spec).strip()
    if spec.isdigit(): return int(spec)-1
    n=0
    for ch in spec.upper():
        if "A"<=ch<="Z": n=n*26+(ord(ch)-64)
    return n-1

def load_mapped(path, sheet, colmap):
    """Headerless / explicitly-mapped files: colmap = {canonical: 0-based col idx}.
    Data rows = those with a numeric value in the arr column (skips title/filter rows)."""
    path=Path(path)
    if path.suffix.lower()==".csv":
        import csv
        with open(path, newline="", encoding="utf-8-sig", errors="ignore") as f:
            rows=[r for r in csv.reader(f)]
        sheetname=None; sheetnames=None
    else:
        wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws=wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows=[[c for c in r] for r in ws.iter_rows(values_only=True)]
        sheetname=ws.title; sheetnames=wb.sheetnames
    arr_idx=colmap.get("arr")
    records=[]
    for r in rows:
        if arr_idx is not None and (arr_idx>=len(r) or to_num(r[arr_idx]) is None):
            continue
        rec={c:(r[i] if i<len(r) else None) for c,i in colmap.items()}
        if _is_totals_row(rec): continue
        records.append(rec)
    return records, dict(colmap), {"sheet":sheetname,"sheetnames":sheetnames,
                                   "title_block":[],"mapped_via":"--map (positional)"}

def load_rows(path, sheet=None, header_row=None):
    path=Path(path)
    if path.suffix.lower()==".csv":
        import csv
        with open(path, newline="", encoding="utf-8-sig", errors="ignore") as f:
            rows=[r for r in csv.reader(f)]
        return _matrix_to_records(rows, header_row)
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    targets=[sheet] if sheet else list(wb.sheetnames)
    all_records=[]; union={}; used=[]; first_meta=None
    for sn in targets:
        ws=wb[sn]
        rows=[[c for c in r] for r in ws.iter_rows(values_only=True)]
        recs,mp,meta=_matrix_to_records(rows, header_row, sheetname=ws.title, sheetnames=wb.sheetnames)
        # include a sheet only if it actually looks like deal data
        looks_like_deals = ("arr" in mp) and bool({"outcome","probability","close_date"} & set(mp))
        if sheet or looks_like_deals:
            all_records+=recs
            for k,v in mp.items(): union.setdefault(k,v)
            used.append(ws.title)
            if first_meta is None: first_meta=meta
    if first_meta is None:  # nothing qualified — fall back to first sheet so caller can refuse cleanly
        ws=wb[wb.sheetnames[0]]
        rows=[[c for c in r] for r in ws.iter_rows(values_only=True)]
        return _matrix_to_records(rows, header_row, sheetname=ws.title, sheetnames=wb.sheetnames)
    first_meta["sheets_used"]=used
    return all_records, union, first_meta

def _detect_header(rows):
    best_i, best_score = 0, -1
    for i, r in enumerate(rows[:25]):
        score=sum(1 for c in r if match_header(c))
        if score>best_score:
            best_score, best_i = score, i
    return best_i, best_score

def _matrix_to_records(rows, header_row, sheetname=None, sheetnames=None):
    rows=[r for r in rows if any(c is not None and str(c).strip()!="" for c in r)]
    if not rows: return [], {}, {"sheet":sheetname,"sheetnames":sheetnames,"header_score":0}
    hi = header_row if header_row is not None else _detect_header(rows)[0]
    headers=rows[hi]
    mapping={}
    for ci,h in enumerate(headers):
        canon=match_header(h)
        if canon and canon not in mapping:  # first wins
            mapping[canon]=ci
    records=[]
    dropped_totals=0
    for r in rows[hi+1:]:
        rec={canon: (r[ci] if ci < len(r) else None) for canon,ci in mapping.items()}
        if not any(v is not None and str(v).strip()!="" for v in rec.values()):
            continue
        if _is_totals_row(rec):
            dropped_totals+=1; continue
        records.append(rec)
    meta_extra_dropped=dropped_totals
    title_block=[]
    for r in rows[:hi]:
        for c in r:
            if c is not None and str(c).strip()!="":
                title_block.append(str(c).strip())
    meta={"sheet":sheetname,"sheetnames":sheetnames,
          "header_row_index":hi,"raw_headers":[str(h) for h in headers],
          "dropped_totals_rows":dropped_totals,"title_block":title_block}
    return records, mapping, meta

# ---------- value parsing ----------
def to_num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    s=re.sub(r"[^0-9.\-]","",str(v))
    try: return float(s) if s not in ("","-",".") else None
    except: return None

def to_date(v):
    if v is None or str(v).strip()=="": return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return datetime.date(v.year,v.month,v.day)
    s=str(v).strip()
    if re.fullmatch(r"\d{4,6}(\.0)?", s):  # excel serial
        try: return datetime.date(1899,12,30)+datetime.timedelta(days=int(float(s)))
        except: pass
    for fmt in ("%Y-%m-%d","%m/%d/%Y","%m/%d/%y","%d/%m/%Y","%Y/%m/%d","%b %Y","%b-%y","%m-%d-%Y","%d-%b-%Y","%Y-%m-%d %H:%M:%S"):
        try: return datetime.datetime.strptime(s,fmt).date()
        except: pass
    return None

def to_prob(v):
    n=to_num(v)
    if n is None: return None
    if n>1.5: n=n/100.0   # "40" -> 0.40
    return max(0.0,min(1.0,n))

def classify_outcome(raw, prob, close_date):
    t=norm(raw)
    if any(k in t for k in ["closed won","won","win"]): return "won"
    if any(k in t for k in ["closed lost","lost","loss"]): return "lost"
    if any(k in t for k in ["no business awarded","no decision","no opportunity","abandoned","disqualified","dead","nurture"]): return "no_decision"
    if t in ("","none"):
        if prob is not None and 0 < prob < 1: return "open"
        if close_date is None: return "open"
    return "open"  # any live stage name

def deal_type_group(dt):
    t=norm(dt)
    if "new" in t: return "New Logo"
    if t: return "Expansion"
    return ""

# ---------- main analysis ----------
def analyze(records, mapping, target=None):
    cols=set(mapping)
    deals=[]
    for r in records:
        arr=to_num(r.get("arr"))
        cd=to_date(r.get("close_date"))
        crd=to_date(r.get("create_date"))
        prob=to_prob(r.get("probability"))
        bucket=classify_outcome(r.get("outcome"), prob, cd)
        cycle=(cd-crd).days if (cd and crd and (cd-crd).days>=0) else None
        deals.append({
            "deal_id": r.get("deal_id"), "account": r.get("account"),
            "arr": arr or 0.0, "outcome_raw": r.get("outcome"), "bucket": bucket,
            "close_date": cd, "create_date": crd, "cycle": cycle,
            "deal_type": r.get("deal_type"), "deal_type_group": deal_type_group(r.get("deal_type")),
            "deal_source": r.get("deal_source"), "product": r.get("product"),
            "owner": r.get("owner"), "probability": prob,
            "age": to_num(r.get("age")),
        })

    def s(bucket, key="arr"):
        return sum(d[key] for d in deals if d["bucket"]==bucket and isinstance(d[key],(int,float)))
    def c(bucket):
        return sum(1 for d in deals if d["bucket"]==bucket)

    W,L,N = s("won"), s("lost"), s("no_decision")
    wc,lc,nc,oc = c("won"),c("lost"),c("no_decision"),c("open")
    closed_arr=W+L+N
    has_closed = (wc+lc+nc)>0
    has_open = oc>0

    # tier
    tier1 = ("arr" in cols) and bool({"outcome","probability","close_date"} & cols)
    tier2 = tier1 and bool({"deal_type","product","owner","create_date","probability"} & cols)
    if tier2: tier="Silver"
    elif tier1: tier="Bronze"
    else: tier="Below threshold"

    m={"tier":tier,"deal_count":len(deals),"has_closed":has_closed,"has_open":has_open,
       "buckets":{"won":wc,"lost":lc,"no_decision":nc,"open":oc},
       "arr":{"won":W,"lost":L,"no_decision":N,"open":s("open")}}

    blocked={}
    # win rates
    if has_closed and (W+L)>0:
        m["competitive_win_rate_arr"]=W/(W+L)
        m["competitive_win_rate_count"]=wc/(wc+lc) if (wc+lc) else None
    else: blocked["competitive_win_rate"]="needs closed Won + Lost outcomes"
    if has_closed and closed_arr>0:
        m["allin_win_rate_arr"]=W/closed_arr
        m["no_decision_rate_arr"]=N/closed_arr
    else: blocked["allin_win_rate"]="needs closed outcomes (won/lost/no-decision)"
    m["won_lost_ratio_arr"]= (W/L) if L>0 else None

    # sales cycle
    cycles=[d["cycle"] for d in deals if d["bucket"]=="won" and d["cycle"] is not None]
    if cycles:
        m["sales_cycle_days_mean"]=round(statistics.mean(cycles),1)
        m["sales_cycle_days_median"]=statistics.median(cycles)
    else: blocked["sales_cycle"]="needs Create Date + Close Date on won deals"

    # new vs upsell (of won)
    if has_closed and "deal_type" in cols and W>0:
        new=sum(d["arr"] for d in deals if d["bucket"]=="won" and d["deal_type_group"]=="New Logo")
        m["pct_new_logo_arr"]=new/W
        m["pct_expansion_arr"]=1-(new/W)
    else: blocked["new_vs_upsell"]="needs Deal Type on won deals"

    # open module
    if has_open:
        openarr=s("open")
        m["open_pipeline_arr"]=openarr
        m["open_deal_count"]=oc
        probs=[d for d in deals if d["bucket"]=="open" and d["probability"] is not None]
        if probs:
            m["weighted_pipeline_arr"]=sum(d["arr"]*d["probability"] for d in probs)
        else: blocked["weighted_pipeline"]="needs % Probability on open deals"
        if target:
            m["pipeline_coverage_x"]=round(openarr/target,2) if target else None
            m["bookings_target"]=target
        else: blocked["pipeline_coverage"]="needs a bookings target (pass --target)"
    # ---- aging of open deals (uses Age column, else today − create_date) ----
    today=datetime.date.today()
    def age_of(d):
        if d.get("age") is not None: return d["age"]
        if d.get("create_date"): return (today-d["create_date"]).days
        return None
    aged=[(d,age_of(d)) for d in deals if d["bucket"]=="open"]
    aged=[(d,a) for d,a in aged if a is not None and a>=0]
    if aged:
        order=["0–30 days","31–60 days","61–90 days","91–180 days","180+ days"]
        bArr={k:0 for k in order}; bCnt={k:0 for k in order}
        def buck(a):
            return order[0] if a<=30 else order[1] if a<=60 else order[2] if a<=90 else order[3] if a<=180 else order[4]
        for d,a in aged:
            k=buck(a); bArr[k]+=d["arr"]; bCnt[k]+=1
        m["aging"]={"order":order,"arr":bArr,"count":bCnt,
                    "stale_over_90_arr":bArr["91–180 days"]+bArr["180+ days"],
                    "stale_over_90_count":bCnt["91–180 days"]+bCnt["180+ days"]}
    elif has_open:
        blocked["aging"]="needs Create Date or an Age column on open deals"

    # ---- bookings by period (closed Won, by year) ----
    won_dated=[d for d in deals if d["bucket"]=="won" and d.get("close_date")]
    if won_dated:
        by_year={}
        for d in won_dated: by_year[d["close_date"].year]=by_year.get(d["close_date"].year,0)+d["arr"]
        m["bookings_by_year"]={str(y):by_year[y] for y in sorted(by_year)}
    elif has_closed:
        blocked["bookings_by_period"]="needs Close Date on won deals"

    # ---- concentration: top deals as % of the relevant pool ----
    pool=[d for d in deals if d["bucket"]=="open"] if (has_open and not has_closed) else deals
    pool_total=sum(d["arr"] for d in pool) or 0
    topN=sorted(pool, key=lambda d:d["arr"], reverse=True)[:10]
    m["concentration"]=[{"account":str(d.get("account") or d.get("deal_id") or "(unnamed)"),
                         "arr":d["arr"],"pct":(d["arr"]/pool_total if pool_total else None)} for d in topN]
    m["top5_pct"]=(sum(d["arr"] for d in topN[:5])/pool_total) if pool_total else None
    m["pool_total"]=pool_total

    m["blocked"]=blocked
    return deals, m

# ---------- workbook build ----------
import shutil
from openpyxl.styles import Border, Side

NAVY="1F3864"; LTband="EAF0F7"; GREY="595959"
TITLE=Font(bold=True,size=14,color=NAVY)
SUB=Font(size=10,color=GREY)
HEADF=Font(bold=True,color="FFFFFF"); HEADFILL=PatternFill("solid",fgColor=NAVY)
SECT=Font(bold=True,size=11,color=NAVY)
TOTF=Font(bold=True,color=NAVY)
NOTE=Font(italic=True,size=9,color=GREY)
BLOCKF=Font(italic=True,color="B00000")
USD='$#,##0;($#,##0);"–"'; PCT='0.0%'; INTF='#,##0'
THIN=Side(style="thin",color="BFBFBF")
BOX=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
TOPLINE=Border(top=Side(style="thin",color=NAVY))

def _pivot(deals, rowkey, colkey):
    """Return (row_labels, col_labels, grid$, row_totals, col_totals, counts_by_row, grand)."""
    def lab(v): return str(v).strip() if v not in (None,"") else "Unspecified"
    rsum={}; csum={}; cell={}; rcount={}
    for d in deals:
        rk=lab(d.get(rowkey)); ck=lab(d.get(colkey)); a=d.get("arr") or 0
        rsum[rk]=rsum.get(rk,0)+a; csum[ck]=csum.get(ck,0)+a
        cell[(rk,ck)]=cell.get((rk,ck),0)+a; rcount[rk]=rcount.get(rk,0)+1
    rows=sorted(rsum,key=lambda k:rsum[k],reverse=True)
    cols=[c for c in sorted(csum,key=lambda k:csum[k],reverse=True) if csum[c]>0]  # drop empty/blank cols
    grand=sum(rsum.values())
    return rows,cols,cell,rsum,csum,rcount,grand

def _matrix(ws, start_row, title, deals, rowkey, colkey, row_hdr):
    rows,cols,cell,rsum,csum,rcount,grand=_pivot(deals,rowkey,colkey)
    r=start_row
    ws.cell(r,1,title).font=SECT; r+=1
    # header
    ws.cell(r,1,row_hdr); ws.cell(r,1).font=HEADF; ws.cell(r,1).fill=HEADFILL; ws.cell(r,1).border=BOX
    for j,c in enumerate(cols,2):
        x=ws.cell(r,j,c); x.font=HEADF; x.fill=HEADFILL; x.border=BOX; x.alignment=Alignment(horizontal="right")
    tcol=len(cols)+2; dcol=tcol+1
    for lbl,cc in [("Total",tcol),("# Deals",dcol)]:
        x=ws.cell(r,cc,lbl); x.font=HEADF; x.fill=HEADFILL; x.border=BOX; x.alignment=Alignment(horizontal="right")
    r+=1
    for i,rk in enumerate(rows):
        band=PatternFill("solid",fgColor=LTband) if i%2 else None
        a=ws.cell(r,1,rk); a.border=BOX
        if band: a.fill=band
        for j,c in enumerate(cols,2):
            v=cell.get((rk,c),0)
            x=ws.cell(r,j,v); x.number_format=USD; x.border=BOX
            if band: x.fill=band
        t=ws.cell(r,tcol,rsum[rk]); t.number_format=USD; t.font=TOTF; t.border=BOX
        dc=ws.cell(r,dcol,rcount[rk]); dc.number_format=INTF; dc.border=BOX
        if band: t.fill=band; dc.fill=band
        r+=1
    # total row
    a=ws.cell(r,1,"Total"); a.font=TOTF; a.border=BOX
    for j,c in enumerate(cols,2):
        x=ws.cell(r,j,csum[c]); x.number_format=USD; x.font=TOTF; x.border=BOX
    x=ws.cell(r,tcol,grand); x.number_format=USD; x.font=TOTF; x.border=BOX
    x=ws.cell(r,dcol,sum(rcount.values())); x.number_format=INTF; x.font=TOTF; x.border=BOX
    return r+2, (len(cols)+3)

def _table(ws, r, title, headers, rows_data, fmts, aligns=None):
    """Generic banker table. rows_data = list of value-tuples; fmts per column."""
    ws.cell(r,1,title).font=SECT; r+=1
    for j,h in enumerate(headers,1):
        x=ws.cell(r,j,h); x.font=HEADF; x.fill=HEADFILL; x.border=BOX
        x.alignment=Alignment(horizontal=("left" if j==1 else "right"))
    r+=1
    for i,row in enumerate(rows_data):
        band=PatternFill("solid",fgColor=LTband) if i%2 else None
        for j,(val,fmt) in enumerate(zip(row,fmts),1):
            x=ws.cell(r,j,val); x.border=BOX
            if fmt and isinstance(val,(int,float)): x.number_format=fmt
            if band: x.fill=band
        r+=1
    return r+1

def build_workbook(deals, m, out, src_meta, target, source_path):
    # 1) Preserve raw EXACTLY: copy the original file, then prepend an Analysis sheet.
    src=Path(source_path)
    if src.suffix.lower() in (".xlsx",".xlsm"):
        shutil.copyfile(src, out)
        wb=openpyxl.load_workbook(out)            # keeps original styles/colors/columns
        an=wb.create_sheet("Analysis", 0)         # index 0 => first tab
    else:
        wb=openpyxl.Workbook(); an=wb.active; an.title="Analysis"
        raw=wb.create_sheet("Raw data")
        cols=["deal_id","account","arr","outcome_raw","deal_type","product",
              "deal_source","owner","probability","close_date","create_date"]
        for j,c in enumerate(cols,1):
            cell=raw.cell(1,j,c); cell.font=HEADF; cell.fill=HEADFILL
        for i,d in enumerate(deals,2):
            for j,c in enumerate(cols,1):
                v=d.get(c)
                raw.cell(i,j,v.isoformat() if isinstance(v,datetime.date) else v)

    # ---- header block ----
    tb=src_meta.get("title_block") or []
    title=tb[0] if tb else "Pipeline Summary"
    asof=next((s for s in tb if "as of" in s.lower()), "")
    an["A1"]=title; an["A1"].font=TITLE
    sub=[]
    if asof: sub.append(asof)
    sub.append("Open pipeline only — no win/loss outcomes in source" if not m["has_closed"]
               else ("Closed + open" if m["has_open"] else "Closed deals only"))
    an["A2"]=("   •   ".join(sub)); an["A2"].font=SUB

    # ---- KPI strip ----
    r=4
    kpis=[]
    if m["has_open"]:
        kpis+=[("Open pipeline", m.get("open_pipeline_arr",0), USD),
               ("Open deals", m.get("open_deal_count",0), INTF)]
    if m["has_closed"]:
        kpis+=[("Won ARR", m["arr"]["won"], USD),
               ("Competitive win rate", m.get("competitive_win_rate_arr"), PCT),
               ("All-in win rate", m.get("allin_win_rate_arr"), PCT)]
    for j,(lbl,val,fmt) in enumerate(kpis):
        c0=1+j*2
        a=an.cell(r,c0,lbl); a.font=NOTE
        v=an.cell(r+1,c0,val if val is not None else "n/a")
        v.font=Font(bold=True,size=12,color=NAVY)
        if val is not None: v.number_format=fmt
    r+=3

    # ---- MAIN: cross-tab summary (the deliverable for thin/open files) ----
    use=[d for d in deals if d["bucket"]=="open"] if (m["has_open"] and not m["has_closed"]) else deals
    rowkey="outcome_raw"  # = Stage
    if any(d.get("product") for d in use):
        r,_=_matrix(an, r, "Pipeline by Stage × Product ($)", use, rowkey, "product", "Stage")
    if any(d.get("deal_type") for d in use):
        r,_=_matrix(an, r, "Pipeline by Stage × Type ($)", use, rowkey, "deal_type", "Stage")

    # ---- by-owner cross-tab (Owner rows × Stage cols) ----
    try:
        if any(d.get("owner") for d in use):
            r,_=_matrix(an, r, "Pipeline by Owner × Stage ($)", use, "owner", rowkey, "Owner")
    except Exception: pass

    # ---- concentration: top deals ----
    try:
        conc=m.get("concentration") or []
        if conc:
            rows=[(d["account"], d["arr"], d["pct"]) for d in conc]
            r=_table(an, r, "Top deals by value (concentration)",
                     ["Deal / Account","Value","% of pool"], rows, [None,USD,PCT])
            if m.get("top5_pct") is not None:
                an.cell(r-1,1,f"Top 5 = {m['top5_pct']*100:.0f}% of ${m.get('pool_total',0):,.0f} pool").font=NOTE; r+=1
    except Exception: pass

    # ---- aging of open deals ----
    try:
        ag=m.get("aging")
        if ag:
            rows=[(k, ag["arr"][k], ag["count"][k]) for k in ag["order"]]
            r=_table(an, r, "Open pipeline aging (days open)",
                     ["Age bucket","Value","# Deals"], rows, [None,USD,INTF])
            an.cell(r-1,1,f"Stale (>90 days): ${ag['stale_over_90_arr']:,.0f} across {ag['stale_over_90_count']} deals").font=BLOCKF; r+=1
    except Exception: pass

    # ---- bookings by period (closed Won) ----
    try:
        by=m.get("bookings_by_year")
        if by:
            rows=[(y, v) for y,v in by.items()]
            r=_table(an, r, "Bookings by year (Won ARR)", ["Year","Won ARR"], rows, [None,USD])
    except Exception: pass

    # ---- closed-deal metrics (only when outcomes exist) ----
    if m["has_closed"]:
        an.cell(r,1,"Conversion context (from historical closed deals)").font=SECT; r+=1
        def metric(lbl,val,fmt=None,note=None):
            nonlocal r
            an.cell(r,1,lbl); c=an.cell(r,3,val if val is not None else "n/a")
            if val is not None and fmt: c.number_format=fmt
            c.font=TOTF
            if note: an.cell(r,4,note).font=NOTE
            r+=1
        metric("Competitive win rate (Won ÷ Won+Lost)", m.get("competitive_win_rate_arr"), PCT)
        metric("All-in win rate (Won ÷ Won+Lost+No-Decision)", m.get("allin_win_rate_arr"), PCT, "show both — never just the top one")
        metric("No-decision rate", m.get("no_decision_rate_arr"), PCT)
        metric("Avg sales cycle (days)", m.get("sales_cycle_days_mean"), INTF)
        metric("% New Logo of Won ARR", m.get("pct_new_logo_arr"), PCT)
        r+=1

    # ---- caveats + follow-ups (compact, at bottom) ----
    an.cell(r,1,"What this data can't tell you").font=SECT; r+=1
    NICE={"weighted_pipeline":"No probability per deal → cannot weight/forecast the pipeline.",
          "pipeline_coverage":"No bookings target → cannot compute pipeline coverage.",
          "aging":"No create date / age → cannot age the open pipeline.",
          "bookings_by_period":"No close date on won deals → no bookings-by-period.",
          "competitive_win_rate":"No closed Won/Lost history → no win rate or competitive win rate.",
          "sales_cycle":"No create+close dates on won deals → no sales-cycle length.",
          "new_vs_upsell":"No deal type on won deals → no new-vs-upsell split."}
    cav=[NICE.get(k, f"{k}: {v}") for k,v in m["blocked"].items()]
    if not cav: cav.append("Full analysis available — every supported metric was computed from this data.")
    for c in cav: an.cell(r,1,"•  "+c).font=BLOCKF; r+=1
    r+=1
    an.cell(r,1,"To complete the analysis, request:").font=SECT; r+=1
    req=[]
    if not m["has_closed"]: req.append("Historical CLOSED deals (Won / Lost / No-Decision) with create & close dates")
    if "weighted_pipeline" in m["blocked"]: req.append("% Probability (or stage→probability map) per open deal")
    if "pipeline_coverage" in m["blocked"]: req.append("Bookings target / quota for the period")
    req.append("Flag recurring software ARR vs. one-time services per deal")
    for x in req: an.cell(r,1,"•  "+x); r+=1

    an.column_dimensions["A"].width=42
    for col in ["B","C","D","E","F","G","H"]: an.column_dimensions[col].width=16
    an.sheet_view.showGridLines=False
    wb.save(out)

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--input",required=True); ap.add_argument("--output",required=True)
    ap.add_argument("--target",type=float,default=None); ap.add_argument("--sheet",default=None)
    ap.add_argument("--header-row",type=int,default=None); ap.add_argument("--map",default=None)
    a=ap.parse_args()
    if not Path(a.input).exists():
        print(json.dumps({"error":"input file not found","path":a.input})); return
    try:
        if a.map:
            raw=json.loads(a.map) if a.map.strip().startswith("{") else json.loads(Path(a.map).read_text())
            colmap={c:col_to_idx(v) for c,v in raw.items()}
            records,mapping,meta=load_mapped(a.input,a.sheet,colmap)
        else:
            records,mapping,meta=load_rows(a.input,a.sheet,a.header_row)
    except Exception as e:
        print(json.dumps({"error":"could not read input file","detail":str(e),"path":a.input})); return
    if not records:
        print(json.dumps({"error":"no data rows found","meta":meta})); return
    # threshold guard: must look like a deal file at all.
    # Require a value column (arr) AND something that conveys outcome/stage or open-ness.
    have=set(mapping)
    if "arr" not in have or not ({"outcome","probability","close_date"} & have):
        print(json.dumps({"error":"below Tier 1 threshold — not a usable deal-level file",
            "found_fields":sorted(have),
            "need":["arr (deal $ value)","plus at least one of: outcome/stage, probability, close_date"],
            "next":"if the file has data but no/garbled headers, re-run with --map "
                   '\'{"arr":"G","outcome":"N",...}\' (column letters); otherwise emit '
                   "references/data-request-template.md to the source",
            "meta":meta},default=str)); return
    try:
        deals,m=analyze(records,mapping,a.target)
        build_workbook(deals,m,a.output,meta,a.target,a.input)
        m["output"]=a.output; m["mapped_fields"]=sorted(mapping); m["source_sheet"]=meta.get("sheet")
        print(json.dumps(m,default=str,indent=2))
    except Exception as e:
        import traceback
        print(json.dumps({"error":"analysis failed gracefully","detail":str(e),
            "trace":traceback.format_exc().splitlines()[-3:],
            "mapped_fields":sorted(mapping)},default=str))

if __name__=="__main__":
    main()
