#!/usr/bin/env python3
"""
normalize.py — the "front door" for ARR-to-bookings.

Turns ANY customer-revenue file (messy workbook or CSV) into ONE canonical grid:
customers (one row each) x a continuous period axis (real dates), clean numbers,
plus a traceability map back to the source and a shape report. The bookings engine
consumes only this canonical form, so every quirk is handled in one place.

Handles, generally (not file-by-file):
  - dates: real dates; formula dates (cached values); text dates in many formats
    (Jan 21, 2021-01, January 2021, Q1 2021, FY21, year-only); reverse order; gaps.
  - frequency: monthly / quarterly / annual (detected from cadence).
  - orientation: wide grid / long-tidy / transaction list; data not on first sheet
    or not at the top (title/summary rows); customer column not in column A.
  - identity: one row per customer via group-sum when a customer spans many rows.
  - values: text-numbers ("9,900", "$1,542"), accounting negatives "(500)", blanks,
    "NA"; mixed text/number cells.
  - negatives: detected and listed (caller decides policy).
  - scale: MRR vs ARR vs annual-revenue inferred from magnitude + cadence.

Importable (normalize_file) and runnable (prints a shape report; --json).
"""
import argparse, calendar, datetime, json, re, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as CL

EPS_ZERO = 0.005  # a period column with |sum| below this is treated as empty
MAX_DATA_ROWS = 60000  # refuse before materializing pathologically large sheets


class RefuseError(ValueError):
    """An intentional, user-facing refusal (e.g. too large) — must propagate, not be
    swallowed by the per-sheet 'skip if unparseable' guard."""

# -----------------------------------------------------------------------------
# date / period parsing
# -----------------------------------------------------------------------------
_MONTHS = {m.lower(): i for i, m in enumerate(calendar.month_abbr) if m}
_MONTHS.update({m.lower(): i for i, m in enumerate(calendar.month_name) if m})

def _eom(y, m):
    return datetime.date(y, m, calendar.monthrange(y, m)[1])

def _yy(y):
    y = int(y)
    if y < 100:
        y += 2000 if y < 70 else 1900
    return y

def parse_period(v):
    """Return (date_at_period_end, freq) where freq in {'M','Q','A'} or (None,None).
    date is the last day of the month/quarter/year the header denotes."""
    if v is None:
        return None, None
    if isinstance(v, datetime.datetime):
        v = v.date()
    if isinstance(v, datetime.date):
        return _eom(v.year, v.month), 'M'
    # pure integer year, e.g. 2021 or 2021.0
    if isinstance(v, (int, float)):
        iv = int(v)
        if 1990 <= iv <= 2100 and float(v) == iv:
            return _eom(iv, 12), 'A'
        return None, None
    s = str(v).strip().replace("\t", " ").replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return None, None
    low = s.lower()
    # strip trailing noise like "or prior", "actual", footnote marks
    low = re.sub(r"\b(or prior|actual|actuals|est\.?|estimate|forecast)\b.*$", "", low).strip(" .*-")
    # quarter forms: Q1 2021 / 2021 Q1 / 2021-Q1 / Q1-21 / FY21 Q1
    mq = re.search(r"(?:fy\s*)?(\d{2,4})[ \-]*q([1-4])", low) or \
         re.search(r"q([1-4])[ \-/]*(?:fy\s*)?(\d{2,4})", low)
    if mq:
        g = mq.groups()
        if g[0] and g[0].isdigit() and len(g[0]) >= 2 and int(g[0]) > 4:  # year first
            y, q = _yy(g[0]), int(g[1])
        else:
            q, y = int(g[0]), _yy(g[1])
        return _eom(y, q * 3), 'Q'
    # ISO-ish: 2021-01 / 2021/01 / 2021-01-31
    m = re.fullmatch(r"(\d{4})[-/](\d{1,2})(?:[-/](\d{1,2}))?", s)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return _eom(y, mo), 'M'
    # 01/2021 or 1-2021
    m = re.fullmatch(r"(\d{1,2})[-/](\d{4})", s)
    if m and 1 <= int(m.group(1)) <= 12:
        return _eom(int(m.group(2)), int(m.group(1))), 'M'
    # Month name + year: "Jan 21", "Jan-2021", "January 2021"
    m = re.fullmatch(r"([A-Za-z]{3,9})[ \-]?(\d{2,4})", s)
    if m and m.group(1).lower() in _MONTHS:
        return _eom(_yy(m.group(2)), _MONTHS[m.group(1).lower()]), 'M'
    # "21-Jan" / "2021 Jan"
    m = re.fullmatch(r"(\d{2,4})[ \-]?([A-Za-z]{3,9})", s)
    if m and m.group(2).lower() in _MONTHS:
        return _eom(_yy(m.group(1)), _MONTHS[m.group(2).lower()]), 'M'
    # FY21 / FY2021 / bare year string
    m = re.fullmatch(r"(?:fy)?\s*(\d{4})", low) or re.fullmatch(r"fy\s*(\d{2})", low)
    if m:
        return _eom(_yy(m.group(1)), 12), 'A'
    return None, None

# -----------------------------------------------------------------------------
# value cleaning
# -----------------------------------------------------------------------------
_NA = {"", "na", "n/a", "n.a.", "-", "--", "—", "–", "nil", "none", "null", "#n/a"}

def clean_num(v):
    """Return float or None. Handles text-numbers, $ , thousands, (parens)=negative."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\t", "").replace("\xa0", "")
    if s.lower() in _NA:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = s.replace("$", "").replace("€", "").replace("£", "").replace(",", "").replace(" ", "")
    s = s.rstrip("%")
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return None

# -----------------------------------------------------------------------------
# helpers
# -----------------------------------------------------------------------------
_SUMMARY_RE = re.compile(r"^\s*(grand\s+)?(total|subtotal|sum|net|summary|all customers|"
                         r"existing customers|new revenue|expansion|contraction|churn|"
                         r"pipeline|unidentified|average|avg|count)\b", re.I)

def is_summary_label(s):
    return bool(s) and bool(_SUMMARY_RE.match(str(s).strip()))

def load_values_only(path, read_only=True):
    """Load CACHED values (data_only=True) so formula headers/values resolve."""
    return load_workbook(path, data_only=True, read_only=read_only)

# -----------------------------------------------------------------------------
# orientation detection on one worksheet
# -----------------------------------------------------------------------------
def scan_sheet(ws, max_scan_rows=60, max_scan_cols=80):
    """Return a dict describing the best interpretation of this sheet, or None."""
    grid = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        grid.append(list(row))
        if i >= max_scan_rows:
            break
    if not grid:
        return None
    ncols = min(max(len(r) for r in grid), max_scan_cols)

    # ---- WIDE: find the best period-header row. Rank by the count of DISTINCT periods
    # (a 'year super-header' like 2021,2021,... above the real month row has few distinct
    # values and must lose to the finer month header), then by total count, then lower row. ----
    best_hdr = None  # (row_idx, [(col_idx,date,freq)], first_col)
    best_key = None
    periodish_rows = 0  # rows containing >=2 parseable periods -> hint of a date COLUMN (long)
    for ri, row in enumerate(grid):
        cols = []
        for ci in range(ncols):
            v = row[ci] if ci < len(row) else None
            d, f = parse_period(v)
            if d:
                cols.append((ci, d, f))
        if len(cols) >= 2:
            periodish_rows += 1
            distinct = len({d for _, d, _ in cols})
            key = (distinct, len(cols), ri)
            if best_key is None or key > best_key:
                best_key = key
                best_hdr = (ri, cols, cols[0][0])

    # ---- LONG: find a header row naming a date column + a value column ----
    long_hit = None
    for ri, row in enumerate(grid[:8]):
        labels = {ci: str(row[ci]).strip().lower() for ci in range(min(ncols, len(row)))
                  if row[ci] not in (None, "")}
        date_col = next((ci for ci, t in labels.items()
                         if re.search(r"\b(date|month|period)\b", t)), None)
        cust_col = next((ci for ci, t in labels.items()
                         if re.search(r"\b(customer|account|client|company|name|id)\b", t)), None)
        val_col = next((ci for ci, t in labels.items()
                        if re.search(r"\b(mrr|arr|amount|revenue|gross|net|value|sales|total|billing)\b", t)), None)
        if date_col is not None and cust_col is not None and val_col is not None and cust_col != date_col:
            long_hit = dict(header_row=ri, date_col=date_col, cust_col=cust_col, val_col=val_col)
            break

    return dict(wide=best_hdr, long=long_hit, ncols=ncols, nrows=len(grid), grid=grid,
                periodish_rows=periodish_rows)

# -----------------------------------------------------------------------------
# build canonical from a WIDE sheet
# -----------------------------------------------------------------------------
_CUSTKEY_RE = re.compile(r"\b(customer|account|client|company|name|id|org)\b", re.I)

def _detect_customer_col(grid, header_ri, first_period_col):
    """Pick the customer column among the non-period columns to the left of the data.
    The customer column is the one with the most DISTINCT text labels (high
    cardinality) — a 'Type'/'Region'/constant label column has low cardinality and
    must not win. A header named customer/account/name/etc. gets a bonus."""
    best_ci, best_score = 0, -1
    hdr = grid[header_ri] if header_ri < len(grid) else []
    for ci in range(0, max(first_period_col, 1)):
        labels = [grid[r][ci] for r in range(header_ri + 1, len(grid))
                  if ci < len(grid[r])]
        nonempty = [x for x in labels if x not in (None, "")]
        texty = [x for x in nonempty if not isinstance(x, (int, float, datetime.date))]
        if len(nonempty) < 3 or len(texty) < 0.5 * len(nonempty):
            continue
        distinct = len(set(str(x).strip().lower() for x in texty))
        score = distinct
        hlabel = str(hdr[ci]).strip() if ci < len(hdr) and hdr[ci] not in (None, "") else ""
        if _CUSTKEY_RE.search(hlabel):
            score += 1000  # explicit header name wins decisively
        if score > best_score:
            best_score, best_ci = score, ci
    return best_ci

def build_wide(ws_title, full_ws, scan, neg_policy, actuals_through=None):
    header_ri, pcols, first_pc = scan["wide"]
    nrows_est = (full_ws.max_row or 0) - (header_ri + 1)
    if nrows_est > MAX_DATA_ROWS:
        raise RefuseError(f"sheet '{ws_title}' has ~{nrows_est:,} data rows (> {MAX_DATA_ROWS:,}); "
                         f"too large to materialize a per-customer model. Pre-aggregate to one row "
                         f"per customer (or sample) before running.")
    grid = scan["grid"]
    cust_ci = _detect_customer_col(grid, header_ri, first_pc)
    # read FULL column extents (scan grid may be truncated) for customer rows
    # gather customer rows from the full sheet
    rows = []
    for r in full_ws.iter_rows(min_row=header_ri + 2, values_only=True):
        rows.append(list(r))
    # period source columns. Dedupe repeated date blocks (actuals + adjusted/derived
    # columns often reuse the same months) — keep the FIRST (leftmost) occurrence.
    period_src = []
    seen_dates = set()
    freqs = []
    dup_blocks = 0
    for (ci, d, f) in pcols:
        if d in seen_dates:
            dup_blocks += 1
            continue
        seen_dates.add(d)
        period_src.append((d, ci))
        freqs.append(f)
    # customer aggregation (group by name; sum each period)
    agg = {}        # name -> {date: value}
    src_rows = {}   # name -> [src_row_indices]
    dropped = []
    negatives = []
    multi = False
    all_numeric = True  # True only if every consumed value was already a real number
    for idx, r in enumerate(rows):
        src_row = header_ri + 2 + idx  # 1-based row in sheet
        name = r[cust_ci] if cust_ci < len(r) else None
        if name in (None, ""):
            continue
        name = str(name).strip()
        if is_summary_label(name):
            dropped.append((src_row, name)); continue
        vals = {}
        any_num = False
        for (d, ci) in period_src:
            raw = r[ci] if ci < len(r) else None
            if isinstance(raw, str) and raw.strip() != "":
                all_numeric = False
            n = clean_num(raw)
            if n is not None:
                any_num = True
                if n < 0:
                    negatives.append((name, d.isoformat(), n))
                vals[d] = n
        if not any_num:
            dropped.append((src_row, name + " (no numeric data)")); continue
        if name in agg:
            multi = True
            for d, n in vals.items():
                agg[name][d] = agg[name].get(d, 0.0) + n
            src_rows[name].append(src_row)
        else:
            agg[name] = vals
            src_rows[name] = [src_row]
    orientation = "wide_multirow" if multi else "wide_direct"
    trace = dict(orientation=orientation, sheet=ws_title, header_row=header_ri + 1,
                 customer_col=cust_ci + 1, period_src=[(d.isoformat(), ci + 1) for d, ci in period_src],
                 src_rows={k: v for k, v in src_rows.items()},
                 all_numeric=all_numeric)
    return _finalize(agg, freqs, trace, negatives, dropped, neg_policy, actuals_through)

# -----------------------------------------------------------------------------
# build canonical from a LONG / transaction sheet
# -----------------------------------------------------------------------------
def build_long(ws_title, full_ws, scan, neg_policy, actuals_through=None):
    lh = scan["long"]
    hr, dc, cc, vc = lh["header_row"], lh["date_col"], lh["cust_col"], lh["val_col"]
    nrows_est = (full_ws.max_row or 0) - (hr + 1)
    if nrows_est > MAX_DATA_ROWS:
        raise RefuseError(f"sheet '{ws_title}' has ~{nrows_est:,} rows (> {MAX_DATA_ROWS:,}); "
                         f"too large to load. Pre-aggregate or sample before running.")
    agg = {}
    negatives = []
    freqs = []
    all_numeric = True  # True -> value column is real numbers -> helper can use SUMIFS
    for r in full_ws.iter_rows(min_row=hr + 2, values_only=True):
        name = r[cc] if cc < len(r) else None
        if name in (None, ""):
            continue
        name = str(name).strip()
        if is_summary_label(name):
            continue
        d, f = parse_period(r[dc] if dc < len(r) else None)
        if not d:
            continue
        freqs.append(f)
        rawv = r[vc] if vc < len(r) else None
        if isinstance(rawv, str) and rawv.strip() != "":
            all_numeric = False
        n = clean_num(rawv)
        if n is None:
            continue
        if n < 0:
            negatives.append((name, d.isoformat(), n))
        agg.setdefault(name, {})
        agg[name][d] = agg[name].get(d, 0.0) + n
    trace = dict(orientation="long", sheet=ws_title, header_row=hr + 1,
                 date_col=dc + 1, customer_col=cc + 1, value_col=vc + 1, all_numeric=all_numeric)
    return _finalize(agg, freqs, trace, negatives, [], neg_policy, actuals_through)

# -----------------------------------------------------------------------------
# finalize: continuous axis, frequency, scale, negative policy
# -----------------------------------------------------------------------------
def _modal_freq(freqs, dates):
    if dates and len(dates) >= 2:
        gaps = []
        sd = sorted(set(dates))
        for a, b in zip(sd, sd[1:]):
            gaps.append((b.year - a.year) * 12 + (b.month - a.month))
        gaps = [g for g in gaps if g > 0]
        if gaps:
            g = sorted(gaps)[len(gaps) // 2]
            if g >= 12:
                return 'A'
            if g >= 3:
                return 'Q'
            return 'M'
    if freqs:
        return max(set(freqs), key=freqs.count)
    return 'M'

def _axis(dates, freq):
    lo, hi = min(dates), max(dates)
    out = []
    if freq == 'A':
        for y in range(lo.year, hi.year + 1):
            out.append(_eom(y, 12))
    elif freq == 'Q':
        y, m = lo.year, ((lo.month - 1) // 3) * 3 + 3
        while (y, m) <= (hi.year, hi.month):
            out.append(_eom(y, m))
            m += 3
            if m > 12:
                m -= 12; y += 1
    else:
        y, m = lo.year, lo.month
        while (y, m) <= (hi.year, hi.month):
            out.append(_eom(y, m))
            m += 1
            if m > 12:
                m = 1; y += 1
    return out

def _parse_cutoff(actuals_through):
    if not actuals_through:
        return None
    s = str(actuals_through)
    d, _ = parse_period(s)
    return d

def _finalize(agg, freqs, trace, negatives, dropped, neg_policy, actuals_through=None):
    if not agg:
        return None
    cutoff = _parse_cutoff(actuals_through)
    all_dates = sorted({d for vals in agg.values() for d in vals})
    if cutoff:
        all_dates = [d for d in all_dates if d <= cutoff]
    if len(all_dates) < 2:
        return None
    freq = _modal_freq(freqs, all_dates)
    axis = _axis(all_dates, freq)
    # trim trailing periods that are entirely empty/zero across all customers (forecast tail
    # or a grid that extends past the last booked month)
    def colsum(d):
        return sum(abs(vals.get(d, 0) or 0) for vals in agg.values())
    while len(axis) > 2 and colsum(axis[-1]) < EPS_ZERO:
        axis = axis[:-1]
    gaps_filled = len([d for d in axis if d not in set(all_dates)])
    # apply negative policy
    def apply(n):
        if n is not None and n < 0:
            if neg_policy == "zero":
                return 0.0
            # "contraction"/"leave" keep the value; engine treats via downsell/churn
        return n
    customers = sorted(agg.keys(), key=_natkey)
    matrix = []
    for name in customers:
        row = [apply(agg[name].get(d)) or 0.0 for d in axis]
        matrix.append(row)
    # scale inference: median of each customer's first nonzero value
    firsts = []
    for row in matrix:
        for x in row:
            if x and x > 0:
                firsts.append(x); break
    firsts.sort()
    med = firsts[len(firsts) // 2] if firsts else 0
    if freq == 'A':
        scale = "REVENUE"; factor = 1
    elif med > 50000:
        scale = "ARR"; factor = 1
    else:
        scale = "MRR"; factor = 12
    return dict(customers=customers, axis=[d.isoformat() for d in axis], matrix=matrix,
                frequency=freq, scale=scale, arr_factor=factor, trace=trace,
                negatives=negatives, dropped=dropped, gaps_filled=gaps_filled,
                n_customers=len(customers), n_periods=len(axis),
                multirow=(trace.get("orientation") == "wide_multirow"))

def _natkey(s):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", str(s))]

# -----------------------------------------------------------------------------
# top level
# -----------------------------------------------------------------------------
def _recalc_to_temp(path):
    """Recalc a workbook with LibreOffice so formula headers/values get cached.
    Returns a new file path, or None if LibreOffice isn't available."""
    import os, shutil, subprocess, tempfile
    soffice = None
    for c in ("soffice", "/Applications/LibreOffice.app/Contents/MacOS/soffice",
              "/usr/bin/soffice", "/opt/libreoffice/program/soffice"):
        if shutil.which(c) or os.path.exists(c):
            soffice = c; break
    if not soffice:
        return None
    d = tempfile.mkdtemp(prefix="a2b_recalc_")
    src = os.path.join(d, "in.xlsx")
    shutil.copy(path, src)
    try:
        subprocess.run([soffice, "--headless", "--calc", "--convert-to", "xlsx",
                        "--outdir", d, src], check=True, capture_output=True, timeout=600)
    except Exception:  # noqa
        return None
    out = os.path.join(d, "in.xlsx")
    return out if os.path.exists(out) else None

def normalize_file(path, sheet=None, neg_policy="leave", prefer_sheet_with=None, actuals_through=None, _recalced=False):
    """Return canonical dict (best sheet) or raise ValueError with the reason."""
    if str(path).lower().endswith(".csv"):
        return _normalize_csv(path, neg_policy, actuals_through)
    wb = load_values_only(path, read_only=True)
    candidates = wb.sheetnames if sheet is None else [sheet]
    results = []
    for sn in candidates:
        ws = wb[sn]
        scan = scan_sheet(ws)
        if not scan:
            continue
        # re-open a non-read-only handle for full extraction? read_only iter is fine.
        try:
            # Prefer LONG when a Customer+Date+Value header exists AND periods repeat down
            # many rows (a date COLUMN, not a header) or the wide candidate is weak — this
            # stops a transaction log's data rows from being misread as a wide header.
            wide_distinct = len({d for _, d, _ in scan["wide"][1]}) if scan["wide"] else 0
            prefer_long = scan["long"] and (scan["periodish_rows"] >= 3 or not scan["wide"]
                                            or wide_distinct <= 4)
            if prefer_long:
                can = build_long(sn, ws, scan, neg_policy, actuals_through)
            elif scan["wide"]:
                can = build_wide(sn, ws, scan, neg_policy, actuals_through)
            elif scan["long"]:
                can = build_long(sn, ws, scan, neg_policy, actuals_through)
            else:
                can = None
        except RefuseError:
            raise
        except Exception as e:  # noqa
            can = None
        if can:
            can["sheet"] = sn
            can["score"] = can["n_customers"] * can["n_periods"]
            results.append(can)
    if not results:
        # Possibly formula headers/values that weren't cached (e.g. =EOMONTH chains in
        # a file never opened by Excel). Recalc once with LibreOffice and retry.
        if not _recalced:
            rp = _recalc_to_temp(path)
            if rp:
                return normalize_file(rp, sheet=sheet, neg_policy=neg_policy,
                                      prefer_sheet_with=prefer_sheet_with, actuals_through=actuals_through, _recalced=True)
        raise ValueError("No sheet looks like customer revenue: need >=2 parseable period "
                         "headers (wide) or Customer+Date+Value columns (long). "
                         "If headers are formulas, install LibreOffice so they can be recalculated.")
    results.sort(key=lambda r: r["score"], reverse=True)
    return results[0]

def _normalize_csv(path, neg_policy, actuals_through=None):
    import csv
    with open(path, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        raise ValueError("empty CSV")
    hdr = rows[0]
    low = [h.strip().lower() for h in hdr]
    def find(pat):
        return next((i for i, t in enumerate(low) if re.search(pat, t)), None)
    dc = find(r"\b(date|month|period)\b")
    cc = find(r"\b(customer|account|client|company|name|id)\b")
    vc = find(r"\b(mrr|arr|amount|revenue|gross|net|value|sales|billing)\b")
    if dc is not None and cc is not None and vc is not None:
        agg = {}; negatives = []; freqs = []
        for r in rows[1:]:
            if len(r) <= max(dc, cc, vc):
                continue
            name = r[cc].strip()
            if not name or is_summary_label(name):
                continue
            d, f = parse_period(r[dc])
            if not d:
                continue
            freqs.append(f)
            n = clean_num(r[vc])
            if n is None:
                continue
            if n < 0:
                negatives.append((name, d.isoformat(), n))
            agg.setdefault(name, {})
            agg[name][d] = agg[name].get(d, 0.0) + n
        # CSV cells are text -> SUMIFS can't sum them; use the cleaned-value helper path.
        trace = dict(orientation="long_csv", header_row=1, date_col=dc + 1, customer_col=cc + 1,
                     value_col=vc + 1, all_numeric=False)
        can = _finalize(agg, freqs, trace, negatives, [], neg_policy, actuals_through)
        if can:
            can["sheet"] = "(csv)"
            return can
        raise ValueError("CSV has Customer/Date/Value columns but <2 distinct periods — looks "
                         "like a current-state snapshot (one row per customer), not a time "
                         "series. Bookings need monthly/quarterly history.")
    raise ValueError("CSV not recognized as long/tidy (need Customer, Date, Value columns).")


def main():
    ap = argparse.ArgumentParser(description="Normalize any customer-revenue file to a canonical grid + report.")
    ap.add_argument("path")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--neg-policy", default="leave", choices=["leave", "zero", "contraction"])
    ap.add_argument("--actuals-through", default=None, help="drop periods after this (YYYY-MM or 'May-26')")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()
    try:
        can = normalize_file(args.path, sheet=args.sheet, neg_policy=args.neg_policy, actuals_through=args.actuals_through)
    except ValueError as e:
        print(f"REFUSE: {e}")
        sys.exit(2)
    if args.json:
        out = {k: v for k, v in can.items() if k != "matrix"}
        out["matrix_shape"] = [len(can["matrix"]), len(can["matrix"][0]) if can["matrix"] else 0]
        print(json.dumps(out, indent=2, default=str))
        return
    print(f"FILE: {args.path}")
    print(f"  sheet: {can['sheet']}  orientation: {can['trace'].get('orientation')}")
    print(f"  frequency: {can['frequency']}  scale: {can['scale']} (factor {can['arr_factor']})")
    print(f"  customers: {can['n_customers']}  periods: {can['n_periods']}  "
          f"axis: {can['axis'][0]} -> {can['axis'][-1]}")
    print(f"  gaps filled: {can['gaps_filled']}  negatives: {len(can['negatives'])}  "
          f"dropped rows: {len(can['dropped'])}")
    if can["frequency"] == "A":
        print("  NOTE: annual frequency -> quarterly bookings not possible; annual-only bookings.")
    if can["negatives"]:
        print(f"  NEGATIVES (first 5): {can['negatives'][:5]}  -> choose --neg-policy")
    print(f"  sample customers: {can['customers'][:4]}")


if __name__ == "__main__":
    main()
