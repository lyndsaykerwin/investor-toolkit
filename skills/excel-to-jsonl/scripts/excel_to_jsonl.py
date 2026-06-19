#!/usr/bin/env python3
# Plain-English header: this file's job is to turn any Excel/CSV file into a
# ".jsonl" sidecar — a plain-text file with one line of JSON per data row, plus
# a first "_meta" line describing the sheet. It's the agent-readable twin of a
# spreadsheet, so an AI can read the data without opening the .xlsx itself.
#
# Usage:
#   python3 excel_to_jsonl.py <file.xlsx|file.csv> [options]
# Options:
#   --sheet NAME        Convert a specific sheet (default: the most data-rich one)
#   --all-sheets        Convert every sheet -> one <name>.<sheet>.jsonl each
#   --header-row N      Force the header to be on row N (1-based). Default: auto-detect
#   --out PATH          Write to PATH instead of <source>.jsonl
#   --formulas          Keep formula text (=A1+B1) instead of the last-saved value
#   --max-rows N        Stop after N data rows (handy for sampling huge files)
#
# Output: <source>.jsonl next to the source file (or --out / per-sheet names).

import sys, os, json, argparse, csv, datetime

def log(*a): print(*a, file=sys.stderr)

# ---------- value normalization ----------
def norm(v):
    """Turn a cell value into something JSON can hold, in a predictable way."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        # ISO date/datetime strings sort and parse cleanly for downstream agents
        if isinstance(v, datetime.datetime) and (v.hour or v.minute or v.second):
            return v.isoformat()
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    if isinstance(v, float):
        # 3.0 -> 3 so whole numbers don't read as floats; keep real decimals
        return int(v) if v.is_integer() else v
    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None
    return v  # int, bool pass through

def infer_type(values):
    """Best-guess type for a column, from its non-empty values."""
    seen = set()
    for v in values:
        if v is None: continue
        if isinstance(v, bool): seen.add("boolean")
        elif isinstance(v, int): seen.add("number")
        elif isinstance(v, float): seen.add("number")
        elif isinstance(v, str):
            # ISO-ish date string?
            seen.add("date" if _looks_date(v) else "string")
        else: seen.add("string")
    if not seen: return "empty"
    if seen == {"number"}: return "number"
    if seen == {"boolean"}: return "boolean"
    if seen == {"date"}: return "date"
    if seen <= {"number", "string"} and "string" in seen: return "string"
    return "mixed" if len(seen) > 1 else seen.pop()

def _looks_date(s):
    try:
        datetime.date.fromisoformat(s[:10]); return len(s) >= 8 and s[4] == "-"
    except Exception:
        return False

# ---------- header detection ----------
def detect_header(rows):
    """Pick the header row. Spreadsheets often stack a title and a repeated
    "banner" row (e.g. 'Unadj.' across every date column) above the real column
    headers. A banner repeats the SAME word many times, while a true header is a
    row of many DISTINCT text labels — so we score each early row by its number
    of distinct string values and take the best one that has data beneath it.
    This beats a width/text-ratio heuristic, which a wide banner row can fool."""
    if not rows: return 0
    best_i, best_score = None, 0
    for i, r in enumerate(rows[:25]):
        distinct_text = {c for c in r if isinstance(c, str) and c != ""}
        if len(distinct_text) < 2:
            continue
        # require at least one populated row beneath it (real data follows)
        if not any(any(c is not None for c in rows[j]) for j in range(i + 1, len(rows))):
            continue
        score = len(distinct_text)
        # strict '>' keeps the EARLIEST row on ties (header sits above its data)
        if score > best_score:
            best_score, best_i = score, i
    if best_i is not None:
        return best_i
    # fallback: first non-empty row
    for i, r in enumerate(rows):
        if any(c is not None for c in r): return i
    return 0

def make_keys(header):
    """Column header text -> JSON keys. Blank/duplicate headers get safe names."""
    keys, seen = [], {}
    for idx, h in enumerate(header):
        name = h.strip() if isinstance(h, str) and h.strip() else f"column_{idx+1}"
        if name in seen:
            seen[name] += 1; name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        keys.append(name)
    return keys

# ---------- readers ----------
def read_xlsx_sheet(ws, formulas):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([None if c is None else (c if formulas else norm(c)) for c in row])
    # trim trailing all-empty rows
    while rows and all(c is None for c in rows[-1]): rows.pop()
    return rows

def read_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        raw = list(csv.reader(f))
    return [[norm(c) for c in r] for r in raw]

# ---------- core conversion ----------
def rows_to_records(rows, header_row=None, max_rows=None):
    h = (header_row - 1) if header_row else detect_header(rows)
    header = rows[h] if h < len(rows) else []
    keys = make_keys(header)
    records, columns_vals = [], {k: [] for k in keys}
    for r in rows[h + 1:]:
        if all(c is None for c in r):  # skip blank rows
            continue
        rec = {}
        for k, c in zip(keys, r):
            rec[k] = c
            columns_vals[k].append(c)
        # carry any cells beyond the header width under overflow keys
        for j in range(len(keys), len(r)):
            if r[j] is not None:
                rec[f"column_{j+1}"] = r[j]
        records.append(rec)
        if max_rows and len(records) >= max_rows:
            break
    columns_meta = [{"name": k, "type": infer_type(columns_vals[k]),
                     "sample": next((v for v in columns_vals[k] if v is not None), None)}
                    for k in keys]
    return records, columns_meta, h

def write_jsonl(out_path, meta, records):
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(json.dumps(meta, ensure_ascii=False) + "\n")
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

def build_meta(source, sheet, columns, n, header_row_idx):
    return {
        "_meta": True,
        "source_file": os.path.basename(source),
        "sheet": sheet,
        "generated": datetime.date.today().isoformat(),
        "header_row": header_row_idx + 1,   # 1-based, for humans
        "row_count": n,
        "columns": columns,
    }

# ---------- main ----------
def main():
    ap = argparse.ArgumentParser(description="Convert an Excel/CSV file to a JSONL sidecar.")
    ap.add_argument("source")
    ap.add_argument("--sheet")
    ap.add_argument("--all-sheets", action="store_true")
    ap.add_argument("--header-row", type=int)
    ap.add_argument("--out")
    ap.add_argument("--formulas", action="store_true")
    ap.add_argument("--max-rows", type=int)
    args = ap.parse_args()

    src = args.source
    base, ext = os.path.splitext(src)
    ext = ext.lower()
    outputs = []

    if ext == ".csv":
        rows = read_csv(src)
        records, cols, h = rows_to_records(rows, args.header_row, args.max_rows)
        out = args.out or base + ".jsonl"
        write_jsonl(out, build_meta(src, None, cols, len(records), h), records)
        outputs.append((out, len(records)))
    elif ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(src, read_only=True, data_only=not args.formulas)
        sheets = wb.sheetnames
        if args.all_sheets:
            targets = sheets
        elif args.sheet:
            if args.sheet not in sheets:
                log(f"Sheet '{args.sheet}' not found. Available: {sheets}"); sys.exit(1)
            targets = [args.sheet]
        else:
            # pick the most data-rich sheet (most non-empty cells)
            def density(name):
                ws = wb[name]; return sum(1 for row in ws.iter_rows(values_only=True)
                                          for c in row if c is not None)
            targets = [max(sheets, key=density)]
            if len(sheets) > 1:
                log(f"Auto-selected sheet '{targets[0]}' (most data). "
                    f"Use --sheet NAME or --all-sheets to change.")
        for name in targets:
            rows = read_xlsx_sheet(wb[name], args.formulas)
            records, cols, h = rows_to_records(rows, args.header_row, args.max_rows)
            if args.out and len(targets) == 1:
                out = args.out
            elif len(targets) > 1:
                safe = "".join(ch if ch.isalnum() else "_" for ch in name)
                out = f"{base}.{safe}.jsonl"
            else:
                out = base + ".jsonl"
            write_jsonl(out, build_meta(src, name, cols, len(records), h), records)
            outputs.append((out, len(records)))
    else:
        log(f"Unsupported file type: {ext}. Use .xlsx, .xlsm, or .csv."); sys.exit(1)

    for out, n in outputs:
        print(f"Wrote {out}  ({n} rows)")

if __name__ == "__main__":
    main()
