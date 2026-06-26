#!/usr/bin/env python3
"""
build_pnl_workbook.py — deterministic P&L workbook builder.

Input : a normalized P&L JSON (schema in references/normalized_schema.md).
Output: an .xlsx with two sheets:
  • Source_Audit — every line transposed verbatim; subtotals/computed lines are
    live formulas checked against the source-stated number (check-to-zero).
  • Output — clean, single-highlight-color summary with margins, growth and an
    actionable follow-up list.

Formatting rules (locked):
  1. NEVER merge & center. Use Center Across Selection (alignment=centerContinuous).
  2. ONE highlight color for all major lines (Revenue / Gross Profit / EBITDA / Net Income).
  3. Notes are actionable FOLLOW-UPS (classification / functional-split / comparability),
     never methodology prose.

Usage: python3 build_pnl_workbook.py <input.json> <output.xlsx>
"""
import sys, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ---------- palette / styles ---------------------------------------------
NAVY = "1F3864"; HILITE = "D6E0F0"; GREY = "F2F2F2"; AMBER = "FCE4D6"; GREEN_OK = "E2EFDA"
FNAME = "Calibri"
thin = Side(style="thin", color="BFBFBF")
top_border = Border(top=thin)
USD = '#,##0.00;(#,##0.00)'      # source: verbatim 2dp
USD0 = '#,##0;(#,##0)'           # output: whole dollars
PCT = '0%;(0%)'                  # whole-number percentages by default
CHK = '#,##0.00;[Red](#,##0.00)'
CENTER = Alignment(horizontal="centerContinuous")   # Center Across Selection
CTR = Alignment(horizontal="center")
LEFT = Alignment(horizontal="left")
RIGHT = Alignment(horizontal="right")

def F(sz=10, b=False, color="000000", italic=False):
    return Font(name=FNAME, size=sz, bold=b, color=color, italic=italic)
def fill(c):
    return PatternFill("solid", fgColor=c)

def center_across(ws, row, c_from, c_to, text, font):
    """Center Across Selection: text in leftmost cell, centerContinuous on the whole span.
    No merging — keeps cells individually addressable (and avoids MergedCell bugs)."""
    i0 = column_index_from_string(c_from); i1 = column_index_from_string(c_to)
    for ci in range(i0, i1 + 1):
        cell = ws.cell(row=row, column=ci)
        cell.alignment = CENTER
        cell.font = font
        if ci == i0:
            cell.value = text

# =========================================================================
def build(data, out_path):
    periods = data["periods"]
    pids = [p["id"] for p in periods]
    n = len(periods)
    VCOLS = [get_column_letter(2 + i) for i in range(n)]              # B,C,...
    gap1 = column_index_from_string(VCOLS[-1]) + 1                    # gap col
    SCOLS = [get_column_letter(gap1 + 1 + i) for i in range(n)]       # stated
    gap2 = column_index_from_string(SCOLS[-1]) + 1
    KCOLS = [get_column_letter(gap2 + 1 + i) for i in range(n)]       # check

    wb = openpyxl.Workbook()
    src = wb.active; src.title = "Source_Audit"
    rowref = {}   # key -> source row

    def S(col, row, val, font=None, numfmt=None, fillc=None, align=None, border=None):
        c = src.cell(row=row, column=column_index_from_string(col), value=val)
        if font: c.font = font
        if numfmt: c.number_format = numfmt
        if fillc: c.fill = fill(fillc)
        if align: c.alignment = align
        if border: c.border = border
        return c

    r = 1
    m = data.get("meta", {})
    center_across(src, r, "A", VCOLS[-1], f"{m.get('company','')} — {m.get('statement','Profit & Loss')}", F(14, True, NAVY)); r += 1
    center_across(src, r, "A", VCOLS[-1], "Source & Audit — every line transposed verbatim from source. Subtotals = live formulas; audit block (right) must read 0.", F(9, color="595959", italic=True)); r += 1
    r += 1

    # header band
    hdr = r
    S("A", hdr, "Account / Line item", F(10, True, "FFFFFF"), fillc=NAVY, align=LEFT)
    for i, c in enumerate(VCOLS):
        S(c, hdr, periods[i]["label"], F(10, True, "FFFFFF"), fillc=NAVY, align=CTR)
    center_across(src, hdr - 1, SCOLS[0], SCOLS[-1], "— per source (stated) —", F(8, True, "595959"))
    center_across(src, hdr - 1, KCOLS[0], KCOLS[-1], "— check: computed − stated (=0) —", F(8, True, "595959"))
    for i in range(n):
        S(SCOLS[i], hdr, periods[i]["label"], F(9, True, "595959"), align=CTR)
        S(KCOLS[i], hdr, periods[i]["label"], F(9, True, "595959"), align=CTR)
    r += 1

    def vals_of(line, pid):
        return (line.get("values") or {}).get(pid)

    # index detail lines by key
    detail_by_key = {ln["key"]: ln for ln in data["lines"] if ln["role"] == "detail" and "key" in ln}

    # period-presence for ANY keyed line (detail OR nested subtotal/computed).
    # detail lines signal presence via `values`; subtotal/computed via `stated`.
    def member_present(key, pid):
        d = detail_by_key.get(key)
        if d is not None:
            return (d.get("values") or {}).get(pid) is not None
        for ln2 in data["lines"]:
            if ln2.get("key") == key and ln2["role"] in ("subtotal", "computed"):
                return (ln2.get("stated") or {}).get(pid) is not None
        return False

    for ln in data["lines"]:
        role = ln["role"]
        if role == "section":
            S("A", r, ln["label"], F(10, True, NAVY), fillc=GREY)
            for c in VCOLS: S(c, r, None, fillc=GREY)
            r += 1
        elif role == "group_header":
            S("A", r, "  " + ln["label"], F(9, True, "404040")); r += 1
        elif role == "detail":
            indent = ln.get("indent", 2)
            S("A", r, ("    " * (indent - 1)) + ln["label"], F(9))
            for i, c in enumerate(VCOLS):
                v = vals_of(ln, pids[i])
                if v is not None:
                    S(c, r, v, F(9), numfmt=USD)
            rowref[ln["key"]] = r; r += 1
        elif role in ("subtotal", "computed"):
            major = ln.get("major", False)
            fc = HILITE if major else None
            fnt = F(10, True) if major else F(9, True)
            S("A", r, ("  " * ln.get("level", 1)) + ln["label"], fnt, fillc=fc, border=top_border)
            for i, c in enumerate(VCOLS):
                if role == "subtotal":
                    cells = [f"{c}{rowref[k]}" for k in ln["members"]
                             if rowref.get(k) and member_present(k, pids[i])]
                    formula = ("=" + "+".join(cells)) if cells else None
                else:  # computed: token list of keys and +/-
                    toks = ln["formula"]; parts = []
                    ok = True
                    for t in toks:
                        if t in ("+", "-"):
                            parts.append(t)
                        else:
                            if rowref.get(t) is None:
                                ok = False; break
                            parts.append(f"{c}{rowref[t]}")
                    formula = ("=" + "".join(parts)) if ok and parts else None
                if formula:
                    S(c, r, formula, fnt, numfmt=USD, fillc=fc, border=top_border)
                else:
                    S(c, r, None, fillc=fc, border=top_border)
                sv = (ln.get("stated") or {}).get(pids[i])
                if sv is not None:
                    S(SCOLS[i], r, sv, F(9, color="595959"), numfmt=USD)
                    if formula:
                        S(KCOLS[i], r, f"=ROUND({c}{r}-{SCOLS[i]}{r},2)", F(9), numfmt=CHK)
            rowref[ln["key"]] = r; r += 1
        elif role == "blank":
            r += 1

    # master check
    r += 1
    mc = r
    center_across(src, mc, "A", VCOLS[-1], "MASTER AUDIT CHECK  (sum of all checks — must = 0)", F(11, True, "FFFFFF"))
    for c in ("A",) + tuple(VCOLS):
        src.cell(row=mc, column=column_index_from_string(c)).fill = fill(NAVY)
    chk_cells = []
    for ln in data["lines"]:
        if ln["role"] in ("subtotal", "computed") and ln.get("stated"):
            rr = rowref.get(ln["key"])
            if rr:
                chk_cells += [f"{c}{rr}" for c in KCOLS]
    fcol = get_column_letter(column_index_from_string(VCOLS[-1]) + 1)
    mcell = src.cell(row=mc, column=column_index_from_string(fcol),
                     value=f"=ROUND(SUM({','.join(chk_cells)}),2)" if chk_cells else 0)
    mcell.font = F(12, True, "006100"); mcell.number_format = CHK; mcell.fill = fill(GREEN_OK); mcell.alignment = CTR

    src.column_dimensions["A"].width = 46
    for c in VCOLS + SCOLS + KCOLS:
        src.column_dimensions[c].width = 13
    src.column_dimensions[get_column_letter(gap1)].width = 3
    src.column_dimensions[get_column_letter(gap2)].width = 3
    src.freeze_panes = f"B{hdr+1}"

    # =====================================================================
    # OUTPUT SHEET
    # =====================================================================
    out = wb.create_sheet("Output", 0)
    OC = [get_column_letter(2 + i) for i in range(n)]
    omap = data["output"]

    def O(col, row, val, font=None, numfmt=None, fillc=None, align=None, border=None):
        c = out.cell(row=row, column=column_index_from_string(col), value=val)
        if font: c.font = font
        if numfmt: c.number_format = numfmt
        if fillc: c.fill = fill(fillc)
        if align: c.alignment = align
        if border: c.border = border
        return c

    def link(key, i):
        # null-safe: a missing/unresolved key renders 0 (e.g. no-COGS P&Ls).
        if not key or rowref.get(key) is None:
            return 0
        return f"=Source_Audit!{VCOLS[i]}{rowref[key]}"
    def link_sum(keys, i):
        present = [k for k in keys if rowref.get(k) is not None]
        if not present: return None
        return "=" + "+".join(f"Source_Audit!{VCOLS[i]}{rowref[k]}" for k in present)

    rr = 1
    O("A", rr, m.get("company", ""), F(16, True, NAVY)); rr += 1
    full_ids = [p["id"] for p in periods if p.get("full_year")]
    sub = m.get("statement", "Profit & Loss")   # meta-driven; set "Consolidated P&L" only when truly consolidated
    center_across(out, rr, "A", OC[-1], sub, F(11, True, "404040")); rr += 1
    O("A", rr, "$ figures. All values link to Source_Audit. Margins & growth computed live.", F(9, italic=True, color="595959")); rr += 1
    partial = [p["label"] for p in periods if not p.get("full_year")]
    if partial:
        center_across(out, rr, "A", OC[-1],
                      f"⚠ Partial period(s): {', '.join(partial)} — not comparable to full years; YoY growth into them is shown n/m.",
                      F(9, True, "C00000"))
        for i in range(n + 1):
            out.cell(row=rr, column=1 + i).fill = fill(AMBER)
        rr += 1
    rr += 1

    h = rr
    O("A", h, "($)", F(10, True, "FFFFFF"), fillc=NAVY)
    for i, c in enumerate(OC):
        O(c, h, periods[i]["label"], F(10, True, "FFFFFF"), fillc=NAVY, align=Alignment(horizontal="center", wrap_text=True))
    out.row_dimensions[h].height = 30
    rr += 1

    def major(label, key, val_formula=None):
        global_rr = None
        O("A", rr_box[0], label, F(11, True), fillc=HILITE, border=top_border)
        for i, c in enumerate(OC):
            v = val_formula(i, c) if val_formula else link(key, i)
            O(c, rr_box[0], v, F(11, True), numfmt=USD0, fillc=HILITE, border=top_border)
        kr = rr_box[0]; rr_box[0] += 1
        return kr

    rr_box = [rr]  # mutable row cursor for helpers below
    def line(label, key=None, formula=None, indent=1, italic=False, bold=False, border=False):
        bd = top_border if border else None
        O("A", rr_box[0], ("   " * indent) + label, F(10, italic=italic, b=bold), border=bd)
        for i, c in enumerate(OC):
            v = formula(i, c) if formula else link(key, i)
            if v is not None:
                O(c, rr_box[0], v, F(10, italic=italic, b=bold), numfmt=USD0, border=bd)
        kr = rr_box[0]; rr_box[0] += 1
        return kr
    def pct_line(label, num, den, indent=1):
        # same color/size as the normal P&L lines (F(10) black)
        O("A", rr_box[0], ("   " * indent) + label, F(10))
        for i, c in enumerate(OC):
            O(c, rr_box[0], f"={c}{num}/{c}{den}", F(10), numfmt=PCT, align=RIGHT)
        rr_box[0] += 1
    def growth_line(label, vr, indent=1):
        if n < 2:
            return  # single period: no prior year to grow from — skip the row entirely
        O("A", rr_box[0], ("   " * indent) + label, F(10))
        for i in range(n):
            c = OC[i]
            if i == 0:
                rr_box[0]  # leftmost: no prior period -> blank
                continue
            this_full = periods[i].get("full_year"); prev_full = periods[i-1].get("full_year")
            if this_full and prev_full:
                p = OC[i-1]
                O(c, rr_box[0], f"=({c}{vr}-{p}{vr})/ABS({p}{vr})", F(10), numfmt=PCT, align=RIGHT)
            else:
                O(c, rr_box[0], "n/m", F(10, color="A6A6A6"), align=RIGHT)  # right-aligned, matches % cells
        rr_box[0] += 1

    # Revenue
    rev = line("Revenue", omap["revenue"], indent=0, bold=True)  # placeholder; recolor as major below
    # turn rev row into major styling
    for col in ["A"] + OC:
        cc = out[f"{col}{rev}"]; cc.fill = fill(HILITE); cc.font = F(11, True); cc.border = top_border
    out[f"A{rev}"].value = "Revenue"
    growth_line("Revenue growth (YoY)", rev)
    cogs = line("Cost of goods sold", omap.get("cogs"))   # 0 when no COGS in source
    # collapse COGS into an expandable group (hidden by default; user can show it)
    out.row_dimensions[cogs].outlineLevel = 1
    out.row_dimensions[cogs].hidden = True
    rr_box[0] += 1   # blank spacer between Revenue growth and Gross Profit
    if omap.get("gross_profit"):
        gp = line("Gross Profit", omap["gross_profit"], indent=0)
    else:
        # no gross-profit line in source → compute Revenue − COGS (honest: = Revenue when COGS is 0)
        gp = line("Gross Profit", formula=lambda i, c: f"={c}{rev}-{c}{cogs}", indent=0)
    for col in ["A"] + OC:
        cc = out[f"{col}{gp}"]; cc.fill = fill(HILITE); cc.font = F(11, True); cc.border = top_border
    out[f"A{gp}"].value = "Gross Profit"
    pct_line("Gross margin %", gp, rev)
    rr_box[0] += 1

    O("A", rr_box[0], "Operating expenses (excl. interest & D&A)", F(10, True, "404040")); rr_box[0] += 1
    opex_rows = []
    for item in omap["opex"]:
        keys = item.get("keys") or [item["key"]]
        kr = line(item["label"], formula=lambda i, c, kk=keys: link_sum(kk, i))
        opex_rows.append(kr)
    totopex = rr_box[0]
    O("A", totopex, "   Total operating expenses (excl. int. & D&A)", F(10, True), border=top_border)
    for i, c in enumerate(OC):
        O(c, totopex, "=" + "+".join(f"{c}{x}" for x in opex_rows), F(10, True), numfmt=USD0, border=top_border)
    rr_box[0] += 2

    ebitda = rr_box[0]
    O("A", ebitda, "EBITDA", F(11, True), fillc=HILITE, border=top_border)
    for i, c in enumerate(OC):
        O(c, ebitda, f"={c}{gp}-{c}{totopex}", F(11, True), numfmt=USD0, fillc=HILITE, border=top_border)
    rr_box[0] += 1
    pct_line("EBITDA margin %", ebitda, rev)
    da = line("Depreciation & amortization", omap.get("da"), formula=(None if omap.get("da") else (lambda i, c: 0)))
    ebit = rr_box[0]
    O("A", ebit, "   EBIT (operating income)", F(10, True), border=top_border)
    for i, c in enumerate(OC):
        O(c, ebit, f"={c}{ebitda}-{c}{da}", F(10, True), numfmt=USD0, border=top_border)
    rr_box[0] += 1
    inter = line("Interest expense", formula=lambda i, c: link(omap.get("interest"), i))
    # Below-the-line items — each row appears ONLY if its key is present in the output map.
    # bridge terms: Net Income = EBIT − interest − tax + other_income − other_expense
    bridge = [("-", inter)]
    if omap.get("income_taxes"):
        tax_row = line("Income taxes", formula=lambda i, c: link(omap.get("income_taxes"), i))
        bridge.append(("-", tax_row))
    if omap.get("other_income"):
        oinc_row = line("Other income", formula=lambda i, c: link(omap.get("other_income"), i))
        bridge.append(("+", oinc_row))
    if omap.get("other_expense"):
        oexp_row = line("Other expense", formula=lambda i, c: link(omap.get("other_expense"), i))
        bridge.append(("-", oexp_row))
    rr_box[0] += 1
    ni = line("Net Income", omap["net_income"], indent=0)
    for col in ["A"] + OC:
        cc = out[f"{col}{ni}"]; cc.fill = fill(HILITE); cc.font = F(11, True); cc.border = top_border
    out[f"A{ni}"].value = "Net Income"
    pct_line("Net margin %", ni, rev)
    # reconciliation: EBIT ± bridge terms − Net Income must = 0
    rr_box[0] += 1
    O("A", rr_box[0], "   Reconciliation (EBIT − interest − tax + other income − other expense − net income, =0)",
      F(8, italic=True, color="595959"))
    for i, c in enumerate(OC):
        expr = f"{c}{ebit}" + "".join(f"{s}{c}{rrf}" for s, rrf in bridge) + f"-{c}{ni}"
        O(c, rr_box[0], f"=ROUND({expr},2)", F(8, italic=True, color="595959"), numfmt=CHK)
    rr_box[0] += 2

    out.column_dimensions["A"].width = 48
    for c in OC:
        out.column_dimensions[c].width = 15
    out.freeze_panes = f"B{h+1}"
    out.sheet_view.showGridLines = False

    # =====================================================================
    # FOLLOW-UPS — separate tab, so the P&L overview stays clean
    # =====================================================================
    fu = wb.create_sheet("Follow-ups", 1)   # tab order: Output → Follow-ups → Source_Audit
    wrap_top = Alignment(wrap_text=True, vertical="top")
    def FU(row, text, font, height=None, align=None):
        cell = fu.cell(row=row, column=1, value=text)
        cell.font = font
        if align: cell.alignment = align
        if height: fu.row_dimensions[row].height = height
    fr = 1
    FU(fr, f"{m.get('company','')} — Follow-ups & data needed", F(14, True, NAVY)); fr += 1
    FU(fr, "Open items from standardizing this P&L — confirm before relying on the numbers.", F(9, italic=True, color="595959")); fr += 2
    bucket_titles = {
        "classification": "Classification to confirm",
        "functional_split": "Data needed to split costs (S&M / R&D / G&A)",
        "comparability": "Comparability / growth gaps",
    }
    fus = data.get("followups", [])
    def emit_bucket(title, items):
        nonlocal fr
        if not items: return
        FU(fr, title, F(11, True, "404040")); fr += 1
        for f in items:
            FU(fr, "• " + f["text"], F(10, color="404040"),
               height=max(15, 15 * (1 + len(f["text"]) // 110)), align=wrap_top); fr += 1
        fr += 1
    for bkey, btitle in bucket_titles.items():
        emit_bucket(btitle, [f for f in fus if f.get("bucket") == bkey])
    emit_bucket("Other", [f for f in fus if f.get("bucket") not in bucket_titles])
    if not fus:
        FU(fr, "No follow-ups flagged.", F(10, italic=True, color="595959"))
    fu.column_dimensions["A"].width = 110
    fu.sheet_view.showGridLines = False

    wb.save(out_path)
    return {"source_rows": r, "keys": len(rowref), "periods": pids}


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 build_pnl_workbook.py <input.json> <output.xlsx>"); sys.exit(1)
    with open(sys.argv[1]) as fh:
        data = json.load(fh)
    info = build(data, sys.argv[2])
    print("Built", sys.argv[2], "|", info)
