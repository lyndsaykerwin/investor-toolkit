#!/usr/bin/env python3
"""
ARR-to-bookings — build a Quarterly ACV Bookings analysis from ANY customer-revenue file.

Pipeline:
  1. normalize.py turns the source into a canonical grid (customers x continuous period
     axis, clean numbers) + a traceability map. All messy-input handling lives there.
  2. This script writes a CLEAN OUTPUT workbook (source file on disk untouched), ordered:
       - "Quarterly/Annual Bookings": new-logo & upsell ACV, logo counts, largest deal per
                               period (+ customer), YoY %, annual block, partial-year stubs,
                               and a reconciliation.
       - "Normalized" (helper): ONLY when a reshape is needed — simple formulas back to Raw
                               Data (SUMIF / SUMIFS / =MAX(Raw,0)), else cleaned values for
                               non-numeric text. A clean wide source skips it entirely.
       - "Bookings Detail"   : per-customer snapshots =ROUND(<ref>xfactor,2) and the
                               new/upsell/down/churn classification, all formulas.
       - "Raw Data"          : the FINAL tab — a verbatim copy of the source data, no changes.
     Everything links back to Raw Data through auditable formulas.
  3. Self-verifies: recomputes every figure from the canonical grid and (if LibreOffice is
     present) recalculates the workbook and asserts the cached values match. Non-zero exit
     on any mismatch.

Frequency: monthly or quarterly -> quarterly bookings (with an annual summary block).
Annual-only sources -> an "Annual Bookings" tab (yearly new-vs-upsell, YoY vs prior year).
"""
import argparse, datetime, json, os, shutil, subprocess, sys
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter as CL, column_index_from_string as CIX
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import normalize as NZ

EPS = 0.01
MAX_CELLS = 4_000_000  # volume guard on the Detail sheet

# styles
NAVY = PatternFill("solid", fgColor="1F4E79"); MBLUE = PatternFill("solid", fgColor="BDD7EE")
GREENOK = PatternFill("solid", fgColor="C6EFCE"); AMBER = PatternFill("solid", fgColor="FFF2CC")
LGREY = PatternFill("solid", fgColor="F2F2F2")
WB = Font(bold=True, color="FFFFFF"); BOLD = Font(bold=True); GREEN = Font(color="008000")
BLUE = Font(color="0000FF"); ITAL = Font(italic=True, size=9, color="404040")
CTR = Alignment(horizontal="center")
USD = "#,##0;(#,##0)"; NUM = "#,##0"; PCT = "0.0%;(0.0%)"


def _d(s):
    return datetime.date.fromisoformat(s) if isinstance(s, str) else s

def quarter_label(d):
    return f"{d.year}-Q{(d.month - 1) // 3 + 1}"


def make_periods(axis, freq):
    """axis: list[date] continuous. Return (opening_idx, periods, lookback, do_fy, do_stub).
    period = dict(label, idx, prev_idx, date, partial). Monthly -> quarter-ends (+trailing
    partial). Quarterly -> each point (lookback 4). Annual -> each year (lookback 1, no FY
    block, no stubs)."""
    if freq == "A":
        periods = [dict(label=str(axis[i].year), idx=i, prev_idx=i - 1, date=axis[i], partial=False)
                   for i in range(1, len(axis))]
        return 0, periods, 1, False, False
    opening = 0
    periods = []
    prev = 0
    qe = {3, 6, 9, 12}
    for i, d in enumerate(axis):
        if i == 0:
            continue
        is_qe = d.month in qe
        is_last = i == len(axis) - 1
        is_partial_q = is_last and not is_qe
        if freq == "Q":
            take = True; is_partial_q = False
        else:
            take = is_qe or is_partial_q
        if take:
            periods.append(dict(label=quarter_label(d) + ("*" if is_partial_q else ""),
                                idx=i, prev_idx=prev, date=d, partial=is_partial_q))
            prev = i
    do_fy = True
    last = periods[-1]["date"] if periods else None
    do_stub = bool(last) and last.month != 12  # trailing year incomplete
    return opening, periods, 4, do_fy, do_stub


def col_for_axis(k):
    """Normalized/Detail snapshot column for axis index k (Customer in col 1, periods from 2)."""
    return 2 + k


def build(args):
    try:
        can = NZ.normalize_file(args.source, sheet=args.sheet, neg_policy=args.neg_policy,
                                actuals_through=args.actuals_through)
    except ValueError as e:
        sys.exit(f"REFUSE: {e}")
    annual = can["frequency"] == "A"
    BOOK = "Annual Bookings" if annual else "Quarterly Bookings"
    customers = can["customers"]
    axis = [_d(x) for x in can["axis"]]
    matrix = can["matrix"]
    arrf = args.arr_factor if args.arr_factor else can["arr_factor"]
    ncust, nper = len(customers), len(axis)

    opening, periods, lookback, do_fy, do_stub = make_periods(axis, can["frequency"])
    if len(periods) < 1:
        sys.exit("REFUSE: not enough periods to form a single booking quarter.")

    est_cells = ncust * (nper + 4 * (nper - 1) + 4 * len(periods))
    if ncust > args.max_customers or est_cells > MAX_CELLS:
        sys.exit(f"REFUSE (volume): {ncust:,} customers x {nper} periods ~= {est_cells:,} helper "
                 f"cells exceeds the safe limit ({args.max_customers:,} customers / {MAX_CELLS:,} "
                 f"cells). Aggregate/segment first, or raise --max-customers deliberately.")

    out = args.out
    trace = can["trace"]
    orientation = trace.get("orientation")
    all_numeric = bool(trace.get("all_numeric"))
    policy_zero = bool(can["negatives"]) and args.neg_policy == "zero"
    # DIRECT: clean wide numeric, no zeroing -> analysis links straight to Raw Data, no helper.
    # HELPER: a reshape/clean is needed -> a Normalized tab links to Raw Data via SUMIF/SUMIFS
    #         (numeric) or holds cleaned values (when the source values are non-numeric text).
    direct = (orientation == "wide_direct") and all_numeric and not policy_zero
    helper = not direct
    value_helper = helper and not all_numeric        # text values Excel can't sum by formula
    formula_helper = helper and all_numeric           # SUMIF/SUMIFS/MAX back to Raw Data
    if value_helper:
        matrix = [[round(v, 2) for v in row] for row in matrix]; can["matrix"] = matrix

    # ---------------- fresh OUTPUT workbook ----------------
    wb = Workbook(); wb.remove(wb.active)

    # ---------------- RAW DATA (verbatim copy of the source data; placed LAST) ----------------
    RAW = "Raw Data"
    rd = wb.create_sheet(RAW)
    if str(args.source).lower().endswith(".csv"):
        import csv
        with open(args.source, newline="", encoding="utf-8-sig") as fh:
            csv_rows = list(csv.reader(fh))
        for ri_, row in enumerate(csv_rows, start=1):
            for ci_, val in enumerate(row, start=1):
                if val != "":
                    rd.cell(ri_, ci_, val)
        src_last_row = len(csv_rows)
    else:
        srcws = load_workbook(args.source, data_only=True)[can["sheet"]]
        for row in srcws.iter_rows():
            for cc in row:
                if cc.value is not None:
                    nc = rd.cell(cc.row, cc.column, cc.value)
                    try: nc.number_format = cc.number_format
                    except Exception: pass
        for key, dim in srcws.column_dimensions.items():
            if dim.width: rd.column_dimensions[key].width = dim.width
        src_last_row = srcws.max_row

    # source coordinate maps (1-based; identical positions in the Raw Data copy)
    period_src = {_d(ds): ci for ds, ci in trace.get("period_src", [])}
    src_rows = trace.get("src_rows", {})
    HR = 4; DF = HR + 1; NL = DF + ncust - 1

    # ---------------- NORMALIZED helper (only when a reshape/clean is needed) ----------------
    if helper:
        nz = wb.create_sheet("Normalized")
        nz["A1"] = "Normalized — canonical customer × period grid (links to Raw Data)"; nz["A1"].font = BOLD
        how = "cleaned values (source values were non-numeric text)" if value_helper else \
              ("SUMIF aggregation of Raw Data" if orientation == "wide_multirow" else
               ("SUMIFS aggregation of Raw Data" if orientation in ("long", "long_csv") else
                "=MAX(Raw Data, 0) (negative policy=zero)"))
        nz["A2"] = (f"Built from 'Raw Data' | orientation {orientation} | freq {can['frequency']} | "
                    f"scale {can['scale']} | {how}. Negatives {len(can['negatives'])} "
                    f"(policy {args.neg_policy}); gap-filled {can['gaps_filled']}; dropped rows {len(can['dropped'])}.")
        nz["A2"].font = ITAL
        nz.cell(HR, 1, "Customer").font = BOLD
        for k, d in enumerate(axis):
            c = nz.cell(HR, col_for_axis(k), datetime.datetime(d.year, d.month, d.day))
            c.number_format = "mmm-yy"; c.font = BOLD; c.alignment = CTR
        # ranges for aggregation formulas
        if formula_helper and orientation == "wide_multirow":
            allrows = [r for rows in src_rows.values() for r in rows]
            c0, c1 = min(allrows), max(allrows)
            custcol = CL(trace["customer_col"]); custrng = f"'{RAW}'!${custcol}${c0}:${custcol}${c1}"
        if formula_helper and orientation in ("long", "long_csv"):
            r0 = trace["header_row"] + 1; r1 = src_last_row
            dcol = CL(trace["date_col"]); ccol = CL(trace["customer_col"]); vcol = CL(trace["value_col"])
            crng = f"'{RAW}'!${ccol}${r0}:${ccol}${r1}"; drng = f"'{RAW}'!${dcol}${r0}:${dcol}${r1}"
            vrng = f"'{RAW}'!${vcol}${r0}:${vcol}${r1}"
        for ri, name in enumerate(customers):
            rr = DF + ri; nz.cell(rr, 1, name)
            nm = str(name).replace('"', '""')
            for k, d in enumerate(axis):
                cell = nz.cell(rr, col_for_axis(k)); cell.number_format = USD
                if value_helper:
                    v = matrix[ri][k]; cell.value = v if v else 0; continue
                base = None
                if orientation == "wide_multirow":
                    if d in period_src:
                        pcol = CL(period_src[d]); prng = f"'{RAW}'!${pcol}${c0}:${pcol}${c1}"
                        base = f'SUMIF({custrng},"{nm}",{prng})'
                elif orientation in ("long", "long_csv"):
                    start = datetime.date(d.year, d.month, 1)
                    base = (f'SUMIFS({vrng},{crng},"{nm}",{drng},">="&DATE({start.year},{start.month},{start.day}),'
                            f'{drng},"<="&DATE({d.year},{d.month},{d.day}))')
                else:  # wide_direct numeric, policy=zero
                    if d in period_src and name in src_rows and src_rows[name]:
                        base = f"'{RAW}'!{CL(period_src[d])}{src_rows[name][0]}"
                if base is None:
                    cell.value = 0; continue
                cell.value = f"=MAX({base},0)" if policy_zero else f"={base}"
                cell.font = GREEN
        trow = NL + 1
        nz.cell(trow, 1, "Period total").font = BOLD
        for k in range(nper):
            L = CL(col_for_axis(k)); c = nz.cell(trow, col_for_axis(k))
            c.value = f"=SUM({L}{DF}:{L}{NL})"; c.number_format = USD; c.font = BOLD; c.fill = LGREY
        nz.freeze_panes = "B5"; nz.column_dimensions["A"].width = 30

    # ---------------- reference builders (used by Detail + reconciliation) ----------------
    def snap_ref(ri, k):
        """formula body (no '=') for customer ri's source revenue at axis period k."""
        if helper:
            return f"Normalized!{CL(col_for_axis(k))}{DF + ri}"
        d = axis[k]; name = customers[ri]
        if d in period_src and name in src_rows and src_rows[name]:
            return f"'{RAW}'!{CL(period_src[d])}{src_rows[name][0]}"
        return "0"

    def name_ref(ri):
        if helper:
            return f"Normalized!A{DF + ri}"
        name = customers[ri]
        if name in src_rows and src_rows[name]:
            return f"'{RAW}'!{CL(trace['customer_col'])}{src_rows[name][0]}"
        return None

    def period_total(k):
        """ARR total of axis period k = sum of the (cents-rounded) Detail snapshots, which
        themselves are simple formulas back to Raw Data / Normalized."""
        L = CL(SNAP0 + k); return f"SUM('Bookings Detail'!{L}{DDF}:{L}{DDL})"

    # ---------------- DETAIL ----------------
    NQ = len(periods)
    d = wb.create_sheet("Bookings Detail")
    d["A1"] = ("Bookings Detail — per-customer per-MONTH classification (mNew/mUp/mDn/mCh = each "
               "month vs prior), then quarter columns = SUM of the months each quarter owns"); d["A1"].font = BOLD
    d.cell(3, 1, "Customer").font = BOLD
    SNAP0 = 2
    for k, dt in enumerate(axis):
        d.cell(3, SNAP0 + k, dt.strftime("%b-%y")).font = Font(bold=True, size=7)
    # MONTHLY bookings math (each period vs the immediately prior period), then quarter rollups =
    # SUM of the months a quarter owns. Period 0 (first in the dataset) is the opening base — there
    # is no prior period to compare it to, so it books nothing. Every comparison is a uniform
    # one-step delta, and any quarter (including the first) is simply the sum of its months. This
    # measures GROSS bookings: a sign-and-churn or up-then-down inside one quarter is captured,
    # not netted away as it would be by an end-to-end quarter snapshot comparison.
    NM = nper - 1                                     # per-step comparison columns (k = 1..nper-1)
    mNew0 = SNAP0 + nper; mUp0 = mNew0 + NM; mDn0 = mUp0 + NM; mCh0 = mDn0 + NM
    new0 = mCh0 + NM; up0 = new0 + NQ; dn0 = up0 + NQ; ch0 = dn0 + NQ

    def qmonths(prev_i, cur_i):
        """axis step indices a period owns: prior period-end+1 .. this period-end (each vs k-1)."""
        return range(prev_i + 1, cur_i + 1)

    def mrange(blk0, ks, dr):
        """Detail row range over a monthly block for (contiguous) step indices ks."""
        ks = list(ks); return f"{CL(blk0 + ks[0] - 1)}{dr}:{CL(blk0 + ks[-1] - 1)}{dr}"

    for blk, t in [(mNew0, "mNew"), (mUp0, "mUp"), (mDn0, "mDn"), (mCh0, "mCh")]:
        for k in range(1, nper):
            d.cell(3, blk + (k - 1), f"{t} {axis[k].strftime('%b-%y')}").font = Font(bold=True, size=7)
    for blk, t in [(new0, "New"), (up0, "Up"), (dn0, "Dn"), (ch0, "Ch")]:
        for qi, p in enumerate(periods):
            d.cell(3, blk + qi, f"{t} {p['label']}").font = Font(bold=True, size=7)
    DDF = 4
    AR = f"'{BOOK}'!$B$2"
    for ri, name in enumerate(customers):
        dr = DDF + ri
        nr = name_ref(ri)
        if nr: d.cell(dr, 1, f"={nr}").font = GREEN
        else: d.cell(dr, 1, customers[ri])
        for k in range(nper):
            d.cell(dr, SNAP0 + k, f"=ROUND({snap_ref(ri, k)}*{AR},2)").font = GREEN
        # per-step classification: period k vs period k-1
        for k in range(1, nper):
            prev = f"{CL(SNAP0 + k - 1)}{dr}"; cur = f"{CL(SNAP0 + k)}{dr}"; o = k - 1
            # The 4 buckets must EXACTLY partition the step delta (cur-prev) for ALL sign
            # combinations of prev/cur, so the reconciliation telescopes to $0 under every
            # neg-policy (not just `zero`). Active = >EPS, inactive = <=EPS (covers negative
            # credits left in by policy=leave/contraction). New = state flips inactive->active;
            # Churn = active->inactive; both record the FULL delta (not just cur / -prev) so no
            # credit residual leaks. Up/Dn = active-state unchanged (both active OR both inactive),
            # split by the sign of the delta. The four conditions are mutually exclusive and the
            # firing bucket's amount always equals cur-prev.
            same = f"OR(AND({prev}>{EPS},{cur}>{EPS}),AND({prev}<={EPS},{cur}<={EPS}))"
            d.cell(dr, mNew0 + o, f"=IF(AND({prev}<={EPS},{cur}>{EPS}),{cur}-{prev},0)")
            d.cell(dr, mUp0 + o, f"=IF(AND({same},{cur}-{prev}>{EPS}),{cur}-{prev},0)")
            d.cell(dr, mDn0 + o, f"=IF(AND({same},{cur}-{prev}<-{EPS}),{cur}-{prev},0)")
            d.cell(dr, mCh0 + o, f"=IF(AND({prev}>{EPS},{cur}<={EPS}),{cur}-{prev},0)")
        # quarter rollups = SUM of the per-step columns the quarter owns
        for qi, p in enumerate(periods):
            ks = qmonths(p["prev_idx"], p["idx"])
            d.cell(dr, new0 + qi, f"=SUM({mrange(mNew0, ks, dr)})")
            d.cell(dr, up0 + qi, f"=SUM({mrange(mUp0, ks, dr)})")
            d.cell(dr, dn0 + qi, f"=SUM({mrange(mDn0, ks, dr)})")
            d.cell(dr, ch0 + qi, f"=SUM({mrange(mCh0, ks, dr)})")
    DDL = DDF + ncust - 1
    d.freeze_panes = "B4"; d.column_dimensions["A"].width = 26  # visible — it's audit work, not hidden machinery

    def RNG(blk, qi): L = CL(blk + qi); return f"'Bookings Detail'!{L}{DDF}:{L}{DDL}"
    NAMES = f"'Bookings Detail'!$A${DDF}:$A${DDL}"

    # ---------------- SUMMARY ----------------
    s = wb.create_sheet(BOOK, 0)
    s["A1"] = (("Annual" if annual else "Quarterly") + " ACV Bookings — with YoY growth"
               + ("" if annual else " & annual summaries")); s["A1"].font = Font(bold=True, size=14)
    s["A2"] = "ARR factor (×):"; s["A2"].font = ITAL
    s["B2"] = arrf; s["B2"].font = BLUE
    if annual:
        s["A3"] = ("Annual bookings: source has annual columns, so each period is a full YEAR and "
                   "values are annual amounts (factor 1). New-logo = customer's first active year; "
                   "Upsell = increase vs prior year. YoY compares to the prior year; first year is the "
                   "opening base. Built from the Normalized tab (see its note).")
    else:
        s["A3"] = ("ACV = annual contract value (period run-rate × factor). Bookings are measured "
                   "period-over-period (each month vs the prior month) and SUMMED into the quarter: "
                   "New-logo = a customer's first active month; Upsell = a month-over-month ARR increase "
                   "for an existing customer (gross — within-quarter rises are not netted against later "
                   "dips). The first month of the dataset is the opening base and books nothing. YoY "
                   "compares to the same quarter a year earlier; blank = no comparable prior. * = partial "
                   "trailing period. See the Bookings Detail tab for the per-month math.")
    s["A3"].font = ITAL
    s["A4"] = ("ESTIMATE — derived from customer-level MRR/ARR, the only data available. Not valid for "
               "usage-based models; a CRM-sourced bookings read, where available, is the most accurate "
               "measure of bookings growth.")
    s["A4"].font = Font(bold=True, italic=True, size=9, color="7F6000")
    HDR = 5; QC0 = 2
    GAP = QC0 + NQ + 1
    # full years (only those with all 4 quarters present)
    yq = {}
    for qi, p in enumerate(periods):
        if not p["partial"]:
            yq.setdefault(p["date"].year, []).append(qi)
    fy_years = sorted(y for y, qs in yq.items() if len(qs) == 4) if do_fy else []
    AC0 = GAP + 1
    fy_cols = {y: AC0 + i for i, y in enumerate(fy_years)}
    GAP2 = AC0 + len(fy_years)

    # stub periods (apples-to-apples) when trailing year incomplete
    stubs = []  # list of dict(label,new_range_idxs...) computed from axis snapshots
    if do_stub and periods:
        last = periods[-1]["date"]; ly, lm = last.year, last.month
        axis_idx = {dt: i for i, dt in enumerate(axis)}
        def dec(y):
            return axis_idx.get(NZ._eom(y, 12))
        def samem(y):
            return axis_idx.get(NZ._eom(y, lm))
        for yr in (ly - 1, ly):
            prev_i = dec(yr - 1); cur_i = samem(yr)
            if prev_i is not None and cur_i is not None:
                stubs.append(dict(label=f"{yr} Jan–{last.strftime('%b')}", prev_i=prev_i, cur_i=cur_i))
    SA = GAP2 + 1 if len(stubs) >= 1 else None
    SB = GAP2 + 2 if len(stubs) >= 2 else (GAP2 + 1 if len(stubs) == 1 else None)
    stub_cols = {}
    if len(stubs) == 1: stub_cols = {0: SA}
    if len(stubs) >= 2: stub_cols = {0: SA, 1: SB}
    ann_cols = [fy_cols[y] for y in fy_years] + [stub_cols[i] for i in range(len(stubs))]

    # detail stub classification columns (append to Detail)
    if stubs:
        base = ch0 + NQ
        for si, st in enumerate(stubs):
            ncol = base + si * 2; ucol = base + si * 2 + 1
            st["new_col"] = ncol; st["up_col"] = ucol
            d.cell(3, ncol, f"New {st['label']}").font = Font(bold=True, size=7)
            d.cell(3, ucol, f"Up {st['label']}").font = Font(bold=True, size=7)
            ks = qmonths(st["prev_i"], st["cur_i"])
            for ri in range(ncust):
                dr = DDF + ri
                d.cell(dr, ncol, f"=SUM({mrange(mNew0, ks, dr)})")
                d.cell(dr, ucol, f"=SUM({mrange(mUp0, ks, dr)})")

    def stub_rng(si, which):
        col = stubs[si]["new_col"] if which == "new" else stubs[si]["up_col"]
        L = CL(col); return f"'Bookings Detail'!{L}{DDF}:{L}{DDL}"

    # headers
    s.cell(HDR, 1, "($ = ACV; % = YoY)").font = BOLD
    for qi, p in enumerate(periods):
        c = s.cell(HDR, QC0 + qi, p["label"]); c.font = WB; c.fill = NAVY; c.alignment = CTR
    for y in fy_years:
        c = s.cell(HDR, fy_cols[y], f"{y} FY"); c.font = WB; c.fill = NAVY; c.alignment = CTR
    for i in range(len(stubs)):
        c = s.cell(HDR, stub_cols[i], stubs[i]["label"]); c.fill = AMBER
        c.font = Font(bold=True, color="7F6000"); c.alignment = CTR
    if fy_years:
        s.cell(HDR - 1, AC0, "ANNUAL SUMMARY").font = Font(bold=True, size=11, color="1F4E79")
    if stubs:
        s.cell(HDR - 1, stub_cols[0], "APPLES-TO-APPLES").font = Font(bold=True, size=10, color="7F6000")

    last_col = (stub_cols[len(stubs) - 1] if stubs else (fy_cols[fy_years[-1]] if fy_years else QC0 + NQ - 1))
    def banner(r, t):
        s.cell(r, 1, t).font = WB; s.cell(r, 1).fill = NAVY
        for c in range(QC0, last_col + 1):
            s.cell(r, c).fill = NAVY

    R = {}
    r = HDR + 1
    s.cell(r, 1, "Opening installed base — ARR at first period").font = Font(italic=True)
    c = s.cell(r, QC0); c.value = f"={period_total(opening)}"; c.number_format = USD; c.font = GREEN
    r += 2

    def fy_formula(row, y):
        qs = yq[y]; a = CL(QC0 + qs[0]); b = CL(QC0 + qs[-1]); return f"=SUM({a}{row}:{b}{row})"

    def yoy_row(rr, drow):
        for qi in range(NQ):
            if qi < lookback: continue
            cur = f"{CL(QC0 + qi)}{drow}"; pri = f"{CL(QC0 + qi - lookback)}{drow}"
            cc = s.cell(rr, QC0 + qi); cc.value = f'=IF({pri}<=0,"",{cur}/{pri}-1)'
            cc.number_format = PCT; cc.font = ITAL
        for i, y in enumerate(fy_years):
            if i == 0: continue
            cur = f"{CL(fy_cols[y])}{drow}"; pri = f"{CL(fy_cols[fy_years[i-1]])}{drow}"
            cc = s.cell(rr, fy_cols[y]); cc.value = f'=IF({pri}<=0,"",{cur}/{pri}-1)'
            cc.number_format = PCT; cc.font = ITAL
        if len(stubs) >= 2:
            cur = f"{CL(stub_cols[1])}{drow}"; pri = f"{CL(stub_cols[0])}{drow}"
            cc = s.cell(rr, stub_cols[1]); cc.value = f'=IF({pri}<=0,"",{cur}/{pri}-1)'
            cc.number_format = PCT; cc.font = Font(italic=True, bold=True, color="7F6000")

    def block(title, det_blk, stub_key, dollar_label, count_label, deal_label):
        nonlocal r
        banner(r, title); r += 1
        rowS = r; s.cell(r, 1, dollar_label).font = BOLD
        for qi in range(NQ):
            cc = s.cell(r, QC0 + qi); cc.value = f"=SUM({RNG(det_blk, qi)})"; cc.number_format = USD
        for y in fy_years:
            cc = s.cell(r, fy_cols[y]); cc.value = fy_formula(r, y); cc.number_format = USD; cc.font = BOLD
        for i in range(len(stubs)):
            cc = s.cell(r, stub_cols[i]); cc.value = f"=SUM({stub_rng(i, stub_key)})"; cc.number_format = USD; cc.font = BOLD
        r += 1; s.cell(r, 1, f"   {dollar_label.split(' (')[0]} — YoY %").font = ITAL; yoy_row(r, rowS); r += 1
        rN = r; s.cell(r, 1, f"   {count_label}")
        for qi in range(NQ):
            cc = s.cell(r, QC0 + qi); cc.value = f'=COUNTIF({RNG(det_blk, qi)},">0")'; cc.number_format = NUM
        for y in fy_years: s.cell(r, fy_cols[y], fy_formula(r, y)).number_format = NUM
        for i in range(len(stubs)): s.cell(r, stub_cols[i], f'=COUNTIF({stub_rng(i, stub_key)},">0")').number_format = NUM
        r += 1; rMax = r; s.cell(r, 1, f"   {deal_label} ($)").font = ITAL
        for qi in range(NQ):
            cc = s.cell(r, QC0 + qi); cc.value = f"=MAX({RNG(det_blk, qi)})"; cc.number_format = USD; cc.font = ITAL
        for y in fy_years:
            qs = yq[y]; a = CL(QC0 + qs[0]); b = CL(QC0 + qs[-1])
            s.cell(r, fy_cols[y], f"=MAX({a}{r}:{b}{r})").number_format = USD; s.cell(r, fy_cols[y]).font = ITAL
        for i in range(len(stubs)): s.cell(r, stub_cols[i], f"=MAX({stub_rng(i, stub_key)})").number_format = USD; s.cell(r, stub_cols[i]).font = ITAL
        r += 1; rWho = r; s.cell(r, 1, f"   {deal_label} (customer)").font = ITAL
        for qi in range(NQ):
            mx = f"{CL(QC0 + qi)}{rMax}"
            s.cell(r, QC0 + qi, f'=IF({mx}=0,"—",INDEX({NAMES},MATCH({mx},{RNG(det_blk, qi)},0)))').font = Font(italic=True, size=8)
        for y in fy_years:
            qs = yq[y]; a = CL(QC0 + qs[0]); b = CL(QC0 + qs[-1]); mx = f"{CL(fy_cols[y])}{rMax}"
            s.cell(r, fy_cols[y], f'=IF({mx}=0,"—",INDEX({a}{rWho}:{b}{rWho},MATCH({mx},{a}{rMax}:{b}{rMax},0)))').font = Font(italic=True, size=8)
        for i in range(len(stubs)):
            mx = f"{CL(stub_cols[i])}{rMax}"
            s.cell(r, stub_cols[i], f'=IF({mx}=0,"—",INDEX({NAMES},MATCH({mx},{stub_rng(i, stub_key)},0)))').font = Font(italic=True, size=8)
        r += 2
        return dict(S=rowS, N=rN, Max=rMax)

    R["new"] = block("NEW BUSINESS (new logos)", new0, "new", "New-logo ACV ($)", "New logos (#)", "Largest new deal")
    R["up"] = block("EXPANSION (upsell to existing customers)", up0, "up", "Upsell ACV ($)", "Accounts upsold (# / annual = events)", "Largest upsell")

    banner(r, "TOTAL BOOKINGS"); r += 1
    rTotS = r; s.cell(r, 1, "Total ACV bookings ($)").font = BOLD
    allcols = list(range(QC0, QC0 + NQ)) + ann_cols
    for col in allcols:
        cc = s.cell(r, col); cc.value = f"={CL(col)}{R['new']['S']}+{CL(col)}{R['up']['S']}"
        cc.number_format = USD; cc.font = BOLD; cc.fill = MBLUE
    r += 1; s.cell(r, 1, "   Total bookings — YoY %").font = ITAL; yoy_row(r, rTotS); r += 1
    s.cell(r, 1, "   Total accounts booked (#)")
    for col in allcols:
        cc = s.cell(r, col); cc.value = f"={CL(col)}{R['new']['N']}+{CL(col)}{R['up']['N']}"; cc.number_format = NUM
    r += 2

    banner(r, "RECONCILIATION  (Beginning + New + Upsell + Downsell + Churn = Ending ARR, vs Normalized)"); r += 1
    rBeg = r; s.cell(r, 1, "Beginning ARR (prior period)")
    for qi, p in enumerate(periods):
        s.cell(r, QC0 + qi, f"={period_total(p['prev_idx'])}").number_format = USD; s.cell(r, QC0 + qi).font = GREEN
    r += 1; rRN = r; s.cell(r, 1, "   + New-logo ACV")
    for qi in range(NQ): s.cell(r, QC0 + qi, f"={CL(QC0 + qi)}{R['new']['S']}").number_format = USD
    r += 1; rRU = r; s.cell(r, 1, "   + Upsell ACV")
    for qi in range(NQ): s.cell(r, QC0 + qi, f"={CL(QC0 + qi)}{R['up']['S']}").number_format = USD
    r += 1; rRD = r; s.cell(r, 1, "   − Downsell ACV")
    for qi in range(NQ): s.cell(r, QC0 + qi, f"=SUM({RNG(dn0, qi)})").number_format = USD
    r += 1; rRC = r; s.cell(r, 1, "   − Churn ACV")
    for qi in range(NQ): s.cell(r, QC0 + qi, f"=SUM({RNG(ch0, qi)})").number_format = USD
    r += 1; rEC = r; s["A" + str(r)] = " = Ending ARR (computed)"; s.cell(r, 1).font = BOLD
    for qi in range(NQ):
        cc = CL(QC0 + qi)
        s.cell(r, QC0 + qi, f"={cc}{rBeg}+{cc}{rRN}+{cc}{rRU}+{cc}{rRD}+{cc}{rRC}").number_format = USD; s.cell(r, QC0 + qi).font = BOLD
    r += 1; rER = r; s.cell(r, 1, "Ending ARR (sum of customer snapshots → Raw Data)")
    for qi, p in enumerate(periods):
        s.cell(r, QC0 + qi, f"={period_total(p['idx'])}").number_format = USD; s.cell(r, QC0 + qi).font = GREEN
    r += 1; rV = r; s.cell(r, 1, "External check (computed − ending) = 0").font = BOLD
    for qi in range(NQ):
        cc = CL(QC0 + qi); x = s.cell(r, QC0 + qi); x.value = f"={cc}{rEC}-{cc}{rER}"
        x.number_format = USD; x.font = BOLD; x.fill = GREENOK

    s.freeze_panes = "B6"; s.column_dimensions["A"].width = 40
    for qi in range(NQ): s.column_dimensions[CL(QC0 + qi)].width = 11
    for col in ann_cols: s.column_dimensions[CL(col)].width = 12

    # sheet order: analysis → helper (if any) → hidden detail → Raw Data LAST
    order = [BOOK] + (["Normalized"] if helper else []) + ["Bookings Detail", RAW]
    wb._sheets.sort(key=lambda ws: order.index(ws.title))
    wb.save(out)
    return dict(out=out, can=can, book_sheet=BOOK, periods=periods, lookback=lookback, fy_years=fy_years,
                stubs=stubs, rows=dict(newS=R["new"]["S"], upS=R["up"]["S"], totS=rTotS, var=rV,
                                       newN=R["new"]["N"], upN=R["up"]["N"], newMax=R["new"]["Max"], upMax=R["up"]["Max"]),
                cols=dict(QC0=QC0, NQ=NQ, fy_cols=fy_cols, stub_cols=stub_cols), arrf=arrf)


# ---------------- verify ----------------
def classify_step(matrix, k, arrf):
    """Per-customer (new, upsell) for one step: period k vs k-1, matching the Detail's ROUND."""
    out = []
    for row in matrix:
        p = round(row[k - 1] * arrf, 2); c = round(row[k] * arrf, 2)
        n = u = 0.0
        # mirror the Detail formulas: New = inactive(<=EPS)->active(>EPS), recorded as the FULL
        # delta (c-p) so a negative prev-credit doesn't leak; Upsell = active-state unchanged
        # (both active OR both inactive) with a positive delta.
        if p <= EPS and c > EPS: n = c - p
        elif ((p > EPS) == (c > EPS)) and c - p > EPS: u = c - p
        out.append((n, u))
    return out


def classify_axis(matrix, prev_i, cur_i, arrf):
    """Period bookings = per-customer SUM of per-step new/upsell over the months it owns
    (prev_i+1..cur_i), then aggregated across customers — mirrors the Detail rollup formulas
    (gross, captures within-period sign-and-reverse activity)."""
    nc = len(matrix); qn = [0.0] * nc; qu = [0.0] * nc
    for k in range(prev_i + 1, cur_i + 1):
        for i, (n, u) in enumerate(classify_step(matrix, k, arrf)):
            qn[i] += n; qu[i] += u
    return dict(new=sum(qn), up=sum(qu),
                nn=sum(1 for x in qn if x > EPS), nu=sum(1 for x in qu if x > EPS),
                mn=max(qn) if qn else 0.0, mu=max(qu) if qu else 0.0)


def verify(meta):
    can = meta["can"]; arrf = meta["arrf"]; matrix = can["matrix"]
    periods = meta["periods"]
    truth = [classify_axis(matrix, p["prev_idx"], p["idx"], arrf) for p in periods]
    soffice = None
    for c in ("soffice", "/Applications/LibreOffice.app/Contents/MacOS/soffice"):
        if shutil.which(c) or os.path.exists(c):
            soffice = c; break
    if not soffice:
        print("WARN: LibreOffice not found; wrote formulas but did not numerically confirm.")
        return True
    outdir = os.path.join(os.path.dirname(os.path.abspath(meta["out"])), "_recalc")
    os.makedirs(outdir, exist_ok=True)
    subprocess.run([soffice, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", outdir, meta["out"]],
                   check=True, capture_output=True)
    rc = load_workbook(os.path.join(outdir, os.path.basename(meta["out"])), data_only=True)
    book = meta["book_sheet"]
    s = rc[book]; R = meta["rows"]; C = meta["cols"]; QC0 = C["QC0"]; NQ = C["NQ"]
    g = lambda r, c: s.cell(r, c).value
    errs = 0
    for qi in range(NQ):
        for k, row in [("new", R["newS"]), ("up", R["upS"]), ("nn", R["newN"]), ("nu", R["upN"]),
                       ("mn", R["newMax"]), ("mu", R["upMax"])]:
            if abs((g(row, QC0 + qi) or 0) - truth[qi][k]) > 0.5:
                errs += 1; print(f"  MISMATCH q{qi} {k}: got {g(row,QC0+qi)} exp {truth[qi][k]}")
    maxvar = max(abs(g(R["var"], QC0 + qi) or 0) for qi in range(NQ))
    # only scan the sheets WE create — source sheets may carry their own pre-existing errors
    ours = [t for t in (book, "Bookings Detail", "Normalized") if t in rc.sheetnames]
    bad = [(t, c.coordinate) for t in ours for row in rc[t].iter_rows()
           for c in row if isinstance(c.value, str) and c.value.startswith("#") and c.value[-1] in "!?"]
    print(f"VERIFY: mismatches={errs}  max reconciliation variance=${maxvar:.4f}  error cells={len(bad)}")
    result = errs == 0 and maxvar < 1.0 and len(bad) == 0
    # On success, deliver the RECALCULATED copy (LibreOffice caches the computed values, so the file
    # opens showing numbers instead of blank formula cells) and remove the temp recalc dir so it
    # never litters the output folder.
    recalced = os.path.join(outdir, os.path.basename(meta["out"]))
    try:
        if result and os.path.exists(recalced):
            shutil.copyfile(recalced, meta["out"])
    finally:
        shutil.rmtree(outdir, ignore_errors=True)
    return result


def main():
    ap = argparse.ArgumentParser(description="Build a Quarterly ACV Bookings analysis from any customer-revenue file.")
    ap.add_argument("--source", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--sheet", default=None, help="force a source sheet (default: auto-detect best)")
    ap.add_argument("--neg-policy", default="leave", choices=["leave", "zero", "contraction"])
    ap.add_argument("--actuals-through", default=None, help="drop periods after this (YYYY-MM or 'May-26')")
    ap.add_argument("--arr-factor", type=float, default=None, help="override the inferred MRR→ARR factor")
    ap.add_argument("--max-customers", type=int, default=5000)
    ap.add_argument("--no-verify", action="store_true")
    args = ap.parse_args()
    meta = build(args)
    can = meta["can"]
    out_abs = os.path.abspath(meta["out"])
    # Clickable link: absolute path + a file:// URL (URL-encoded so spaces in the
    # path stay clickable). Surface this so the caller can hand the user a link.
    from urllib.parse import quote
    out_url = "file://" + quote(out_abs)
    print(f"Built {out_abs}")
    print(f"  link: {out_url}")
    print(f"  source: sheet '{can['sheet']}' / {can['trace'].get('orientation')} / {can['frequency']} / {can['scale']}")
    print(f"  customers: {can['n_customers']}  periods(booking): {len(meta['periods'])}  "
          f"axis: {can['axis'][0]}→{can['axis'][-1]}")
    print(f"  full years: {meta['fy_years']}  stubs: {[s['label'] for s in meta['stubs']]}")
    if can["negatives"]:
        print(f"  negatives: {len(can['negatives'])} (policy={args.neg_policy})")
    if not args.no_verify:
        if not verify(meta):
            sys.exit("FAILED verification")
        print("PASSED verification")


if __name__ == "__main__":
    main()
