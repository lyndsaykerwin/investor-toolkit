#!/usr/bin/env python3
"""
deliver.py — retention-analysis Phase 5 (deliverable).

Reads compute.py JSON + the long-format CSV and writes a three-sheet
formula-driven Excel workbook.

Three sheet-layout modes:

* aggregating  — source has many rows per customer (one per product line / type).
                 Helper sheet = "Raw Data with Analysis" with SUMIFS aggregating by
                 customer + type filter. Includes the self-validation block at
                 the top (active counts, type-decomposed MRR totals, recon
                 against Raw Data direct sums).
* passthrough  — source already has one row per customer. Helper sheet =
                 "Raw Data with Analysis" with live 1:1 refs back to Raw Data.
* twotab       — no source workbook; Raw Data tab is built from the CSV.

Sheet 3 ("Raw Data") is ALWAYS a verbatim copy of the source workbook in
aggregating/passthrough modes — zero edits, zero reformatting, no color
changes. The skill's reconciliation guarantee depends on this.

Formatting follows the skill's finance convention:
  * Blue font (#0000FF)  — hardcoded inputs (ARR factor, methodology label values)
  * Green font (#006100) — formulas that reference another sheet
  * Black font           — formulas internal to the current sheet
  * $ symbol only on the top + bottom row of a vertical numeric block;
    interior rows use #,##0 with no $
  * Title fill #1F4E79, sub-header fill #D9E1F2, no merged cells (uses
    centerContinuous alignment)

CLI:
    python3 deliver.py <compute-output.json> <long-format-csv> <output.xlsx> \\
        [--company "<name>"] \\
        [--source <source.xlsx> --source-sheet "<sheet>" \\
         --source-customer-col <letter> --source-first-data-row <n> \\
         --source-first-date-col <letter>] \\
        [--source-type-col <letter> --type-filter "Recurring,Re-occurring"] \\
        [--lookback 12]

With --source-type-col, the aggregating mode is selected automatically.
With --source but no --source-type-col, passthrough mode is used.
With neither, two-tab fallback is used.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import sys
from copy import copy as _copy
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


# ---------------------------------------------------------------------------
# Layout constants (Corkscrew)
# ---------------------------------------------------------------------------

ROW_TITLE = 1
ROW_GENERATED = 2
ROW_ARR_FACTOR = 3
ROW_DATES = 5
ROW_VS = 6  # optional "vs prior year" label row

# Rollforward block
ROW_BEGIN = 8
ROW_NEW = 9
ROW_UPSELL = 10
ROW_DOWNSELL = 11
ROW_CHURN = 12
ROW_END = 13
ROW_CHECK = 14

# Customer count block
ROW_CC_BANNER = 16
ROW_N_ACTIVE_PRIOR = 17
ROW_N_ACTIVE_CURR = 18
ROW_N_CHURNED = 19
ROW_N_NEW = 20

# Retention metrics
ROW_RR_BANNER = 22
ROW_GRR = 23
ROW_NRR = 24
ROW_LOGO = 25

# Per-customer metrics
ROW_PC_BANNER = 27
ROW_AVG_ARR = 28
ROW_AVG_NEW = 29

# Decomposed reconciliation (only when multi-type scope)
ROW_RECON_BANNER = 31
ROW_REC_RECURRING = 32
ROW_REC_REOCCURRING = 33
ROW_REC_SUM = 35
ROW_REC_VARIANCE = 36

COL_LABEL = 1
FIRST_DATA_COL = 2  # column B = first date column


# ---------------------------------------------------------------------------
# Layout constants (Raw Data with Analysis / helper)
# ---------------------------------------------------------------------------

ANALYSIS_ROW_HDR = 1
ANALYSIS_ROW_ACTIVE = 2
ANALYSIS_ROW_RETAINED = 3
ANALYSIS_ROW_CHECK_ACTIVE = 4
# row 5 blank divider
ANALYSIS_ROW_REC = 6
ANALYSIS_ROW_REOCC = 7
ANALYSIS_ROW_NONREC = 8
ANALYSIS_ROW_TOTAL = 9
ANALYSIS_ROW_CHECK_TOTAL = 10
ANALYSIS_ROW_CHECK_INSCOPE = 11
ANALYSIS_FIRST_CUST_ROW = 12

ANALYSIS_LABEL_COL = 1
ANALYSIS_FIRST_MONTH_COL = 2  # column B


# ---------------------------------------------------------------------------
# Colors & formats
# ---------------------------------------------------------------------------

TITLE_FILL = "1F4E79"      # dark blue, white bold text
BANNER_FILL = "1F4E79"     # section banners use same dark blue
SUBHEADER_FILL = "D9E1F2"  # light blue, black bold
KEY_METRIC_FILL = "BDD7EE" # medium blue for output rows (Ending, retention rates)

COLOR_BLUE = "0000FF"      # hardcoded inputs
COLOR_GREEN = "006100"     # cross-sheet references
COLOR_BLACK = "000000"     # formulas internal to current sheet
COLOR_WHITE = "FFFFFF"

# Number formats. Top-and-bottom-of-block rows get $; interior rows do not.
FMT_DOLLAR = '"$"#,##0;("$"#,##0);"-"'
FMT_NUMBER = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_COUNT = '#,##0;(#,##0);"-"'
FMT_DATE = "mmm-yy"


def font_hardcode(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=COLOR_BLUE)


def font_xsheet(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=COLOR_GREEN)


def font_formula(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=COLOR_BLACK)


def font_title() -> Font:
    return Font(name="Calibri", size=14, bold=True, color=COLOR_WHITE)


def font_banner() -> Font:
    return Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)


def font_subheader() -> Font:
    return Font(name="Calibri", size=10, bold=True, color=COLOR_BLACK)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def center_continuous_across(ws, row: int, first_col: int, last_col: int,
                             text: str, font_obj: Font, fill_obj: PatternFill) -> None:
    """Write `text` to the leftmost cell and apply centerContinuous alignment
    across all cells in the span. Avoids merge_cells, which breaks
    selection/sort/filter/copy-paste."""
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.value = text if c == first_col else None
        cell.font = font_obj
        cell.fill = fill_obj
        cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")


# ---------------------------------------------------------------------------
# Input loading
# ---------------------------------------------------------------------------


def load_compute_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def load_long_csv(path: str) -> Tuple[List[str], List[str], Dict[Tuple[str, str], float]]:
    customers: set = set()
    months: set = set()
    cell: Dict[Tuple[str, str], float] = {}
    with open(path, "r", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            cust = str(row["customer_id"]).strip()
            month_raw = str(row["month"]).strip()
            month = month_raw[:7] if len(month_raw) >= 7 else month_raw
            try:
                mrr = float(row["mrr"])
            except (TypeError, ValueError):
                mrr = 0.0
            customers.add(cust)
            months.add(month)
            cell[(cust, month)] = mrr

    def cust_key(c: str):
        # Parse trailing integer if present (e.g. "Customer 178" -> 178)
        import re
        m = re.search(r"(\d+)\s*$", c)
        if m:
            return (0, int(m.group(1)), c)
        try:
            return (0, int(c), c)
        except ValueError:
            return (1, 0, c)

    customer_list = sorted(customers, key=cust_key)
    month_list = sorted(months)
    return customer_list, month_list, cell


def parse_month_cutoff(s: str) -> str:
    """Parse an --actuals-through value into a canonical 'YYYY-MM' string.
    Accepts 'YYYY-MM', 'YYYY-MM-DD', 'May-26', 'May 2026', '2026-M5', etc.
    (survey.py reports the cutoff as e.g. 'May-26')."""
    s = str(s).strip()
    for fmt in ("%Y-%m", "%Y-%m-%d", "%b-%y", "%b-%Y", "%B %Y", "%b %Y"):
        try:
            d = dt.datetime.strptime(s, fmt).date()
            return f"{d.year:04d}-{d.month:02d}"
        except ValueError:
            continue
    if "-M" in s:  # '2026-M5'
        try:
            y, mm = s.split("-M")
            return f"{int(y):04d}-{int(mm):02d}"
        except ValueError:
            pass
    raise ValueError(
        f"Could not parse --actuals-through {s!r}. Use 'YYYY-MM' (e.g. 2026-05) "
        f"or a month label like 'May-26'."
    )


def month_to_date(month_str: str) -> dt.date:
    y, m = month_str.split("-")[:2]
    return dt.date(int(y), int(m), 1)


def fmt_month_label(month_str: str) -> str:
    """e.g. '2021-01' -> '2021-M1' (matches example output style)."""
    y, m = month_str.split("-")[:2]
    return f"{int(y):04d}-M{int(m)}"


# ---------------------------------------------------------------------------
# Raw Data tab — verbatim copy of source
# ---------------------------------------------------------------------------


def copy_source_sheet_verbatim(src_path: str, src_ws, dest_ws,
                               src_ws_formulas=None) -> None:
    """Copy a source worksheet (or CSV) into dest_ws preserving values, number
    formats, fonts, fills, alignment, borders, merged ranges, column widths,
    row heights, and cell comments. Critical Rule 7: zero edits, no
    reformatting, no color changes.

    Change #1 (formula verbatim): when `src_ws_formulas` (a data_only=False load
    of the same sheet) is supplied, each cell's VALUE is taken from it so live
    formulas (e.g. =EOMONTH(B4,1)) are preserved as formulas instead of being
    flattened to their cached values. Styles still come from `src_ws`. Theme
    colors are carried separately via wb.loaded_theme in deliver()."""
    if src_path.lower().endswith(".csv"):
        with open(src_path, "r", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            for r_idx, row in enumerate(reader, start=1):
                for c_idx, val in enumerate(row, start=1):
                    cast: Any = val
                    if isinstance(val, str):
                        s = val.strip()
                        if s == "":
                            cast = None
                        else:
                            try:
                                cast = float(s) if "." in s else int(s)
                            except (TypeError, ValueError):
                                cast = val
                    dest_ws.cell(row=r_idx, column=c_idx, value=cast)
        return

    for row in src_ws.iter_rows():
        for cell in row:
            # Value comes from the formula load when available (preserves live
            # formulas); style metadata always comes from `cell` (data_only).
            fcell = (src_ws_formulas.cell(row=cell.row, column=cell.column)
                     if src_ws_formulas is not None else cell)
            val = fcell.value
            if val is None and not cell.has_style:
                continue
            dest_cell = dest_ws.cell(row=cell.row, column=cell.column,
                                     value=val)
            if cell.has_style:
                dest_cell.font = _copy(cell.font)
                dest_cell.fill = _copy(cell.fill)
                dest_cell.border = _copy(cell.border)
                dest_cell.alignment = _copy(cell.alignment)
                dest_cell.number_format = cell.number_format
                dest_cell.protection = _copy(cell.protection)
            if cell.comment is not None:
                dest_cell.comment = Comment(cell.comment.text or "",
                                            cell.comment.author or "source")

    for merged_range in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(merged_range))

    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dest_ws.column_dimensions[col_letter].width = dim.width
    for row_num, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dest_ws.row_dimensions[row_num].height = dim.height


# ---------------------------------------------------------------------------
# Two-tab Raw Data tab (no source — built from long CSV)
# ---------------------------------------------------------------------------


def write_raw_from_csv(ws, customers: List[str], months: List[str],
                       cell: Dict[Tuple[str, str], float]) -> None:
    """Two-tab fallback: build Raw Data sheet directly from the long CSV.
    Used only when no --source is supplied."""
    ws.cell(row=1, column=1, value="Customer ID").font = font_subheader()
    ws.cell(row=1, column=1).fill = fill(SUBHEADER_FILL)
    for j, m in enumerate(months):
        c = ws.cell(row=1, column=2 + j, value=month_to_date(m))
        c.number_format = FMT_DATE
        c.font = font_subheader()
        c.fill = fill(SUBHEADER_FILL)
        c.alignment = Alignment(horizontal="center")

    for i, cust in enumerate(customers):
        r = 2 + i
        ws.cell(row=r, column=1, value=cust).font = font_subheader()
        for j, m in enumerate(months):
            v = cell.get((cust, m), 0.0)
            c = ws.cell(row=r, column=2 + j, value=float(v))
            c.number_format = FMT_NUMBER
            c.font = font_hardcode()

    ws.column_dimensions["A"].width = 14
    for j in range(len(months)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 12
    ws.freeze_panes = ws.cell(row=2, column=2)


# ---------------------------------------------------------------------------
# Helper sheet — Raw Data with Analysis (aggregating mode)
# ---------------------------------------------------------------------------


def write_analysis_sheet(
    ws,
    customers: List[str],
    months_analysis: List[str],
    src_sheet_name: str,
    src_customer_col: str,
    src_type_col: str,
    src_first_data_row: int,
    src_last_data_row: int,
    src_first_date_col: str,
    in_scope_types: List[str],
    raw_sheet_name: str = "Raw Data",
) -> None:
    """Build the Raw Data with Analysis helper sheet.

    Layout (per the new SKILL.md):
      Row 1   Month headers          col A = "Customer ID"
      Row 2   # Active customers     COUNTIF on customer rows
      Row 3   # Retained vs N prior  SUMPRODUCT of two-month >0 masks (array)
      Row 4   Check # Active vs Raw  independent recount against Raw Data
      Row 5   blank divider
      Row 6   Recurring MRR total    SUMIFS direct column ref by type
      Row 7   Re-occurring MRR total
      Row 8   Non-recurring MRR total
      Row 9   Total MRR (all types)  = row6+row7+row8
      Row 10  Check vs Raw Data      direct SUM of column on Raw Data
      Row 11  Check (in-scope sum)   = in-scope rows − SUM(customer rows)
      Row 12+ Customer-level data    direct-column SUMIFS by customer + type

    All cells that pull from Raw Data are green (cross-sheet ref). Section
    label cells are black bold.
    """
    # CRITICAL: formulas reference the DESTINATION sheet name ("Raw Data"),
    # not the original source sheet name. The verbatim Raw Data tab keeps the
    # source's contents but is named "Raw Data" in this workbook.
    src_sheet_name = raw_sheet_name
    n_months = len(months_analysis)
    n_cust = len(customers)
    last_cust_row = ANALYSIS_FIRST_CUST_ROW + n_cust - 1

    src_first_date_idx = column_index_from_string(src_first_date_col)

    # Source helper strings — used in every formula
    type_rng = f"'{src_sheet_name}'!${src_type_col}${src_first_data_row}:${src_type_col}${src_last_data_row}"
    cust_rng = f"'{src_sheet_name}'!${src_customer_col}${src_first_data_row}:${src_customer_col}${src_last_data_row}"

    # Header row 1
    hdr_a = ws.cell(row=ANALYSIS_ROW_HDR, column=ANALYSIS_LABEL_COL, value="Customer ID")
    hdr_a.font = font_subheader()
    hdr_a.fill = fill(SUBHEADER_FILL)
    hdr_a.alignment = Alignment(horizontal="left")

    for j, m in enumerate(months_analysis):
        col = ANALYSIS_FIRST_MONTH_COL + j
        cell = ws.cell(row=ANALYSIS_ROW_HDR, column=col, value=fmt_month_label(m))
        cell.font = font_subheader()
        cell.fill = fill(SUBHEADER_FILL)
        cell.alignment = Alignment(horizontal="center")

    # Row 2: # Active customers (COUNTIF on customer rows of THIS sheet)
    ws.cell(row=ANALYSIS_ROW_ACTIVE, column=ANALYSIS_LABEL_COL,
            value="# Active customers").font = font_subheader()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        f = f"=COUNTIF({col_letter}${ANALYSIS_FIRST_CUST_ROW}:{col_letter}${last_cust_row},\">0\")"
        c = ws.cell(row=ANALYSIS_ROW_ACTIVE, column=col, value=f)
        c.font = font_formula()
        c.number_format = FMT_COUNT

    # Row 3: # Retained vs N prior. For first <lookback> cols, value is "n/a".
    LOOKBACK = 12  # YoY; if the dataset is shorter the model is degenerate but harmless
    ws.cell(row=ANALYSIS_ROW_RETAINED, column=ANALYSIS_LABEL_COL,
            value=f"# Retained vs {LOOKBACK}mo prior").font = font_subheader()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        if j < LOOKBACK:
            ws.cell(row=ANALYSIS_ROW_RETAINED, column=col, value="n/a").font = font_formula()
        else:
            prior_letter = get_column_letter(col - LOOKBACK)
            f = (f"=SUMPRODUCT(({col_letter}${ANALYSIS_FIRST_CUST_ROW}:{col_letter}${last_cust_row}>0)"
                 f"*({prior_letter}${ANALYSIS_FIRST_CUST_ROW}:{prior_letter}${last_cust_row}>0))")
            c = ws.cell(row=ANALYSIS_ROW_RETAINED, column=col, value=f)
            c.font = font_formula()
            c.number_format = FMT_COUNT

    # Row 4: Check # Active vs Raw Data — independent recount.
    # Uses SUMPRODUCT over the in-scope type filter against the customer list.
    ws.cell(row=ANALYSIS_ROW_CHECK_ACTIVE, column=ANALYSIS_LABEL_COL,
            value="  Check # Active vs Raw Data").font = font_formula()
    type_filter_or = ",".join(in_scope_types)
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        src_col_letter = get_column_letter(src_first_date_idx + j)
        src_rng = f"'{src_sheet_name}'!${src_col_letter}${src_first_data_row}:${src_col_letter}${src_last_data_row}"
        # An "active" customer is one with a positive in-scope sum.
        # The check is: COUNTIF on analysis sheet row above = count of customers with
        # positive sum on raw direct path.
        # For multi-type filter, use SUMPRODUCT(--(SUMIFS-array > 0)).
        # Since openpyxl can't easily write CSE array formulas, we use a
        # simpler equivalent: count customers where the in-scope SUM is > 0.
        # For single in-scope type — use COUNTIFS directly.
        if len(in_scope_types) == 1:
            t = in_scope_types[0]
            # COUNTIFS counts source rows matching type AND positive — but a
            # customer can have multiple positive rows (multi-product). To
            # count CUSTOMERS we need an aggregation. We approximate with
            # SUMPRODUCT against a unique-customer list. Simpler approach:
            # SUMPRODUCT(1/COUNTIFS) — but that fails on zero rows.
            # Use the array form via SUMPRODUCT with SUMIFS, which works in
            # Excel/LibreOffice as an implicit array context.
            f = (f"={col_letter}{ANALYSIS_ROW_ACTIVE}"
                 f" - SUMPRODUCT(--("
                 f"SUMIFS({src_rng},{type_rng},\"{t}\",{cust_rng},$A${ANALYSIS_FIRST_CUST_ROW}:$A${last_cust_row})>0))")
        else:
            # Multi-type: chain SUMIFS sums per type, then OR via sum.
            # SUMPRODUCT(--((SUMIFS_type1 + SUMIFS_type2 + ...) > 0))
            sumifs_parts = []
            for t in in_scope_types:
                sumifs_parts.append(
                    f"SUMIFS({src_rng},{type_rng},\"{t}\",{cust_rng},$A${ANALYSIS_FIRST_CUST_ROW}:$A${last_cust_row})"
                )
            inner = " + ".join(sumifs_parts)
            f = (f"={col_letter}{ANALYSIS_ROW_ACTIVE}"
                 f" - SUMPRODUCT(--(({inner}) > 0))")
        c = ws.cell(row=ANALYSIS_ROW_CHECK_ACTIVE, column=col, value=f)
        c.font = font_xsheet()  # cross-sheet, green
        c.number_format = FMT_COUNT

    # Row 5 — blank divider (intentionally empty)

    # Rows 6/7/8: per-type totals. Direct column reference per month — each
    # SUMIFS sums the source's date column for rows matching this type.
    type_rows = [
        (ANALYSIS_ROW_REC, "Recurring"),
        (ANALYSIS_ROW_REOCC, "Re-occurring"),
        (ANALYSIS_ROW_NONREC, "Non-recurring"),
    ]

    for row, type_name in type_rows:
        # Label in column A (black formula font)
        ws.cell(row=row, column=ANALYSIS_LABEL_COL,
                value=f"  {type_name}").font = font_formula()
        # If this type isn't in the source at all, still write the row but with
        # SUMIFS that yield 0; useful for the row 10 full-type recon.
        for j in range(n_months):
            col = ANALYSIS_FIRST_MONTH_COL + j
            src_col_letter = get_column_letter(src_first_date_idx + j)
            src_rng_j = (f"'{src_sheet_name}'!"
                         f"${src_col_letter}${src_first_data_row}:"
                         f"${src_col_letter}${src_last_data_row}")
            f = f"=SUMIFS({src_rng_j},{type_rng},\"{type_name}\")"
            c = ws.cell(row=row, column=col, value=f)
            c.font = font_xsheet()
            c.number_format = FMT_NUMBER

    # Row 9: Total MRR (all types)
    ws.cell(row=ANALYSIS_ROW_TOTAL, column=ANALYSIS_LABEL_COL,
            value="Total MRR (all types)").font = font_subheader()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        f = (f"={col_letter}{ANALYSIS_ROW_REC}+{col_letter}{ANALYSIS_ROW_REOCC}"
             f"+{col_letter}{ANALYSIS_ROW_NONREC}")
        c = ws.cell(row=ANALYSIS_ROW_TOTAL, column=col, value=f)
        c.font = font_formula(bold=True)
        c.number_format = FMT_NUMBER

    # Row 10: Check vs Raw Data direct column sum (must = 0)
    ws.cell(row=ANALYSIS_ROW_CHECK_TOTAL, column=ANALYSIS_LABEL_COL,
            value="  Check vs Raw Data").font = font_formula()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        src_col_letter = get_column_letter(src_first_date_idx + j)
        src_rng_j = (f"'{src_sheet_name}'!"
                     f"${src_col_letter}${src_first_data_row}:"
                     f"${src_col_letter}${src_last_data_row}")
        f = f"={col_letter}{ANALYSIS_ROW_TOTAL} - SUM({src_rng_j})"
        c = ws.cell(row=ANALYSIS_ROW_CHECK_TOTAL, column=col, value=f)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER

    # Row 11: Check (in-scope sum) — sum of in-scope type rows above − sum of
    # customer rows below; must = 0 if the customer-row SUMIFS uses the same
    # type filter.
    label = "  Check (" + " + ".join(in_scope_types) + ") vs customer rows"
    ws.cell(row=ANALYSIS_ROW_CHECK_INSCOPE, column=ANALYSIS_LABEL_COL,
            value=label).font = font_formula()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        in_scope_sum = " + ".join(
            f"{col_letter}{ANALYSIS_ROW_REC if t == 'Recurring' else (ANALYSIS_ROW_REOCC if t == 'Re-occurring' else ANALYSIS_ROW_NONREC)}"
            for t in in_scope_types
        )
        f = (f"=({in_scope_sum}) - SUM({col_letter}${ANALYSIS_FIRST_CUST_ROW}:"
             f"{col_letter}${last_cust_row})")
        c = ws.cell(row=ANALYSIS_ROW_CHECK_INSCOPE, column=col, value=f)
        c.font = font_formula()
        c.number_format = FMT_NUMBER

    # Rows 12+: per-customer monthly in-scope MRR.
    # SUMIFS with two criteria: customer id + type filter (one row per in-scope
    # type, summed). When there's >1 in-scope type we add the SUMIFS terms.
    for i, cust in enumerate(customers):
        r = ANALYSIS_FIRST_CUST_ROW + i
        # Customer ID literal (or formula to source? — leave literal since the
        # source has many rows per customer)
        ws.cell(row=r, column=ANALYSIS_LABEL_COL, value=cust).font = font_subheader()
        for j in range(n_months):
            col = ANALYSIS_FIRST_MONTH_COL + j
            src_col_letter = get_column_letter(src_first_date_idx + j)
            src_rng_j = (f"'{src_sheet_name}'!"
                         f"${src_col_letter}${src_first_data_row}:"
                         f"${src_col_letter}${src_last_data_row}")
            sumifs_parts = []
            for t in in_scope_types:
                sumifs_parts.append(
                    f"SUMIFS({src_rng_j},{cust_rng},$A{r},{type_rng},\"{t}\")"
                )
            f = "=" + " + ".join(sumifs_parts)
            c = ws.cell(row=r, column=col, value=f)
            c.font = font_xsheet()
            c.number_format = FMT_NUMBER

    # Column widths
    ws.column_dimensions[get_column_letter(ANALYSIS_LABEL_COL)].width = 38
    for j in range(n_months):
        ws.column_dimensions[get_column_letter(ANALYSIS_FIRST_MONTH_COL + j)].width = 12

    ws.freeze_panes = ws.cell(row=ANALYSIS_FIRST_CUST_ROW, column=ANALYSIS_FIRST_MONTH_COL)


# ---------------------------------------------------------------------------
# Corkscrew sheet — aggregating mode (references Raw Data with Analysis)
# ---------------------------------------------------------------------------


def write_corkscrew_sheet_aggregating(
    ws,
    customers: List[str],
    months_analysis: List[str],
    arr_factor: float,
    company: str,
    in_scope_types: List[str],
    analysis_sheet_name: str,
    lookback: int = 12,
    raw_geometry: dict | None = None,
) -> None:
    """Write the Corkscrew with YoY rollforward (or N-period lookback).

    Two referencing modes:
    - HELPER mode (raw_geometry=None, default): rollforward ranges point at the
      `analysis_sheet_name` helper at its fixed canonical geometry (customer rows
      from ANALYSIS_FIRST_CUST_ROW, months from ANALYSIS_FIRST_MONTH_COL), and
      customer counts come via HLOOKUP into the helper's summary rows 2/3.
    - TWO-SHEET mode (raw_geometry given, Change #2): no helper exists; the
      Corkscrew references Raw Data directly. `raw_geometry` carries the source
      block geometry as written into Raw Data verbatim:
        {sheet, first_row, last_row, first_date_col_idx}
      Rollforward ranges become 'Raw Data'!<srcMonthCol>$<first>:$<last>; customer
      counts are computed INLINE (COUNTIF / SUMPRODUCT) against Raw Data rather
      than pulled from a helper. Only valid for a clean contiguous single-type
      block (the caller gates this).

    All movement formulas reference the Raw Data with Analysis prior and current
    columns. Multi-type recon block written when len(in_scope_types) > 1."""
    n_months = len(months_analysis)
    n_cust = len(customers)
    n_periods = n_months - lookback  # number of comparison periods
    if n_periods <= 0:
        raise ValueError(
            f"Not enough months for {lookback}-period lookback: {n_months} months."
        )

    # Geometry the rollforward formulas reference. HELPER mode uses the canonical
    # helper grid; TWO-SHEET mode (Change #2) uses the source block as written
    # into Raw Data verbatim.
    if raw_geometry is not None:
        ref_sheet = raw_geometry["sheet"]            # e.g. "Raw Data"
        ref_first_row = raw_geometry["first_row"]    # source first customer row
        ref_last_row = raw_geometry["last_row"]      # source last customer row
        ref_first_date_idx = raw_geometry["first_date_col_idx"]  # 1-based col idx

        def month_col_letter(month_index: int) -> str:
            return get_column_letter(ref_first_date_idx + month_index)
    else:
        ref_sheet = analysis_sheet_name
        ref_first_row = ANALYSIS_FIRST_CUST_ROW
        ref_last_row = ANALYSIS_FIRST_CUST_ROW + n_cust - 1

        def month_col_letter(month_index: int) -> str:
            return get_column_letter(ANALYSIS_FIRST_MONTH_COL + month_index)

    last_cust_row = ref_last_row  # used by HLOOKUP header ranges (helper mode)

    months_periods = months_analysis[lookback:]  # comparison-period labels

    # Title (row 1) — centerContinuous, navy fill, white bold
    title_text = (
        f"{company} — YoY ARR Corkscrew & Retention Analysis"
        if company else "YoY ARR Corkscrew & Retention Analysis"
    )
    center_continuous_across(
        ws, ROW_TITLE, 1, FIRST_DATA_COL + n_periods - 1,
        title_text, font_title(), fill(TITLE_FILL),
    )
    ws.row_dimensions[ROW_TITLE].height = 22

    # Generated row
    c = ws.cell(row=ROW_GENERATED, column=1, value="Generated:")
    c.font = font_subheader()
    c2 = ws.cell(row=ROW_GENERATED, column=2, value=dt.date.today().isoformat())
    c2.font = font_hardcode()

    # ARR factor (hardcode — blue)
    label = ws.cell(row=ROW_ARR_FACTOR, column=1, value="ARR Factor (MRR × N):")
    label.font = font_subheader()
    af = ws.cell(row=ROW_ARR_FACTOR, column=2, value=int(arr_factor))
    af.font = font_hardcode(bold=True)
    af.number_format = "0"
    af.comment = Comment(
        f"Source: User-confirmed in Phase 1. Data is "
        f"{'MRR (annualize ×12)' if int(arr_factor) == 12 else 'ARR (factor 1)'}.",
        "retention-analysis"
    )
    arr_ref = f"$B${ROW_ARR_FACTOR}"

    # Date row 5
    lbl5 = ws.cell(row=ROW_DATES, column=COL_LABEL, value="Item")
    lbl5.font = font_banner()
    lbl5.fill = fill(BANNER_FILL)
    lbl5.alignment = Alignment(horizontal="left")
    for j, m in enumerate(months_periods):
        col = FIRST_DATA_COL + j
        c = ws.cell(row=ROW_DATES, column=col, value=fmt_month_label(m))
        c.font = font_banner()
        c.fill = fill(BANNER_FILL)
        c.alignment = Alignment(horizontal="center")

    # Row 6: "(vs. prior year)" prefix
    vs_label = ws.cell(row=ROW_VS, column=COL_LABEL, value="(vs. prior year)")
    vs_label.font = font_subheader()
    vs_label.fill = fill(SUBHEADER_FILL)
    vs_label.alignment = Alignment(horizontal="left")
    for j, m in enumerate(months_periods):
        col = FIRST_DATA_COL + j
        prior_label = fmt_month_label(months_analysis[j])  # prior = T-N
        c = ws.cell(row=ROW_VS, column=col, value=f"vs {prior_label}")
        c.font = font_subheader()
        c.fill = fill(SUBHEADER_FILL)
        c.alignment = Alignment(horizontal="center")

    # Rollforward block. Exactly ONE external check is emitted per column:
    #   - single in-scope type  → ROW_CHECK (row 14) = Ending − period total summed
    #     independently from the analysis/Raw Data sheet × factor.
    #   - multiple in-scope types → the variance at the bottom of the decomposed
    #     reconciliation block (row 36) is the check, with components shown above it.
    # DRY: one check, never two paths to the same algebra.
    rollforward_labels = {
        ROW_BEGIN: ("Beginning ARR (prior year)", True),     # top-of-block: $
        ROW_NEW: ("  + New customer ARR", False),
        ROW_UPSELL: ("  + Expansion (Upsell)", False),
        ROW_DOWNSELL: ("  - Contraction (Downsell)", False),
        ROW_CHURN: ("  - Churn", False),
        ROW_END: ("Ending ARR", True),                        # bottom-of-block: $
    }
    for r, (lbl, _) in rollforward_labels.items():
        c = ws.cell(row=r, column=COL_LABEL, value=lbl)
        c.font = font_subheader()
        c.alignment = Alignment(horizontal="left")
        if r in (ROW_BEGIN, ROW_END):
            c.fill = fill(KEY_METRIC_FILL)

    for j in range(n_periods):
        col = FIRST_DATA_COL + j
        col_letter = get_column_letter(col)
        curr_idx = lookback + j           # source month index for current period
        prior_idx = j                     # source month index for prior period
        curr = month_col_letter(curr_idx)
        prior = month_col_letter(prior_idx)
        # Reference-sheet data ranges (helper grid OR Raw Data block)
        rc = f"'{ref_sheet}'!{curr}${ref_first_row}:{curr}${ref_last_row}"
        rp = f"'{ref_sheet}'!{prior}${ref_first_row}:{prior}${ref_last_row}"

        # Beginning ARR = SUMPRODUCT((prior > 0) * prior) * ARR_factor
        f_beg = f"=SUMPRODUCT(({rp}>0)*{rp})*{arr_ref}"
        c = ws.cell(row=ROW_BEGIN, column=col, value=f_beg)
        c.font = font_xsheet(bold=True)
        c.number_format = FMT_DOLLAR  # top of block — $
        c.fill = fill(KEY_METRIC_FILL)

        # New = SUMPRODUCT((prior=0)*(curr>0)*curr) * factor
        f_new = f"=SUMPRODUCT(({rp}=0)*({rc}>0)*{rc})*{arr_ref}"
        c = ws.cell(row=ROW_NEW, column=col, value=f_new)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER  # interior — no $

        # Upsell = SUMPRODUCT((prior>0)*(curr>prior)*(curr-prior)) * factor
        f_up = f"=SUMPRODUCT(({rp}>0)*({rc}>{rp})*({rc}-{rp}))*{arr_ref}"
        c = ws.cell(row=ROW_UPSELL, column=col, value=f_up)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER

        # Downsell (negative) = SUMPRODUCT((prior>0)*(curr>0)*(curr<prior)*(curr-prior)) * factor
        f_down = f"=SUMPRODUCT(({rp}>0)*({rc}>0)*({rc}<{rp})*({rc}-{rp}))*{arr_ref}"
        c = ws.cell(row=ROW_DOWNSELL, column=col, value=f_down)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER

        # Churn (negative) = SUMPRODUCT((prior>0)*(curr=0)*(-prior)) * factor
        f_ch = f"=SUMPRODUCT(({rp}>0)*({rc}=0)*(-{rp}))*{arr_ref}"
        c = ws.cell(row=ROW_CHURN, column=col, value=f_ch)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER

        # Ending = sum of rollforward
        f_end = (f"={col_letter}{ROW_BEGIN}+{col_letter}{ROW_NEW}"
                 f"+{col_letter}{ROW_UPSELL}+{col_letter}{ROW_DOWNSELL}"
                 f"+{col_letter}{ROW_CHURN}")
        c = ws.cell(row=ROW_END, column=col, value=f_end)
        c.font = font_formula(bold=True)
        c.number_format = FMT_DOLLAR  # bottom of block — $
        c.fill = fill(KEY_METRIC_FILL)

        # Customer counts.
        if raw_geometry is not None:
            # TWO-SHEET mode: compute inline against Raw Data (no helper summary
            # rows to HLOOKUP). One COUNTIF per active count; the retained count
            # is the single legitimate SUMPRODUCT (differential across periods).
            f_n_prior = f"=COUNTIF({rp},\">0\")"
            f_n_curr = f"=COUNTIF({rc},\">0\")"
            f_retained_expr = f"SUMPRODUCT(({rc}>0)*({rp}>0))"
            ws.cell(row=ROW_N_ACTIVE_PRIOR, column=col, value=f_n_prior).font = font_xsheet()
            ws.cell(row=ROW_N_ACTIVE_CURR, column=col, value=f_n_curr).font = font_xsheet()
            f_n_ch = f"={col_letter}{ROW_N_ACTIVE_PRIOR} - {f_retained_expr}"
            f_n_new = f"={col_letter}{ROW_N_ACTIVE_CURR} - {f_retained_expr}"
        else:
            # HELPER mode: HLOOKUP into the helper's summary rows 2/3.
            analysis_hdr_range = (f"'{analysis_sheet_name}'!"
                              f"$B${ANALYSIS_ROW_HDR}:${get_column_letter(ANALYSIS_FIRST_MONTH_COL + n_months - 1)}${ANALYSIS_ROW_ACTIVE}")
            f_n_prior = f"=HLOOKUP(SUBSTITUTE({col_letter}${ROW_VS},\"vs \",\"\"),{analysis_hdr_range},2,FALSE)"
            f_n_curr = f"=HLOOKUP({col_letter}${ROW_DATES},{analysis_hdr_range},2,FALSE)"
            ws.cell(row=ROW_N_ACTIVE_PRIOR, column=col, value=f_n_prior).font = font_xsheet()
            ws.cell(row=ROW_N_ACTIVE_CURR, column=col, value=f_n_curr).font = font_xsheet()
            analysis_ret_range = (f"'{analysis_sheet_name}'!"
                              f"$B${ANALYSIS_ROW_HDR}:${get_column_letter(ANALYSIS_FIRST_MONTH_COL + n_months - 1)}${ANALYSIS_ROW_RETAINED}")
            f_retained = f"=HLOOKUP({col_letter}${ROW_DATES},{analysis_ret_range},3,FALSE)"
            f_n_ch = f"={col_letter}{ROW_N_ACTIVE_PRIOR} - {f_retained[1:]}"  # = prior − retained
            f_n_new = f"={col_letter}{ROW_N_ACTIVE_CURR} - {f_retained[1:]}"  # = current − retained

        ws.cell(row=ROW_N_ACTIVE_PRIOR, column=col).number_format = FMT_COUNT
        ws.cell(row=ROW_N_ACTIVE_CURR, column=col).number_format = FMT_COUNT
        ws.cell(row=ROW_N_CHURNED, column=col, value=f_n_ch).font = font_formula()
        ws.cell(row=ROW_N_CHURNED, column=col).number_format = FMT_COUNT
        ws.cell(row=ROW_N_NEW, column=col, value=f_n_new).font = font_formula()
        ws.cell(row=ROW_N_NEW, column=col).number_format = FMT_COUNT

        # Retention metrics (with IFERROR)
        beg = f"{col_letter}{ROW_BEGIN}"
        f_grr = f"=IFERROR(({beg}+{col_letter}{ROW_DOWNSELL}+{col_letter}{ROW_CHURN})/{beg},0)"
        f_nrr = f"=IFERROR(({beg}+{col_letter}{ROW_UPSELL}+{col_letter}{ROW_DOWNSELL}+{col_letter}{ROW_CHURN})/{beg},0)"
        f_logo = f"=IFERROR(({col_letter}{ROW_N_ACTIVE_PRIOR}-{col_letter}{ROW_N_CHURNED})/{col_letter}{ROW_N_ACTIVE_PRIOR},0)"
        for r, fx in ((ROW_GRR, f_grr), (ROW_NRR, f_nrr), (ROW_LOGO, f_logo)):
            cc = ws.cell(row=r, column=col, value=fx)
            cc.font = font_formula()
            cc.number_format = FMT_PCT

        # Per-customer metrics
        f_avg = f"=IFERROR({col_letter}{ROW_END}/{col_letter}{ROW_N_ACTIVE_CURR},0)"
        f_avg_new = f"=IFERROR({col_letter}{ROW_NEW}/{col_letter}{ROW_N_NEW},0)"
        cc = ws.cell(row=ROW_AVG_ARR, column=col, value=f_avg)
        cc.font = font_formula()
        cc.number_format = FMT_DOLLAR  # singleton numeric row — top & bottom $
        cc = ws.cell(row=ROW_AVG_NEW, column=col, value=f_avg_new)
        cc.font = font_formula()
        cc.number_format = FMT_DOLLAR

        # Decomposed reconciliation (only when multi-type scope)
        if len(in_scope_types) > 1:
            for t in in_scope_types:
                if t == "Recurring":
                    f_rec = f"='{analysis_sheet_name}'!{curr}{ANALYSIS_ROW_REC}*{arr_ref}"
                    c = ws.cell(row=ROW_REC_RECURRING, column=col, value=f_rec)
                    c.font = font_xsheet()
                    c.number_format = FMT_DOLLAR
                elif t == "Re-occurring":
                    f_reocc = f"='{analysis_sheet_name}'!{curr}{ANALYSIS_ROW_REOCC}*{arr_ref}"
                    c = ws.cell(row=ROW_REC_REOCCURRING, column=col, value=f_reocc)
                    c.font = font_xsheet()
                    c.number_format = FMT_NUMBER
            f_sum = f"={col_letter}{ROW_REC_RECURRING}+{col_letter}{ROW_REC_REOCCURRING}"
            c = ws.cell(row=ROW_REC_SUM, column=col, value=f_sum)
            c.font = font_formula(bold=True)
            c.number_format = FMT_DOLLAR
            f_var = f"={col_letter}{ROW_REC_SUM}-{col_letter}{ROW_END}"
            c = ws.cell(row=ROW_REC_VARIANCE, column=col, value=f_var)
            c.font = font_formula()
            c.number_format = FMT_NUMBER
        else:
            # Single in-scope type → external reconciliation right under Ending.
            # Ending must equal the period's in-scope total summed independently
            # from the analysis/Raw Data sheet (a different formula path → a real
            # check, not the tautological Beginning+moves=Ending identity).
            f_chk = f"={col_letter}{ROW_END} - SUM({rc})*{arr_ref}"
            c = ws.cell(row=ROW_CHECK, column=col, value=f_chk)
            c.font = font_xsheet()
            c.number_format = FMT_NUMBER

    # Section banners (banner row above each block)
    banners = [
        (ROW_CC_BANNER, "CUSTOMER COUNTS", n_periods),
        (ROW_RR_BANNER, "RETENTION RATES", n_periods),
        (ROW_PC_BANNER, "PER-CUSTOMER METRICS", n_periods),
    ]
    if len(in_scope_types) > 1:
        banners.append((ROW_RECON_BANNER, "RECONCILIATION CHECKS", n_periods))
    for row, txt, span_cols in banners:
        center_continuous_across(
            ws, row, 1, FIRST_DATA_COL + span_cols - 1,
            txt, font_banner(), fill(BANNER_FILL),
        )

    # Row labels for CC / RR / PC / Recon sections
    rr_labels = {
        ROW_N_ACTIVE_PRIOR: "# Active (prior period)",
        ROW_N_ACTIVE_CURR: "# Active (current period)",
        ROW_N_CHURNED: "# Churned",
        ROW_N_NEW: "# New",
        ROW_GRR: "Gross Dollar Retention (GRR)",
        ROW_NRR: "Net Dollar Retention (NRR)",
        ROW_LOGO: "Logo Retention",
        ROW_AVG_ARR: "Avg ARR per Active Customer",
        ROW_AVG_NEW: "Avg ARR per New Customer",
    }
    if len(in_scope_types) > 1:
        rr_labels.update({
            ROW_REC_RECURRING: "Recurring ARR",
            ROW_REC_REOCCURRING: "Re-occurring ARR",
            ROW_REC_SUM: "Sum customer ARR",
            ROW_REC_VARIANCE: "Variance vs Ending ARR (= 0)",
        })
    else:
        rr_labels[ROW_CHECK] = "External Check (Ending - Raw Data) = 0"
    for r, txt in rr_labels.items():
        c = ws.cell(row=r, column=COL_LABEL, value=txt)
        c.font = font_subheader()
        c.alignment = Alignment(horizontal="left")

    # Column widths
    ws.column_dimensions[get_column_letter(COL_LABEL)].width = 38
    for j in range(n_periods):
        ws.column_dimensions[get_column_letter(FIRST_DATA_COL + j)].width = 13

    ws.freeze_panes = ws.cell(row=ROW_DATES + 2, column=FIRST_DATA_COL)


# ---------------------------------------------------------------------------
# Pass-through "Raw Data with Analysis" sheet — legacy (one row per customer source)
# ---------------------------------------------------------------------------


def write_analysis_passthrough_sheet(
    ws,
    customers: List[str],
    months: List[str],
    customer_to_src_row: Dict[str, int],
    src_customer_col: str,
    src_first_date_col: str,
    excluded_customers: List[str] | None = None,
) -> None:
    """Pass-through helper: one row per customer, formulas pulling Raw Data
    cells one-to-one. Used when the source already has one row per customer."""
    excluded_customers = excluded_customers or []
    src_date_col_idx = column_index_from_string(src_first_date_col)
    n_months = len(months)
    excl_col = ANALYSIS_FIRST_MONTH_COL + n_months
    last_cust_row = ANALYSIS_FIRST_CUST_ROW + len(customers) - 1

    # ---- Summary block (rows 1-3) — same shape the Corkscrew HLOOKUPs expect.
    # Row 1: string month headers ("2022-M1") so HLOOKUP matches the Corkscrew
    #        date row (also string labels). Row 2: # Active. Row 3: # Retained.
    hdr_a = ws.cell(row=ANALYSIS_ROW_HDR, column=ANALYSIS_LABEL_COL, value="Customer ID")
    hdr_a.font = font_subheader(); hdr_a.fill = fill(SUBHEADER_FILL)
    hdr_a.alignment = Alignment(horizontal="left")
    for j, m in enumerate(months):
        c = ws.cell(row=ANALYSIS_ROW_HDR, column=ANALYSIS_FIRST_MONTH_COL + j,
                    value=fmt_month_label(m))
        c.font = font_subheader(); c.fill = fill(SUBHEADER_FILL)
        c.alignment = Alignment(horizontal="center")
    ex_hdr = ws.cell(row=ANALYSIS_ROW_HDR, column=excl_col, value="Excluded?")
    ex_hdr.font = font_subheader(); ex_hdr.fill = fill(SUBHEADER_FILL)

    # Row 2: # Active customers (COUNTIF on this sheet's customer rows)
    ws.cell(row=ANALYSIS_ROW_ACTIVE, column=ANALYSIS_LABEL_COL,
            value="# Active customers").font = font_subheader()
    for j in range(n_months):
        cl = get_column_letter(ANALYSIS_FIRST_MONTH_COL + j)
        c = ws.cell(row=ANALYSIS_ROW_ACTIVE, column=ANALYSIS_FIRST_MONTH_COL + j,
                    value=f"=COUNTIF({cl}${ANALYSIS_FIRST_CUST_ROW}:{cl}${last_cust_row},\">0\")")
        c.font = font_formula(); c.number_format = FMT_COUNT

    # Row 3: # Retained vs N-mo prior ("n/a" for the first <lookback> columns)
    LOOKBACK = 12
    ws.cell(row=ANALYSIS_ROW_RETAINED, column=ANALYSIS_LABEL_COL,
            value=f"# Retained vs {LOOKBACK}mo prior").font = font_subheader()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        cl = get_column_letter(col)
        if j < LOOKBACK:
            ws.cell(row=ANALYSIS_ROW_RETAINED, column=col, value="n/a").font = font_formula()
        else:
            pl = get_column_letter(col - LOOKBACK)
            c = ws.cell(row=ANALYSIS_ROW_RETAINED, column=col,
                        value=(f"=SUMPRODUCT(({cl}${ANALYSIS_FIRST_CUST_ROW}:{cl}${last_cust_row}>0)"
                               f"*({pl}${ANALYSIS_FIRST_CUST_ROW}:{pl}${last_cust_row}>0))"))
            c.font = font_formula(); c.number_format = FMT_COUNT

    first_data_row = ANALYSIS_FIRST_CUST_ROW

    for i, cust in enumerate(customers):
        r = first_data_row + i
        src_row = customer_to_src_row.get(str(cust))
        if src_row is not None:
            ws.cell(row=r, column=1, value=f"='Raw Data'!{src_customer_col}{src_row}").font = font_xsheet()
        else:
            ws.cell(row=r, column=1, value=cust).font = font_subheader()
        for j in range(n_months):
            col = ANALYSIS_FIRST_MONTH_COL + j
            if src_row is not None:
                src_col_letter = get_column_letter(src_date_col_idx + j)
                f = (f"=IF('Raw Data'!{src_col_letter}{src_row}=\"\",0,"
                     f"'Raw Data'!{src_col_letter}{src_row})")
                c = ws.cell(row=r, column=col, value=f)
                c.font = font_xsheet()
            else:
                c = ws.cell(row=r, column=col, value=0.0)
                c.font = font_hardcode()
            c.number_format = FMT_NUMBER
        ws.cell(row=r, column=excl_col, value=False).font = font_formula()

    # Excluded rows below
    base = first_data_row + len(customers)
    for i, cust in enumerate(excluded_customers):
        r = base + i
        src_row = customer_to_src_row.get(str(cust))
        if src_row is not None:
            ws.cell(row=r, column=1, value=f"='Raw Data'!{src_customer_col}{src_row}").font = font_xsheet()
        else:
            ws.cell(row=r, column=1, value=cust).font = font_subheader()
        for j in range(n_months):
            col = ANALYSIS_FIRST_MONTH_COL + j
            c = ws.cell(row=r, column=col, value=0.0)
            c.font = font_hardcode()
            c.number_format = FMT_NUMBER
        c2 = ws.cell(row=r, column=excl_col, value=True)
        c2.font = Font(name="Calibri", size=10, bold=True, color="C00000")

    ws.column_dimensions["A"].width = 18
    for j in range(n_months):
        ws.column_dimensions[get_column_letter(ANALYSIS_FIRST_MONTH_COL + j)].width = 12
    ws.column_dimensions[get_column_letter(excl_col)].width = 12
    ws.freeze_panes = ws.cell(row=first_data_row, column=ANALYSIS_FIRST_MONTH_COL)


def build_customer_to_src_row_map(src_path: str, src_customer_col: str,
                                  src_first_data_row: int, src_ws=None,
                                  src_last_data_row: int | None = None) -> Dict[str, int]:
    if src_path.lower().endswith(".csv"):
        out: Dict[str, int] = {}
        with open(src_path, "r", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            seen = set()
            r = src_first_data_row
            for row in reader:
                cid = str(row.get("customer_id", "")).strip()
                if cid and cid not in seen:
                    out[cid] = r
                    seen.add(cid)
                    r += 1
        return out

    col_idx = column_index_from_string(src_customer_col)
    # Stop at src_last_data_row when given, so a summary/total block BELOW the
    # customer list (within the sheet) isn't scooped up as bogus "customers".
    last_row = src_last_data_row if src_last_data_row else src_ws.max_row
    out = {}
    for r in range(src_first_data_row, last_row + 1):
        v = src_ws.cell(row=r, column=col_idx).value
        if v is None or v == "":
            continue
        out[str(v).strip()] = r
    return out


# ---------------------------------------------------------------------------
# Source-sheet introspection helpers (aggregating mode)
# ---------------------------------------------------------------------------


def find_source_last_data_row(src_ws, src_customer_col: str,
                              src_first_data_row: int) -> int:
    """Find the last row in the source where the customer column has a value."""
    col_idx = column_index_from_string(src_customer_col)
    last = src_first_data_row
    for r in range(src_first_data_row, src_ws.max_row + 1):
        if src_ws.cell(row=r, column=col_idx).value not in (None, ""):
            last = r
    return last


def get_source_months(src_ws, src_first_date_col: str,
                      header_row: int = None) -> List[str]:
    """Read the date headers from the source sheet (the row above first data
    row, or an explicit header_row). Returns 'YYYY-MM' strings."""
    ws = src_ws
    first_idx = column_index_from_string(src_first_date_col)
    # If header_row not given, try row 1 then row 2.
    rows_to_try = [header_row] if header_row else [1, 2]
    for hr in rows_to_try:
        if hr is None or hr < 1:
            continue
        months = []
        c = first_idx
        while True:
            v = ws.cell(row=hr, column=c).value
            if v is None:
                break
            if isinstance(v, (dt.datetime, dt.date)):
                months.append(f"{v.year:04d}-{v.month:02d}")
            else:
                s = str(v).strip()
                # try parse
                parsed = None
                for fmt in ("%Y-%m-%d", "%Y-%m", "%b-%y", "%b %Y", "%B %Y", "%Y-M%m"):
                    try:
                        parsed = dt.datetime.strptime(s, fmt).date()
                        break
                    except ValueError:
                        continue
                if parsed is None and "-M" in s:
                    # "2021-M1" style
                    try:
                        y, mm = s.split("-M")
                        parsed = dt.date(int(y), int(mm), 1)
                    except ValueError:
                        pass
                if parsed is None:
                    break
                months.append(f"{parsed.year:04d}-{parsed.month:02d}")
            c += 1
        if len(months) >= 2:
            return months
    return []


# ---------------------------------------------------------------------------
# Top-level deliver()
# ---------------------------------------------------------------------------


def deliver(
    long_csv_path: str,
    output_xlsx_path: str,
    arr_factor: float = 12.0,
    compute_json_path: str | None = None,
    company: str = "",
    source_path: str | None = None,
    source_sheet: str | None = None,
    source_customer_col: str = "A",
    source_first_data_row: int = 2,
    source_first_date_col: str = "B",
    source_type_col: str | None = None,
    type_filter: List[str] | None = None,
    lookback: int = 12,
    source_header_row: int | None = None,
    source_last_data_row: int | None = None,
    actuals_through: str | None = None,
    two_sheet: bool = False,
) -> str:
    customers, months, cell = load_long_csv(long_csv_path)

    # Actuals cutoff (#2): drop the in-progress current month and any forecast
    # tail so projections aren't counted as retention. Applies to every mode —
    # the helper and corkscrew iterate this month list.
    if actuals_through:
        cutoff = parse_month_cutoff(actuals_through)
        kept = [m for m in months if m <= cutoff]
        if len(kept) < 2:
            raise ValueError(
                f"--actuals-through {actuals_through!r} (= {cutoff}) leaves only "
                f"{len(kept)} month(s) of the {len(months)} in the CSV — need >= 2."
            )
        months = kept
    # compute.py output is OPTIONAL. The workbook is built independently from the
    # CSV as live formulas; if a compute.json is supplied we only adopt its
    # arr_factor (the agent can use the rest as an out-of-band cross-check).
    arr_factor = float(arr_factor)
    if compute_json_path:
        compute = load_compute_json(compute_json_path)
        arr_factor = float(compute.get("config", {}).get("arr_factor", arr_factor))

    wb = Workbook()
    ws_cork = wb.active
    ws_cork.title = "Corkscrew"

    mode = "twotab"
    if source_path and source_type_col:
        mode = "aggregating"
    elif source_path:
        mode = "passthrough"

    # Change #2 — two-sheet / no-helper path. Opt-in (--two-sheet), and ONLY
    # valid for a clean passthrough source (one contiguous row per customer, a
    # single in-scope revenue type). The caller sets the flag after survey.py
    # confirms a clean contiguous block; here we hard-gate so a misuse fails
    # loudly rather than silently shipping a wrong file.
    if two_sheet:
        if not source_path or source_path.lower().endswith(".csv"):
            raise ValueError(
                "--two-sheet needs an Excel --source (it references Raw Data "
                "directly). For a tidy CSV the two-sheet layout is already the "
                "default."
            )
        if source_type_col:
            raise ValueError(
                "--two-sheet is incompatible with --source-type-col: a type "
                "column means aggregation is needed, which requires the helper "
                "sheet. Drop --two-sheet or drop the type column."
            )
        mode = "twosheet"

    # Load the source workbook for layout discovery (data_only=True so computed
    # customer names / date headers like ="Customer "&ROW() and =EOMONTH(...)
    # resolve to the values the scans need). A SECOND data_only=False load
    # supplies the verbatim Raw Data copy so formulas survive (Change #1) — the
    # data_only worksheet would hand us cached values, flattening live formulas.
    src_ws = None          # values (discovery)
    src_ws_formulas = None  # formula strings (verbatim copy)
    src_theme = None        # source theme1.xml bytes (verbatim theme colors)
    if source_path and not source_path.lower().endswith(".csv"):
        _src_wb = load_workbook(source_path, data_only=True, read_only=False)
        if source_sheet is None or source_sheet not in _src_wb.sheetnames:
            raise ValueError(
                f"Source sheet {source_sheet!r} not found in {source_path}. "
                f"Available: {_src_wb.sheetnames}"
            )
        src_ws = _src_wb[source_sheet]
        _src_wb_f = load_workbook(source_path, data_only=False, read_only=False)
        src_ws_formulas = _src_wb_f[source_sheet]
        # Carry the SOURCE theme palette so theme-indexed colors (e.g. a header
        # fill referencing accent5 = #A02B93) render with the source's resolved
        # RGB instead of openpyxl's default theme (which would re-render teal).
        # Corkscrew/helper banners use explicit hardcoded RGB, so this does NOT
        # change them.
        src_theme = _src_wb_f.loaded_theme
    if src_theme is not None:
        wb.loaded_theme = src_theme

    if mode == "aggregating":
        # Three sheets: Corkscrew, Raw Data with Analysis, Raw Data
        ws_analysis = wb.create_sheet("Raw Data with Analysis")
        ws_raw = wb.create_sheet("Raw Data")

        # 1. Raw Data — verbatim
        copy_source_sheet_verbatim(source_path, src_ws, ws_raw, src_ws_formulas)

        # 2. Discover source layout. An explicit --source-last-data-row wins over
        # auto-detection (#3) — caps the customer block so a summary/total block
        # below it isn't aggregated in.
        src_last_row = source_last_data_row or find_source_last_data_row(
            src_ws, source_customer_col, source_first_data_row
        )
        in_scope = type_filter or ["Recurring", "Re-occurring"]
        # Use the analysis sheet's month list = full source month range
        months_analysis = get_source_months(
            src_ws, source_first_date_col, header_row=source_header_row,
        )
        if not months_analysis:
            # Fallback to the long-CSV month list
            months_analysis = months
        # Honor the actuals cutoff (#2) on the source-derived month list too.
        if actuals_through:
            months_analysis = [m for m in months_analysis if m <= parse_month_cutoff(actuals_through)]

        # 3. Raw Data with Analysis
        write_analysis_sheet(
            ws_analysis,
            customers=customers,
            months_analysis=months_analysis,
            src_sheet_name=source_sheet,
            src_customer_col=source_customer_col,
            src_type_col=source_type_col,
            src_first_data_row=source_first_data_row,
            src_last_data_row=src_last_row,
            src_first_date_col=source_first_date_col,
            in_scope_types=in_scope,
        )

        # 4. Corkscrew
        write_corkscrew_sheet_aggregating(
            ws_cork,
            customers=customers,
            months_analysis=months_analysis,
            arr_factor=arr_factor,
            company=company,
            in_scope_types=in_scope,
            analysis_sheet_name="Raw Data with Analysis",
            lookback=lookback,
        )

    elif mode == "passthrough":
        ws_helper = wb.create_sheet("Raw Data with Analysis")
        ws_raw = wb.create_sheet("Raw Data")
        copy_source_sheet_verbatim(source_path, src_ws, ws_raw, src_ws_formulas)

        customer_to_src_row = build_customer_to_src_row_map(
            source_path, source_customer_col, source_first_data_row, src_ws,
            src_last_data_row=source_last_data_row,
        )
        src_customers_in_order = list(customer_to_src_row.keys())
        post_excl = set(map(str, customers))
        excluded = [c for c in src_customers_in_order if c not in post_excl]

        write_analysis_passthrough_sheet(
            ws_helper, customers, months,
            customer_to_src_row=customer_to_src_row,
            src_customer_col=source_customer_col,
            src_first_date_col=source_first_date_col,
            excluded_customers=excluded,
        )
        # Corkscrew references the pass-through helper — same formulas as
        # aggregating but pointing at a different helper sheet, with single-
        # type recon (passthrough = one type in scope by definition).
        write_corkscrew_sheet_aggregating(
            ws_cork,
            customers=customers,
            months_analysis=months,
            arr_factor=arr_factor,
            company=company,
            in_scope_types=["Recurring"],  # treated as a single-type bucket
            analysis_sheet_name="Raw Data with Analysis",
            lookback=lookback,
        )

    elif mode == "twosheet":
        # Change #2 — two sheets only: Corkscrew + Raw Data (verbatim). No
        # helper. The Corkscrew references Raw Data directly via raw_geometry.
        ws_raw = wb.create_sheet("Raw Data")
        copy_source_sheet_verbatim(source_path, src_ws, ws_raw, src_ws_formulas)

        # Determine the source customer block bounds written into Raw Data.
        src_last_row = source_last_data_row or find_source_last_data_row(
            src_ws, source_customer_col, source_first_data_row
        )
        # Months iterated by the corkscrew = source month headers (honoring the
        # actuals cutoff). They live in Raw Data starting at source_first_date_col,
        # so index 0 = source_first_date_col, index 1 = next column, etc.
        months_ts = get_source_months(
            src_ws, source_first_date_col, header_row=source_header_row,
        ) or months
        if actuals_through:
            months_ts = [m for m in months_ts
                         if m <= parse_month_cutoff(actuals_through)]

        raw_geometry = {
            "sheet": "Raw Data",
            "first_row": source_first_data_row,
            "last_row": src_last_row,
            "first_date_col_idx": column_index_from_string(source_first_date_col),
        }
        write_corkscrew_sheet_aggregating(
            ws_cork,
            customers=customers,
            months_analysis=months_ts,
            arr_factor=arr_factor,
            company=company,
            in_scope_types=["Recurring"],  # single-type bucket by gate
            analysis_sheet_name="Raw Data",
            lookback=lookback,
            raw_geometry=raw_geometry,
        )

    else:
        # twotab
        ws_raw = wb.create_sheet("Raw Data")
        write_raw_from_csv(ws_raw, customers, months, cell)
        write_corkscrew_sheet_aggregating(
            ws_cork,
            customers=customers,
            months_analysis=months,
            arr_factor=arr_factor,
            company=company,
            in_scope_types=["Recurring"],
            analysis_sheet_name="Raw Data",
            lookback=lookback,
        )

    wb.save(output_xlsx_path)
    return output_xlsx_path


# ---------------------------------------------------------------------------
# Multi-segment (Change #3)
# ---------------------------------------------------------------------------


def write_blended_corkscrew(
    ws,
    seg_specs: List[dict],
    n_periods: int,
    months_periods: List[str],
    months_prior: List[str],
    arr_factor: float,
    company: str,
) -> None:
    """Write the Blended Corkscrew.

    Each `seg_specs` entry carries the per-segment Corkscrew sheet name plus the
    geometry of that segment's Raw Data block (sheet, first_row, last_row,
    first_date_col_idx, lookback) so the blended sheet can compute an INDEPENDENT
    reconciliation.

    Real reconciliation (not summing the segments' own check rows):
      * Blended Beginning(t) = Σ  '<seg> Corkscrew'!<col>ROW_BEGIN
      * Blended Ending(t)    = Σ  '<seg> Corkscrew'!<col>ROW_END
        (New/Upsell/Downsell/Churn likewise summed so the rollforward identity holds)
      * Variance(t) = Blended Ending(t)
                      − Σ_seg( SUM('<seg> Raw Data'!<curr month col over its rows>) × factor )
        The right-hand side is an INDEPENDENT path straight from each segment's
        Raw Data — it never references the segment Beginning+moves=Ending identity
        — so Variance == 0 every period proves the blend ties to source.
    Blended retention is computed from the blended dollar rows and Σ counts."""
    arr_ref = f"$B${ROW_ARR_FACTOR}"

    # Title + ARR factor + date rows (same shape as a segment corkscrew).
    title_text = (f"{company} — Blended ARR Corkscrew & Retention Analysis"
                  if company else "Blended ARR Corkscrew & Retention Analysis")
    center_continuous_across(ws, ROW_TITLE, 1, FIRST_DATA_COL + n_periods - 1,
                             title_text, font_title(), fill(TITLE_FILL))
    ws.row_dimensions[ROW_TITLE].height = 22
    ws.cell(row=ROW_GENERATED, column=1, value="Generated:").font = font_subheader()
    ws.cell(row=ROW_GENERATED, column=2, value=dt.date.today().isoformat()).font = font_hardcode()
    ws.cell(row=ROW_ARR_FACTOR, column=1, value="ARR Factor (MRR × N):").font = font_subheader()
    af = ws.cell(row=ROW_ARR_FACTOR, column=2, value=int(arr_factor))
    af.font = font_hardcode(bold=True); af.number_format = "0"

    lbl5 = ws.cell(row=ROW_DATES, column=COL_LABEL, value="Item")
    lbl5.font = font_banner(); lbl5.fill = fill(BANNER_FILL)
    lbl5.alignment = Alignment(horizontal="left")
    for j, m in enumerate(months_periods):
        c = ws.cell(row=ROW_DATES, column=FIRST_DATA_COL + j, value=fmt_month_label(m))
        c.font = font_banner(); c.fill = fill(BANNER_FILL)
        c.alignment = Alignment(horizontal="center")
    vs = ws.cell(row=ROW_VS, column=COL_LABEL, value="(vs. prior year)")
    vs.font = font_subheader(); vs.fill = fill(SUBHEADER_FILL)
    for j, m in enumerate(months_prior):
        c = ws.cell(row=ROW_VS, column=FIRST_DATA_COL + j, value=f"vs {fmt_month_label(m)}")
        c.font = font_subheader(); c.fill = fill(SUBHEADER_FILL)
        c.alignment = Alignment(horizontal="center")

    rollforward_labels = {
        ROW_BEGIN: ("Beginning ARR (prior year)", True),
        ROW_NEW: ("  + New customer ARR", False),
        ROW_UPSELL: ("  + Expansion (Upsell)", False),
        ROW_DOWNSELL: ("  - Contraction (Downsell)", False),
        ROW_CHURN: ("  - Churn", False),
        ROW_END: ("Ending ARR", True),
    }
    for r, (lbl, _) in rollforward_labels.items():
        c = ws.cell(row=r, column=COL_LABEL, value=lbl)
        c.font = font_subheader(); c.alignment = Alignment(horizontal="left")
        if r in (ROW_BEGIN, ROW_END):
            c.fill = fill(KEY_METRIC_FILL)

    summed_rows = (ROW_BEGIN, ROW_NEW, ROW_UPSELL, ROW_DOWNSELL, ROW_CHURN)
    count_rows = (ROW_N_ACTIVE_PRIOR, ROW_N_ACTIVE_CURR, ROW_N_CHURNED, ROW_N_NEW)

    for j in range(n_periods):
        col = FIRST_DATA_COL + j
        col_letter = get_column_letter(col)

        # Blended rollforward movement rows = Σ segment corkscrew cells.
        for r in summed_rows:
            parts = [f"'{s['corkscrew_sheet']}'!{col_letter}{r}" for s in seg_specs]
            f = "=" + "+".join(parts)
            c = ws.cell(row=r, column=col, value=f)
            c.font = font_xsheet(bold=(r == ROW_BEGIN))
            c.number_format = FMT_DOLLAR if r == ROW_BEGIN else FMT_NUMBER
            if r == ROW_BEGIN:
                c.fill = fill(KEY_METRIC_FILL)

        # Blended Ending = Σ segment Ending (cross-sheet refs).
        end_parts = [f"'{s['corkscrew_sheet']}'!{col_letter}{ROW_END}" for s in seg_specs]
        c = ws.cell(row=ROW_END, column=col, value="=" + "+".join(end_parts))
        c.font = font_xsheet(bold=True); c.number_format = FMT_DOLLAR
        c.fill = fill(KEY_METRIC_FILL)

        # INDEPENDENT variance: Blended Ending − Σ_seg(SUM(seg Raw Data curr col)×factor).
        indep_terms = []
        for s in seg_specs:
            curr_letter = get_column_letter(s["first_date_col_idx"] + s["lookback"] + j)
            indep_terms.append(
                f"SUM('{s['raw_sheet']}'!{curr_letter}{s['first_row']}:"
                f"{curr_letter}{s['last_row']})")
        indep = "+".join(indep_terms)
        f_var = f"={col_letter}{ROW_END}-({indep})*{arr_ref}"
        cv = ws.cell(row=ROW_CHECK, column=col, value=f_var)
        cv.font = font_xsheet(); cv.number_format = FMT_NUMBER

        # Blended counts = Σ segment counts.
        for r in count_rows:
            parts = [f"'{s['corkscrew_sheet']}'!{col_letter}{r}" for s in seg_specs]
            c = ws.cell(row=r, column=col, value="=" + "+".join(parts))
            c.font = font_xsheet(); c.number_format = FMT_COUNT

        # Blended retention from blended dollar rows + Σ counts.
        beg = f"{col_letter}{ROW_BEGIN}"
        ws.cell(row=ROW_GRR, column=col,
                value=f"=IFERROR(({beg}+{col_letter}{ROW_DOWNSELL}+{col_letter}{ROW_CHURN})/{beg},0)"
                ).number_format = FMT_PCT
        ws.cell(row=ROW_NRR, column=col,
                value=f"=IFERROR(({beg}+{col_letter}{ROW_UPSELL}+{col_letter}{ROW_DOWNSELL}+{col_letter}{ROW_CHURN})/{beg},0)"
                ).number_format = FMT_PCT
        ws.cell(row=ROW_LOGO, column=col,
                value=f"=IFERROR(({col_letter}{ROW_N_ACTIVE_PRIOR}-{col_letter}{ROW_N_CHURNED})/{col_letter}{ROW_N_ACTIVE_PRIOR},0)"
                ).number_format = FMT_PCT
        for r in (ROW_GRR, ROW_NRR, ROW_LOGO):
            ws.cell(row=r, column=col).font = font_formula()

    # Banners + labels.
    for row, txt in ((ROW_CC_BANNER, "CUSTOMER COUNTS"),
                     (ROW_RR_BANNER, "RETENTION RATES")):
        center_continuous_across(ws, row, 1, FIRST_DATA_COL + n_periods - 1,
                                 txt, font_banner(), fill(BANNER_FILL))
    labels = {
        ROW_CHECK: "Blended Variance vs Source (= 0)",
        ROW_N_ACTIVE_PRIOR: "# Active (prior period)",
        ROW_N_ACTIVE_CURR: "# Active (current period)",
        ROW_N_CHURNED: "# Churned",
        ROW_N_NEW: "# New",
        ROW_GRR: "Gross Dollar Retention (GRR)",
        ROW_NRR: "Net Dollar Retention (NRR)",
        ROW_LOGO: "Logo Retention",
    }
    for r, txt in labels.items():
        c = ws.cell(row=r, column=COL_LABEL, value=txt)
        c.font = font_subheader(); c.alignment = Alignment(horizontal="left")

    ws.column_dimensions[get_column_letter(COL_LABEL)].width = 38
    for j in range(n_periods):
        ws.column_dimensions[get_column_letter(FIRST_DATA_COL + j)].width = 13
    ws.freeze_panes = ws.cell(row=ROW_DATES + 2, column=FIRST_DATA_COL)


def deliver_segments(
    segments: List[dict],
    output_xlsx_path: str,
    arr_factor: float = 12.0,
    company: str = "",
    lookback: int = 12,
) -> str:
    """Build a multi-segment workbook (Change #3): one Corkscrew + one Raw Data
    per segment, plus a Blended Corkscrew with a real (independent) reconciliation.

    Each `segments` entry is a dict:
      {name, long_csv,
       source, source_sheet, source_customer_col, source_first_data_row,
       source_last_data_row, source_first_date_col, source_header_row,
       actuals_through}
    A segment with an Excel `source` gets a verbatim Raw Data sheet and a
    Corkscrew that references it directly (two-sheet style). A segment with only
    a `long_csv` gets a CSV-built Raw Data sheet. All segments must share the same
    comparison-period count so the blend lines up column-for-column."""
    if len(segments) < 2:
        raise ValueError("deliver_segments needs >= 2 segments; use deliver() for one.")

    wb = Workbook()
    # Drop the default sheet; we create named sheets per segment.
    default_ws = wb.active

    seg_specs: List[dict] = []
    period_counts = set()
    blended_company = company

    for seg in segments:
        name = seg["name"]
        cork_name = f"{name} Corkscrew"
        raw_name = f"{name} Raw Data"
        customers, months, cell = load_long_csv(seg["long_csv"])
        actuals_through = seg.get("actuals_through")
        if actuals_through:
            cutoff = parse_month_cutoff(actuals_through)
            months = [m for m in months if m <= cutoff]

        ws_cork = wb.create_sheet(cork_name)
        ws_raw = wb.create_sheet(raw_name)

        source_path = seg.get("source")
        if source_path and not str(source_path).lower().endswith(".csv"):
            source_sheet = seg["source_sheet"]
            _src_wb = load_workbook(source_path, data_only=True)
            src_ws = _src_wb[source_sheet]
            _src_wb_f = load_workbook(source_path, data_only=False)
            src_ws_formulas = _src_wb_f[source_sheet]
            # Carry the FIRST segment's theme (all Metazoa sheets share one theme).
            if wb.loaded_theme is None and _src_wb_f.loaded_theme is not None:
                wb.loaded_theme = _src_wb_f.loaded_theme
            copy_source_sheet_verbatim(source_path, src_ws, ws_raw, src_ws_formulas)

            first_date_col = seg.get("source_first_date_col", "B")
            first_data_row = seg.get("source_first_data_row", 2)
            last_data_row = seg.get("source_last_data_row") or find_source_last_data_row(
                src_ws, seg.get("source_customer_col", "A"), first_data_row)
            months_ts = get_source_months(
                src_ws, first_date_col, header_row=seg.get("source_header_row")) or months
            if actuals_through:
                months_ts = [m for m in months_ts if m <= parse_month_cutoff(actuals_through)]
            first_date_idx = column_index_from_string(first_date_col)
            raw_geometry = {
                "sheet": raw_name,
                "first_row": first_data_row,
                "last_row": last_data_row,
                "first_date_col_idx": first_date_idx,
            }
            write_corkscrew_sheet_aggregating(
                ws_cork, customers=customers, months_analysis=months_ts,
                arr_factor=arr_factor, company=name,
                in_scope_types=["Recurring"], analysis_sheet_name=raw_name,
                lookback=lookback, raw_geometry=raw_geometry)
            months_used = months_ts
        else:
            # CSV-only segment: Raw Data built from CSV (customers at rows 2..,
            # months at col B..). Corkscrew references it directly.
            write_raw_from_csv(ws_raw, customers, months, cell)
            first_data_row = 2
            last_data_row = 2 + len(customers) - 1
            first_date_idx = 2  # col B
            raw_geometry = {
                "sheet": raw_name, "first_row": first_data_row,
                "last_row": last_data_row, "first_date_col_idx": first_date_idx,
            }
            write_corkscrew_sheet_aggregating(
                ws_cork, customers=customers, months_analysis=months,
                arr_factor=arr_factor, company=name,
                in_scope_types=["Recurring"], analysis_sheet_name=raw_name,
                lookback=lookback, raw_geometry=raw_geometry)
            months_used = months

        n_periods = len(months_used) - lookback
        period_counts.add(n_periods)
        seg_specs.append({
            "name": name, "corkscrew_sheet": cork_name, "raw_sheet": raw_name,
            "first_row": first_data_row, "last_row": last_data_row,
            "first_date_col_idx": first_date_idx, "lookback": lookback,
            "months_periods": months_used[lookback:],
            "months_prior": months_used[:n_periods],
        })

    if len(period_counts) != 1:
        raise ValueError(
            f"Segments produce differing comparison-period counts {period_counts}; "
            "they must share the same month range for a column-aligned blend.")
    n_periods = period_counts.pop()

    ws_blended = wb.create_sheet("Blended Corkscrew", 0)  # leftmost
    write_blended_corkscrew(
        ws_blended, seg_specs, n_periods,
        months_periods=seg_specs[0]["months_periods"],
        months_prior=seg_specs[0]["months_prior"],
        arr_factor=arr_factor, company=blended_company)

    wb.remove(default_ws)
    wb.save(output_xlsx_path)
    return output_xlsx_path


# ---------------------------------------------------------------------------
# Self-test harness (Change #1/#2/#3 — TDD)
# ---------------------------------------------------------------------------


def _st_make_themed_source(path: str) -> None:
    """Build a tiny one-row-per-customer source workbook that exercises:
    - a LIVE date-header formula (=EOMONTH) and a computed customer-name formula,
    - a header fill that references a THEME color (accent5) whose source palette
      resolves to magenta #A02B93 (openpyxl's default theme would render teal).
    Used by the verbatim self-tests."""
    import zipfile
    import shutil
    # 1. Build the data + a theme-indexed fill with openpyxl.
    wb = Workbook()
    ws = wb.active
    ws.title = "Seg"
    from openpyxl.styles.colors import Color
    themed = PatternFill("solid", fgColor=Color(theme=8, tint=0.0))  # accent5
    # Header row 1: customer label + 15 month columns (live EOMONTH formulas).
    ws["A1"] = "Customer ID"
    ws["A1"].fill = themed
    ws["B1"] = dt.datetime(2024, 1, 1)
    ws["B1"].fill = themed
    for j in range(1, 15):
        col = get_column_letter(2 + j)
        prev = get_column_letter(2 + j - 1)
        c = ws[f"{col}1"]
        c.value = f"=EOMONTH({prev}1,1)"   # LIVE formula
        c.fill = themed
    # 4 customers, one row each, flat revenue (clean contiguous block).
    revs = [1000, 2000, 1500, 3000]
    for i, rev in enumerate(revs):
        r = 2 + i
        ws[f"A{r}"] = f'="Customer " & (ROW()-1)'   # LIVE formula
        for j in range(15):
            ws.cell(row=r, column=2 + j, value=rev)
    wb.save(path)
    # 2. Rewrite theme1.xml so accent5 = magenta A02B93 (simulating a real
    #    Office theme), AND inject cached <v> values for the formula cells so a
    #    data_only=True load resolves them (exactly as a real Excel save would —
    #    openpyxl can't recalc, so without a cache the discovery scans see None).
    import re as _re
    tmp = path + ".tmp"
    with zipfile.ZipFile(path, "r") as zin:
        names = zin.namelist()
        theme = zin.read("xl/theme/theme1.xml").decode()
        theme2 = _re.sub(r"<a:accent5>.*?</a:accent5>",
                         '<a:accent5><a:srgbClr val="A02B93"/></a:accent5>',
                         theme, flags=_re.S)
        # Patch the sheet XML: add cached values to formula cells.
        sheet_name = next(n for n in names if _re.match(r"xl/worksheets/sheet\d+\.xml$", n))
        sheet = zin.read(sheet_name).decode()

        def _cache_dates(m):
            # <c r="C1" ...><f>=EOMONTH(B1,1)</f></c> -> add Excel serial date <v>
            cell = m.group(0)
            ref = _re.search(r'r="([A-Z]+)(\d+)"', cell)
            col_letters, row = ref.group(1), int(ref.group(2))
            col_idx = column_index_from_string(col_letters)
            # month index 0-based from col B(=2)
            mi = col_idx - 2
            d = dt.date(2024, 1, 1)
            # add mi months
            y = 2024 + (mi // 12)
            mo = 1 + (mi % 12)
            d = dt.date(y, mo, 1)
            serial = (d - dt.date(1899, 12, 30)).days
            return cell.replace("</f>", f"</f><v>{serial}</v>")

        # date header formulas in row 1 (=EOMONTH ...)
        sheet = _re.sub(r'<c r="[A-Z]+1"[^>]*><f>=EOMONTH[^<]*</f></c>',
                        _cache_dates, sheet)

        # customer-name formulas in col A rows 2..5 → cache "Customer N"
        def _cache_names(m):
            cell = m.group(0)
            ref = _re.search(r'r="A(\d+)"', cell)
            n = int(ref.group(1)) - 1
            # inline string cache: set t="str" and add <v>
            cell = cell.replace("<c ", '<c t="str" ', 1) if 't="' not in cell else cell
            return cell.replace("</f>", f"</f><v>Customer {n}</v>")
        sheet = _re.sub(r'<c r="A[2-9]"[^>]*>(?:<f>[^<]*</f>)</c>', _cache_names, sheet)

        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for n in names:
                if n == "xl/theme/theme1.xml":
                    zout.writestr(n, theme2)
                elif n == sheet_name:
                    zout.writestr(n, sheet)
                else:
                    zout.writestr(n, zin.read(n))
    shutil.move(tmp, path)


def _st_resolve_accent5(xlsx_path: str) -> str:
    """Read the saved workbook's theme1.xml and return accent5's srgb hex."""
    import zipfile
    import re as _re
    with zipfile.ZipFile(xlsx_path, "r") as z:
        theme = z.read("xl/theme/theme1.xml").decode()
    m = _re.search(r"<a:accent5>\s*<a:srgbClr val=\"([0-9A-Fa-f]{6})\"", theme)
    return m.group(1).upper() if m else ""


def _find_soffice() -> str | None:
    """Locate a LibreOffice headless binary for recalc, if installed."""
    import shutil
    for cand in ("soffice", "libreoffice",
                 "/Applications/LibreOffice.app/Contents/MacOS/soffice"):
        p = shutil.which(cand) if "/" not in cand else (
            cand if os.path.exists(cand) else None)
        if p:
            return p
    return None


def _st_recalc(xlsx_path: str, outdir: str) -> str | None:
    """Recalculate a workbook with LibreOffice headless and return the path to
    the recalced copy, or None if LibreOffice isn't available."""
    import subprocess
    soffice = _find_soffice()
    if not soffice:
        return None
    sub = os.path.join(outdir, "recalc")
    os.makedirs(sub, exist_ok=True)
    try:
        subprocess.run([soffice, "--headless", "--calc", "--convert-to", "xlsx",
                        "--outdir", sub, xlsx_path],
                       check=True, capture_output=True, timeout=120)
    except Exception:
        return None
    out = os.path.join(sub, os.path.basename(xlsx_path))
    return out if os.path.exists(out) else None


def _self_test() -> int:
    import tempfile
    import os
    import traceback
    failures: List[str] = []

    def check(name, cond, detail=""):
        if cond:
            print(f"  PASS — {name}")
        else:
            print(f"  FAIL — {name}: {detail}")
            failures.append(name)

    tmpd = tempfile.mkdtemp(prefix="deliver_selftest_")

    # ---- Change #1: theme + formula verbatim ----------------------------
    print("[#1] theme + formula verbatim")
    try:
        src = os.path.join(tmpd, "themed_src.xlsx")
        _st_make_themed_source(src)
        check("source theme accent5 is magenta", _st_resolve_accent5(src) == "A02B93",
              _st_resolve_accent5(src))

        # Build a long CSV from the source (4 customers x 15 months, flat).
        # The fixture's month headers are live =EOMONTH formulas with no cached
        # value (openpyxl can't recalc), so derive the month list from the known
        # start (2024-01, monthly) rather than reading cached header values.
        months_dt = [dt.date(2024, 1 + (j % 12) if j < 12 else (j - 12) + 1,
                             1).replace(year=2024 + (j // 12)) for j in range(15)]
        revs = [1000, 2000, 1500, 3000]
        csv_path = os.path.join(tmpd, "seg.csv")
        with open(csv_path, "w") as fh:
            fh.write("customer_id,month,mrr\n")
            for i in range(4):
                cust = f"Customer {i + 1}"
                for j, mdt in enumerate(months_dt):
                    fh.write(f"{cust},{mdt.strftime('%Y-%m')},{revs[i]}\n")

        out = os.path.join(tmpd, "verbatim_out.xlsx")
        deliver(csv_path, out, arr_factor=12.0, company="ThemeTest",
                source_path=src, source_sheet="Seg",
                source_customer_col="A", source_first_data_row=2,
                source_first_date_col="B", source_header_row=1,
                source_last_data_row=5, lookback=12)

        # 1a. Output carries the SOURCE theme (magenta), not openpyxl default.
        check("output theme accent5 carried from source (magenta)",
              _st_resolve_accent5(out) == "A02B93", _st_resolve_accent5(out))

        # 1b. Raw Data preserves LIVE formulas (not flattened to values).
        wb_f = load_workbook(out, data_only=False)
        raw = wb_f["Raw Data"]
        c1 = raw["C1"].value
        check("Raw Data date header stays a live formula",
              isinstance(c1, str) and c1.startswith("=EOMONTH"), repr(c1))
        a2 = raw["A2"].value
        check("Raw Data customer name stays a live formula",
              isinstance(a2, str) and a2.startswith('="Customer'), repr(a2))

        # 1c. Corkscrew explicit-RGB banner untouched by the theme carry.
        cork = wb_f["Corkscrew"]
        title_fill = cork.cell(row=ROW_TITLE, column=1).fill.fgColor
        check("Corkscrew title banner still explicit RGB 1F4E79",
              (title_fill.rgb or "").endswith(TITLE_FILL), repr(title_fill.rgb))
    except Exception:
        failures.append("#1 raised")
        traceback.print_exc()

    # ---- Change #2: two-sheet / no-helper path --------------------------
    print("[#2] two-sheet / no-helper path")
    try:
        # Reuse the clean themed source from #1 (one contiguous row per
        # customer, no aggregation needed). 15 months → lookback 12 → 3 periods.
        src2 = os.path.join(tmpd, "themed_src.xlsx")
        if not os.path.exists(src2):
            _st_make_themed_source(src2)
        csv2 = os.path.join(tmpd, "seg.csv")
        if not os.path.exists(csv2):
            months_dt = [dt.date(2024, 1, 1).replace(
                year=2024 + (j // 12), month=1 + (j % 12)) for j in range(15)]
            with open(csv2, "w") as fh:
                fh.write("customer_id,month,mrr\n")
                for i, rev in enumerate([1000, 2000, 1500, 3000]):
                    for mdt in months_dt:
                        fh.write(f"Customer {i + 1},{mdt.strftime('%Y-%m')},{rev}\n")

        out2 = os.path.join(tmpd, "two_sheet_out.xlsx")
        deliver(csv2, out2, arr_factor=12.0, company="TwoSheet",
                source_path=src2, source_sheet="Seg",
                source_customer_col="A", source_first_data_row=2,
                source_first_date_col="B", source_header_row=1,
                source_last_data_row=5, lookback=12, two_sheet=True)

        wb2 = load_workbook(out2, data_only=False)
        check("two-sheet output has exactly [Corkscrew, Raw Data]",
              wb2.sheetnames == ["Corkscrew", "Raw Data"], wb2.sheetnames)
        cork2 = wb2["Corkscrew"]
        chk_formula = cork2.cell(row=ROW_CHECK, column=FIRST_DATA_COL).value
        check("external check references 'Raw Data' directly",
              isinstance(chk_formula, str) and "'Raw Data'!" in chk_formula,
              repr(chk_formula))
        beg_formula = cork2.cell(row=ROW_BEGIN, column=FIRST_DATA_COL).value
        check("Beginning ARR references 'Raw Data' directly (no helper)",
              isinstance(beg_formula, str) and "'Raw Data'!" in beg_formula
              and "Analysis" not in beg_formula, repr(beg_formula))

        # Gate: aggregating source + two_sheet must REFUSE (raise).
        gate_ok = False
        try:
            deliver(csv2, os.path.join(tmpd, "gate.xlsx"), arr_factor=12.0,
                    source_path=src2, source_sheet="Seg", source_type_col="Z",
                    two_sheet=True)
        except ValueError:
            gate_ok = True
        check("two_sheet refused when aggregating (type col present)", gate_ok)

        # Recalc with LibreOffice (if available) → external check == 0 every period.
        recalc = _st_recalc(out2, tmpd)
        if recalc:
            wbv = load_workbook(recalc, data_only=True)
            corkv = wbv["Corkscrew"]
            n_per = 3
            checks = [corkv.cell(row=ROW_CHECK, column=FIRST_DATA_COL + j).value
                      for j in range(n_per)]
            allzero = all(abs(v or 0) < 1e-6 for v in checks)
            check("recalc: external check == 0 every period (two-sheet)",
                  allzero, str(checks))
            # Ending for last period: flat revenue → ending == sum(rev)*12.
            end_last = corkv.cell(row=ROW_END, column=FIRST_DATA_COL + n_per - 1).value
            expected = (1000 + 2000 + 1500 + 3000) * 12
            check("recalc: two-sheet Ending matches expected",
                  abs((end_last or 0) - expected) < 1e-6,
                  f"{end_last} vs {expected}")
        else:
            print("  SKIP — LibreOffice not available; recalc checks skipped")
    except Exception:
        failures.append("#2 raised")
        traceback.print_exc()

    # ---- Change #3: multi-segment + blended reconciliation --------------
    print("[#3] multi-segment + blended corkscrew")
    try:
        # Two clean segments, each a one-row-per-customer themed source sheet.
        months_dt = [dt.date(2024, 1, 1).replace(
            year=2024 + (j // 12), month=1 + (j % 12)) for j in range(15)]

        def _make_seg_source(path, revs):
            wbx = Workbook(); wsx = wbx.active; wsx.title = "Seg"
            wsx["A1"] = "Customer ID"
            for j in range(15):
                wsx.cell(row=1, column=2 + j, value=months_dt[j])
            for i, rev in enumerate(revs):
                wsx.cell(row=2 + i, column=1, value=f"Customer {i + 1}")
                for j in range(15):
                    wsx.cell(row=2 + i, column=2 + j, value=rev)
            wbx.save(path)

        def _make_seg_csv(path, revs):
            with open(path, "w") as fh:
                fh.write("customer_id,month,mrr\n")
                for i, rev in enumerate(revs):
                    for mdt in months_dt:
                        fh.write(f"Customer {i + 1},{mdt.strftime('%Y-%m')},{rev}\n")

        seg_a_src = os.path.join(tmpd, "seg_a.xlsx"); _make_seg_source(seg_a_src, [1000, 2000, 1500])
        seg_b_src = os.path.join(tmpd, "seg_b.xlsx"); _make_seg_source(seg_b_src, [500, 800])
        seg_a_csv = os.path.join(tmpd, "seg_a.csv"); _make_seg_csv(seg_a_csv, [1000, 2000, 1500])
        seg_b_csv = os.path.join(tmpd, "seg_b.csv"); _make_seg_csv(seg_b_csv, [500, 800])

        segments = [
            {"name": "Core", "long_csv": seg_a_csv, "source": seg_a_src,
             "source_sheet": "Seg", "source_customer_col": "A",
             "source_first_data_row": 2, "source_last_data_row": 4,
             "source_first_date_col": "B", "source_header_row": 1},
            {"name": "PLG", "long_csv": seg_b_csv, "source": seg_b_src,
             "source_sheet": "Seg", "source_customer_col": "A",
             "source_first_data_row": 2, "source_last_data_row": 3,
             "source_first_date_col": "B", "source_header_row": 1},
        ]
        out3 = os.path.join(tmpd, "blended_out.xlsx")
        deliver_segments(segments, out3, arr_factor=12.0, company="Metazoa",
                         lookback=12)

        wb3 = load_workbook(out3, data_only=False)
        names = wb3.sheetnames
        check("per-segment corkscrews present",
              "Core Corkscrew" in names and "PLG Corkscrew" in names, names)
        check("blended corkscrew present", "Blended Corkscrew" in names, names)
        check("per-segment Raw Data present",
              "Core Raw Data" in names and "PLG Raw Data" in names, names)
        blended = wb3["Blended Corkscrew"]
        beg_f = blended.cell(row=ROW_BEGIN, column=FIRST_DATA_COL).value
        check("blended Beginning sums segment corkscrews (cross-sheet refs)",
              isinstance(beg_f, str) and "Corkscrew'!" in beg_f, repr(beg_f))

        recalc = _st_recalc(out3, tmpd)
        if recalc:
            wbv = load_workbook(recalc, data_only=True)
            bl = wbv["Blended Corkscrew"]
            n_per = 3
            variances = [bl.cell(row=ROW_CHECK, column=FIRST_DATA_COL + j).value
                         for j in range(n_per)]
            check("recalc: blended variance == 0 every period",
                  all(abs(v or 0) < 1e-6 for v in variances), str(variances))
            # Blended ending = sum of both segments' totals × 12.
            end_last = bl.cell(row=ROW_END, column=FIRST_DATA_COL + n_per - 1).value
            expected = (1000 + 2000 + 1500 + 500 + 800) * 12
            check("recalc: blended Ending = sum of segment endings",
                  abs((end_last or 0) - expected) < 1e-6,
                  f"{end_last} vs {expected}")
            # Cross-check each segment ending too.
            core = wbv["Core Corkscrew"]
            core_end = core.cell(row=ROW_END, column=FIRST_DATA_COL + n_per - 1).value
            check("recalc: Core segment Ending correct",
                  abs((core_end or 0) - (1000 + 2000 + 1500) * 12) < 1e-6,
                  str(core_end))
        else:
            print("  SKIP — LibreOffice not available; blended recalc skipped")
    except Exception:
        failures.append("#3 raised")
        traceback.print_exc()

    print("=" * 60)
    if failures:
        print(f"SELF-TEST: FAIL ({len(failures)} failing): {failures}")
        return 1
    print("SELF-TEST: PASS")
    return 0


def _parse_segment_spec(spec: str) -> dict:
    """Parse a --segment SPEC into a segment dict.
    'Name=long.csv' or
    'Name=long.csv:src.xlsx:Sheet:custCol:firstRow:lastRow:firstDateCol:headerRow'.
    Trailing source params are optional (left to right)."""
    if "=" not in spec:
        raise ValueError(f"--segment {spec!r} missing 'Name='. Use 'Name=long.csv[:...]'.")
    name, rest = spec.split("=", 1)
    parts = rest.split(":")
    seg = {"name": name.strip(), "long_csv": parts[0]}
    keys = ["source", "source_sheet", "source_customer_col",
            "source_first_data_row", "source_last_data_row",
            "source_first_date_col", "source_header_row"]
    for k, v in zip(keys, parts[1:]):
        if v == "":
            continue
        if k in ("source_first_data_row", "source_last_data_row", "source_header_row"):
            seg[k] = int(v)
        else:
            seg[k] = v
    return seg


def parse_args(argv):
    p = argparse.ArgumentParser(description="Retention-analysis Phase 5.")
    p.add_argument("long_csv", nargs="?", default=None,
                   help="Long-format CSV (omit when using >= 2 --segment specs).")
    p.add_argument("output_xlsx")
    p.add_argument("--arr-factor", type=float, default=12.0,
                   help="MRR->ARR multiplier (12 for MRR input, 1 for ARR input)")
    p.add_argument("--compute-json", default=None,
                   help="OPTIONAL compute.py output. Not required — deliver builds "
                        "independently from the CSV. If given, its arr_factor "
                        "overrides --arr-factor.")
    p.add_argument("--company", default="")
    p.add_argument("--config", default=None,
                   help="JSON written by survey.py --emit-config. Fills the source-* "
                        "options and --actuals-through so survey's findings flow "
                        "straight in; any explicit flag still overrides it.")
    p.add_argument("--source", default=None)
    p.add_argument("--source-sheet", default=None)
    p.add_argument("--source-customer-col", default="A")
    p.add_argument("--source-first-data-row", type=int, default=2)
    p.add_argument("--source-last-data-row", type=int, default=None,
                   help="Last row of the customer block in the source (#3). Caps the "
                        "scan so a summary/total block below the customers isn't "
                        "treated as customers. Default: scan to the last non-empty row.")
    p.add_argument("--actuals-through", default=None,
                   help="Last COMPLETE actual month, e.g. '2026-05' or 'May-26' (#2). "
                        "Drops the in-progress current month and any forecast tail so "
                        "projections aren't counted as retention. Feed survey.py's "
                        "actuals_through here.")
    p.add_argument("--source-first-date-col", default="B")
    p.add_argument("--source-header-row", type=int, default=None,
                   help="Explicit row number of the date-header row "
                        "(default: try row above first-data-row)")
    p.add_argument("--source-type-col", default=None,
                   help="Column letter of revenue-type column (triggers "
                        "aggregating mode)")
    p.add_argument("--type-filter", default=None,
                   help="Comma-separated list of in-scope types "
                        "(default: 'Recurring,Re-occurring')")
    p.add_argument("--lookback", type=int, default=12)
    p.add_argument("--two-sheet", action="store_true",
                   help="Change #2: build a 2-sheet deliverable (Corkscrew + Raw "
                        "Data, no helper). Opt-in; ONLY pass after survey.py "
                        "confirms a clean contiguous one-row-per-customer block "
                        "(customer_row_range.contiguous, no section rows inside) "
                        "with a single in-scope revenue type. The Corkscrew then "
                        "references Raw Data directly.")
    p.add_argument("--segment", action="append", default=None, metavar="SPEC",
                   help="Change #3: declare a segment for a multi-segment "
                        "deliverable. Repeatable (one per segment). SPEC is "
                        "'Name=long.csv' optionally followed by colon-separated "
                        "source params: "
                        "'Name=long.csv:src.xlsx:Sheet:custCol:firstRow:lastRow:firstDateCol:headerRow'. "
                        "NOTE: colon-delimited specs break if a path/sheet name "
                        "contains a ':' — prefer --segment-config (JSON) for "
                        "anything non-trivial. With 2+ segments deliver_segments() "
                        "runs; output_xlsx is the destination. Produces a Corkscrew "
                        "+ Raw Data per segment plus a Blended Corkscrew (blended "
                        "variance = 0 every period).")
    p.add_argument("--segment-config", default=None, metavar="JSON",
                   help="Change #3 (preferred): path to a JSON file holding a list "
                        "of segment dicts (name, long_csv, source, source_sheet, "
                        "source_customer_col, source_first_data_row, "
                        "source_last_data_row, source_first_date_col, "
                        "source_header_row, actuals_through). Unambiguous for paths "
                        "and sheet names with spaces or colons. Overrides --segment.")
    # Apply survey.py --emit-config values as defaults; explicit CLI flags win.
    pre, _ = p.parse_known_args(argv)
    if pre.config:
        with open(pre.config, "r", encoding="utf-8") as fh:
            cfg = json.load(fh)
        p.set_defaults(**{k: v for k, v in cfg.items() if v is not None})
    return p.parse_args(argv)


def main(argv=None):
    import time
    _t0 = time.perf_counter()
    raw_argv = argv if argv is not None else sys.argv[1:]
    if "--self-test" in raw_argv:
        return _self_test()
    args = parse_args(raw_argv)

    # Change #3 — multi-segment route. JSON config wins over colon specs.
    segments = None
    if args.segment_config:
        with open(args.segment_config, "r", encoding="utf-8") as fh:
            segments = json.load(fh)
    elif args.segment and len(args.segment) >= 2:
        segments = [_parse_segment_spec(s) for s in args.segment]
    if segments and len(segments) >= 2:
        for seg in segments:
            seg.setdefault("actuals_through", args.actuals_through)
        out_xlsx = deliver_segments(
            segments, args.output_xlsx, arr_factor=args.arr_factor,
            company=args.company, lookback=args.lookback)
        print(f"Wrote: {out_xlsx}")
        print(f"[deliver.py] built {len(segments)}-segment workbook in "
              f"{time.perf_counter() - _t0:.2f}s", file=sys.stderr)
        return 0
    if args.segment and len(args.segment) == 1:
        raise SystemExit("--segment needs >= 2 specs for a blended workbook; "
                         "use the single-source flags for one segment.")
    if args.long_csv is None:
        raise SystemExit("long_csv is required unless >= 2 --segment specs are given.")

    type_filter = (
        [t.strip() for t in args.type_filter.split(",")]
        if args.type_filter else None
    )
    out_xlsx = deliver(
        args.long_csv, args.output_xlsx,
        arr_factor=args.arr_factor,
        compute_json_path=args.compute_json,
        company=args.company,
        source_path=args.source,
        source_sheet=args.source_sheet,
        source_customer_col=args.source_customer_col,
        source_first_data_row=args.source_first_data_row,
        source_first_date_col=args.source_first_date_col,
        source_type_col=args.source_type_col,
        type_filter=type_filter,
        lookback=args.lookback,
        source_header_row=args.source_header_row,
        source_last_data_row=args.source_last_data_row,
        actuals_through=args.actuals_through,
        two_sheet=args.two_sheet,
    )
    print(f"Wrote: {out_xlsx}")
    print(f"[deliver.py] built workbook in {time.perf_counter() - _t0:.2f}s", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
