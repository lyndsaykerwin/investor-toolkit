#!/usr/bin/env python3
"""
detect_unit.py — decide whether a customer-revenue grid is ARR, MRR, annual
revenue, or transaction/invoice-level, and locate the customer column + the
current (latest) recurring-revenue period.

THE HARD-WON RULE (encoded here as priority order):
    Read the workbook's own LABELS first — sheet name, title/header text in the
    top rows, the value-column header. A tab named "...ARR" or a header saying
    "annual recurring" means the values are ALREADY ANNUAL; do NOT multiply by
    12. Dollar MAGNITUDE is the LAST resort and only a tiebreaker — it may
    CONFIRM a label but must never OVERRIDE one. (A real deal once had a tab
    literally named "...ARR" with an "annual recurring" header, yet a
    magnitude-only heuristic saw a ~$13.5K median, called it MRR, and annualized
    it 12x — overstating ARR by an order of magnitude. Labels first, always.)

Verdicts:
    ARR              — values are already annual recurring revenue (factor 1)
    MRR              — values are monthly recurring revenue (annualize x12 for ARR)
    ANNUAL_REVENUE   — annual revenue buckets (yearly columns); ARR-scale, factor 1
    TRANSACTIONAL    — invoice/transaction-level revenue, NOT normalized recurring;
                       concentration needs a period aggregation — confirm scope
    AMBIGUOUS        — no label signal; magnitude-only hypothesis, MUST confirm

Every verdict carries `evidence` (the three signals, reported separately) and a
`confidence` ("decisive" = from a label; "hypothesis" = magnitude tiebreaker).
Nothing here writes files. Run with --json for structured output, --self-test
to check against the bundled fixture.

Usage:
    python3 detect_unit.py <path.xlsx|.csv> [--sheet NAME] [--json]
    python3 detect_unit.py --self-test
"""
import argparse
import csv
import json
import re
import sys
from statistics import median

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    raise

# --- token vocabularies -----------------------------------------------------
ARR_TOKENS = [r"\barr\b", r"annual recurring", r"annualized recurring",
              r"annual contract", r"\bacv\b"]
MRR_TOKENS = [r"\bmrr\b", r"\bcmrr\b", r"monthly recurring", r"\bmonthly\b"]
# Headers that signal transaction/invoice-level (not normalized recurring).
TXN_TOKENS = [r"\binvoice", r"\bgross\b", r"\bamount\b", r"\bbilling",
              r"transaction", r"\bsales\b", r"\bpayment"]
SUMMARY_LABEL = re.compile(
    r"^(grand\s+)?(total|subtotal|sum|average|avg|count)\b", re.I)


def _norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _any(patterns, text):
    t = _norm(text)
    return [p for p in patterns if re.search(p, t)]


# --- CSV path ---------------------------------------------------------------
def _read_csv_header_and_samples(path):
    with open(path, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        return [], [], []
    header = [c.strip() for c in rows[0]]
    body = rows[1:]
    return header, body, rows


def detect_csv(path):
    header, body, _ = _read_csv_header_and_samples(path)
    header_text = " | ".join(header)
    arr_hits = _any(ARR_TOKENS, header_text)
    mrr_hits = _any(MRR_TOKENS, header_text)
    txn_hits = _any(TXN_TOKENS, header_text)

    # find a numeric "value" column: prefer one whose header carries a recurring
    # token, else the first mostly-numeric column.
    value_col_idx, value_col_name = None, None
    for pri in (MRR_TOKENS + ARR_TOKENS):
        for i, h in enumerate(header):
            if re.search(pri, _norm(h)):
                value_col_idx, value_col_name = i, h
                break
        if value_col_idx is not None:
            break
    if value_col_idx is None:
        for i, h in enumerate(header):
            nums = 0
            for r in body[:200]:
                if i < len(r):
                    try:
                        float(str(r[i]).replace(",", "").replace("$", ""))
                        nums += 1
                    except ValueError:
                        pass
            if nums > 100:
                value_col_idx, value_col_name = i, h
                break

    samples = []
    if value_col_idx is not None:
        for r in body:
            if value_col_idx < len(r):
                try:
                    v = float(str(r[value_col_idx]).replace(",", "").replace("$", ""))
                    if v > 0:
                        samples.append(v)
                except ValueError:
                    pass

    # CSV has no sheet name — header tokens count as title-level signals.
    verdict, conf, why = _decide([], [], arr_hits, mrr_hits, txn_hits, samples,
                                 has_period_grid=False)
    return {
        "source_type": "csv",
        "value_column": value_col_name,
        "layout": "long/tidy (one row per customer; snapshot value column)",
        "verdict": verdict,
        "confidence": conf,
        "signals": {
            "label_arr": arr_hits,
            "label_mrr": mrr_hits,
            "label_txn": txn_hits,
            "value_median": round(median(samples), 2) if samples else None,
            "value_samples": [round(x, 2) for x in samples[:8]],
        },
        "reasoning": why,
    }


# --- XLSX path --------------------------------------------------------------
def _pick_sheet(wb, want):
    if want and want in wb.sheetnames:
        return wb[want]
    # score by name token: prefer "raw"/"customer"/"mrr"/"arr"; avoid "pivot"/"key"
    best, best_score = None, -1
    for name in wb.sheetnames:
        n = _norm(name)
        score = 0
        for kw, w in (("raw", 3), ("customer", 3), ("mrr", 2), ("arr", 2),
                      ("revenue", 1), ("appended", 1)):
            if kw in n:
                score += w
        for kw in ("pivot", "key", "cache", "rollforward", "summary", "dashboard"):
            if kw in n:
                score -= 2
        if score > best_score:
            best, best_score = name, score
    return wb[best] if best else wb[wb.sheetnames[0]]


def detect_xlsx(path, want_sheet=None):
    """Single sequential pass (read_only-safe — no ws.max_row/random ws.cell,
    which are unreliable/slow in read_only mode and crash when a file omits its
    stored dimensions)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = _pick_sheet(wb, want_sheet)
    sheet_name = ws.title

    rows_iter = ws.iter_rows(values_only=True)
    BUF = 15
    buf = []
    for _ in range(BUF):
        try:
            buf.append(next(rows_iter))
        except StopIteration:
            break

    # header row = the buffered row with the most filled cells (titles/banners
    # above it are typically 1-2 cells); 1-based index.
    header_idx, best_filled = 0, -1
    for i, row in enumerate(buf):
        filled = sum(1 for v in (row or ()) if v not in (None, ""))
        if filled > best_filled:
            best_filled, header_idx = filled, i
    header_row = header_idx + 1
    header_vals = buf[header_idx] if buf else ()

    # title zone = all strings in rows 1..header_row
    chunks = []
    for row in buf[:header_idx + 1]:
        for v in (row or ()):
            if isinstance(v, str) and v.strip():
                chunks.append(v.strip())
    title_text = " | ".join(chunks)

    sheet_arr = _any(ARR_TOKENS, sheet_name)
    sheet_mrr = _any(MRR_TOKENS, sheet_name)
    title_arr = _any(ARR_TOKENS, title_text)
    title_mrr = _any(MRR_TOKENS, title_text)
    title_txn = _any(TXN_TOKENS, title_text)
    arr_hits = sorted(set(sheet_arr + title_arr))
    mrr_hits = sorted(set(sheet_mrr + title_mrr))

    # latest-period column from the header row: prefer date-like (skip per-row
    # summary/total columns); else the most-numeric column seen in the buffer.
    date_idx = [i for i, v in enumerate(header_vals or ()) if _looks_date(v)]
    summary_idx = {i for i, v in enumerate(header_vals or ())
                   if isinstance(v, str) and SUMMARY_LABEL.match(v.strip())}
    if date_idx:
        keep = [i for i in date_idx if i not in summary_idx] or date_idx
        latest_idx = max(keep)
        layout = "wide (customers in rows, periods in columns)"
    else:
        num_count = {}
        for row in buf[header_idx + 1:]:
            for i, v in enumerate(row or ()):
                if isinstance(v, (int, float)) and v != 0 and i not in summary_idx:
                    num_count[i] = num_count.get(i, 0) + 1
        latest_idx = max(num_count, key=num_count.get) if num_count else None
        layout = "long/tidy or single-value column"
    latest_label = (header_vals[latest_idx]
                    if latest_idx is not None and header_vals
                    and latest_idx < len(header_vals) else None)

    # sample the latest column: finish the buffer, then continue the generator.
    samples = []

    def _grab(row):
        if latest_idx is not None and row and latest_idx < len(row):
            v = row[latest_idx]
            if isinstance(v, (int, float)) and v > 0:
                samples.append(float(v))
    for row in buf[header_idx + 1:]:
        _grab(row)
    for row in rows_iter:
        if len(samples) >= 500:
            break
        _grab(row)

    yearly = bool(date_idx) and all(
        isinstance(header_vals[i], (str, int)) and re.match(r"^\d{4}$", str(header_vals[i]).strip())
        for i in date_idx) and len(date_idx) >= 2
    verdict, conf, why = _decide(sheet_arr, sheet_mrr, title_arr, title_mrr,
                                 title_txn, samples,
                                 has_period_grid=bool(date_idx), yearly=yearly)
    wb.close()
    latest_col = (latest_idx + 1) if latest_idx is not None else None
    return {
        "source_type": "xlsx",
        "sheet": sheet_name,
        "header_row": header_row,
        "latest_period_col": get_column_letter(latest_col) if latest_col else None,
        "latest_period_label": str(latest_label) if latest_label is not None else None,
        "layout": layout,
        "verdict": verdict,
        "confidence": conf,
        "signals": {
            "label_arr_sheet": sheet_arr,
            "label_mrr_sheet": sheet_mrr,
            "label_arr_title": title_arr,
            "label_mrr_title": title_mrr,
            "label_txn_title": title_txn,
            "value_median": round(median(samples), 2) if samples else None,
            "value_samples": [round(x, 2) for x in samples[:8]],
            "title_zone": title_text[:200],
        },
        "reasoning": why,
    }


def _looks_date(v):
    import datetime as dt
    if isinstance(v, (dt.date, dt.datetime)):
        return True
    if not isinstance(v, str):
        return False
    s = v.strip().lower()
    if re.match(r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\s\-/]*\d", s):
        return True
    if re.match(r"^\d{4}$", s) and 1990 <= int(s) <= 2099:
        return True
    if re.match(r"^q[1-4][\s\-]?\d", s):
        return True
    if re.match(r"^\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}$", s):
        return True
    return False


# --- the decision -----------------------------------------------------------
def _decide(sheet_arr, sheet_mrr, title_arr, title_mrr, txn_hits, samples,
            has_period_grid, yearly=False):
    med = median(samples) if samples else None
    arr_hits = sorted(set(sheet_arr + title_arr))
    mrr_hits = sorted(set(sheet_mrr + title_mrr))

    def arr(src):
        return ("ARR", "decisive",
                f"{src} says annual recurring → values are ARR, factor 1, do NOT "
                f"annualize." + (f" (median sample {med:,.0f} — magnitude is NOT "
                f"consulted; the label decides.)" if med else ""))

    def mrr(src):
        return ("MRR", "decisive",
                f"{src} says monthly recurring → values are MRR; annualize x12 to "
                f"show ARR." + (f" (median sample {med:,.0f})" if med else ""))

    # 0. SHEET NAME is the strongest signal — it disambiguates outright, even if
    #    a stray 'arr'/'mrr' token appears in a masked cell in the title zone.
    #    (The real-world miss this skill exists to prevent hinged on a sheet
    #    literally named '...ARR' being overridden by magnitude.)
    if sheet_mrr and not sheet_arr:
        return mrr(f"Sheet name {sheet_mrr}")
    if sheet_arr and not sheet_mrr:
        return arr(f"Sheet name {sheet_arr}")

    # 1. Otherwise combined labels (sheet + title) decide when unambiguous.
    if arr_hits and not mrr_hits:
        return arr(f"Label(s) {arr_hits}")
    if mrr_hits and not arr_hits:
        return mrr(f"Label(s) {mrr_hits}")
    if arr_hits and mrr_hits:
        # both present, sheet name didn't disambiguate (e.g. 'Total MRR' AND
        # 'Total ARR' headers) — magnitude breaks the tie but flag to confirm.
        guess = "ARR" if (med and med >= 25_000) else "MRR"
        return (guess, "hypothesis",
                f"Both ARR {arr_hits} and MRR {mrr_hits} labels present and the sheet "
                f"name didn't disambiguate; per-customer value matches {guess} by "
                f"magnitude (median {med:,.0f}) — CONFIRM which column the grid pulls "
                f"from before annualizing.")

    # 2. No recurring label. Yearly buckets → annual revenue (ARR-scale).
    if yearly:
        return ("ANNUAL_REVENUE", "decisive",
                "Period headers are calendar years → annual revenue buckets; "
                "ARR-scale, factor 1, do NOT annualize.")

    # 3. Transaction/invoice tokens with no recurring label → transactional.
    if txn_hits and not has_period_grid:
        return ("TRANSACTIONAL", "decisive",
                f"Value label(s) {txn_hits} are invoice/transaction-level with no "
                f"recurring token → NOT normalized ARR/MRR. Customer concentration "
                f"needs a period aggregation (e.g. trailing-12-month revenue per "
                f"customer) — confirm scope with the user before building.")

    # 4. No label at all → magnitude HYPOTHESIS only, must confirm.
    if med is not None:
        guess = "ARR" if med >= 25_000 else "MRR"
        return (f"{guess}", "hypothesis",
                f"No ARR/MRR label found anywhere (sheet name, title, headers). "
                f"Magnitude-only guess: median per-customer value {med:,.0f} "
                f"{'≥' if guess=='ARR' else '<'} 25,000 → {guess} (hypothesis). "
                f"MUST confirm with the user — magnitude alone is not decisive.")
    return ("AMBIGUOUS", "hypothesis",
            "No recurring label and no numeric samples found — confirm the value "
            "column and unit with the user.")


def _column_label(verdict, confidence):
    """The honest column label for build_concentration.py --unit. Only commit to
    'ARR'/'MRR' when the source unit is CERTAIN (a decisive label); otherwise use
    the neutral 'Run-Rate' and never annualize. Concentration %s are identical
    regardless, so the label is purely about not mis-stating the dollar scale."""
    if confidence == "decisive" and verdict == "ARR":
        return "ARR"
    if confidence == "decisive" and verdict == "MRR":
        return "MRR"
    return "Run-Rate"


def detect(path, want_sheet=None):
    res = detect_csv(path) if path.lower().endswith(".csv") \
        else detect_xlsx(path, want_sheet)
    res["column_label"] = _column_label(res["verdict"], res["confidence"])
    return res


# --- self-test --------------------------------------------------------------
def _self_test():
    import os
    here = os.path.dirname(os.path.abspath(__file__))
    fx = os.path.join(here, "..", "fixtures", "example_saas_arr.xlsx")
    fx = os.path.abspath(fx)
    if not os.path.exists(fx):
        print("FAIL: fixture missing — run make_fixture.py first:", fx)
        return 1
    res = detect(fx)
    ok = res["verdict"] == "ARR" and res["confidence"] == "decisive"
    print(json.dumps(res, indent=2))
    print("SELF-TEST:", "PASS" if ok else "FAIL "
          "(expected decisive ARR from the 'Annual Recurring Revenue' label)")
    return 0 if ok else 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path", nargs="?")
    ap.add_argument("--sheet")
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()
    if a.self_test:
        sys.exit(_self_test())
    if not a.path:
        ap.error("path required (or --self-test)")
    res = detect(a.path, a.sheet)
    if a.json:
        print(json.dumps(res, indent=2))
    else:
        print(f"VERDICT: {res['verdict']}  ({res['confidence']})")
        print(f"  → column label for build (--unit): {res['column_label']}  "
              f"(values are NEVER annualized; label only)")
        print(f"  {res['reasoning']}")
        print(f"  layout: {res['layout']}")
        if res["source_type"] == "xlsx":
            print(f"  sheet: {res['sheet']}  header_row: {res['header_row']}  "
                  f"latest period: {res['latest_period_label']} "
                  f"(col {res['latest_period_col']})")
        print("  signals:")
        for k, v in res["signals"].items():
            print(f"    {k}: {v}")


if __name__ == "__main__":
    main()
