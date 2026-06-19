#!/usr/bin/env python3
"""
Render a theme's Market Map JSONL into a banker-grade .xlsx.

The JSONL is the working database (one record per company, first record _theme_meta).
This script is the deterministic render step: JSONL in -> styled Excel out. It builds the
sheet dynamically, so it handles any number of companies (not a fixed row count), and applies
the house styling from templates/market_map_template.manifest.json.

Usage:
    python render_market_map.py --in <landscape.jsonl> --out <Market_Map.xlsx> \
        [--firm "Firm Name"] [--date MM/DD/YYYY]
    python render_market_map.py --selftest

Companies are placed into three sections. A record can force its section with
"map_section": "prospects" | "leaders" | "funded"; otherwise it is inferred:
  - public ticker / Pass-OverFunded / category landmark -> Market Leaders
  - has a funding round or >= $20M raised               -> Funded Companies
  - everything else (targets, bootstrapped)             -> Firm Fit Prospects
"""

import argparse, hashlib, json, math, re, sys, tempfile, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- house style (from market_map_template.manifest.json) ----
NAVY, MIDBLUE, SOFTBLUE = "FF0E2841", "FF1F4E78", "FFD9E1F2"
WHITE, BLUE_INPUT, GREY = "FFFFFFFF", "FF0000FF", "FF595959"
COL_WIDTHS = {"A": 3.0, "B": 24.0, "C": 22.0, "D": 11.0, "E": 14.0,
              "F": 14.0, "G": 13.0, "H": 38.0, "I": 50.0, "J": 20.0, "K": 3.0}
THIN = Side(style="thin", color="FF000000")
MED = Side(style="medium", color="FF000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def font(size=10, bold=False, color="FF000000", italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def fill(color):
    return PatternFill(fill_type="solid", fgColor=color)


def put(ws, coord, value, *, f=None, bg=None, align="left", border=None, numfmt=None, wrap=None):
    c = ws[coord]
    c.value = value
    c.font = f or font()
    if bg:
        c.fill = fill(bg)
    # Wrap long body text by default; pass wrap=False for titles/labels that must stay on one line.
    do_wrap = (align == "left") if wrap is None else wrap
    c.alignment = Alignment(horizontal=align, vertical=("top" if do_wrap else "center"),
                            wrap_text=do_wrap)
    if border:
        c.border = border
    if numfmt:
        c.number_format = numfmt
    return c


# ---- value helpers ----
def g(rec, key):
    v = rec.get(key)
    return "" if v is None else v


def money(v):
    if not v:
        return ""
    v = float(v)
    if v >= 1e9:
        return f"${v/1e9:.1f}B"
    if v >= 1e6:
        return f"${v/1e6:.1f}M"
    return f"${v:,.0f}"


def funding_disp(v):
    """Funding cell: confirmed $0 reads as 'Bootstrapped' (not blank); None stays blank."""
    if v == 0:
        return "Bootstrapped"
    return money(v)


def headcount(rec):
    """Exact FTE if known, else a published band via headcount_display, else blank."""
    v = rec.get("fte")
    if v is not None:
        return v
    return g(rec, "headcount_display")


def ticker(rec):
    m = re.search(r"public\s*\(([^)]+)\)", str(rec.get("ownership_status", "")), re.I)
    return m.group(1) if m else ""


# header, extractor, alignment, number-format
PROSPECTS = [
    ("Company", lambda r: g(r, "name"), "left", None),
    ("HQ Location", lambda r: g(r, "hq"), "left", None),
    ("Year Founded", lambda r: g(r, "founded_year"), "center", "0"),
    ("Headcount", headcount, "center", "#,##0"),
    ("Total Funding", lambda r: funding_disp(r.get("total_funding_usd")), "center", None),
    ("ARR $M (if known)", lambda r: g(r, "arr_m"), "right", None),
    ("Product / Positioning", lambda r: g(r, "what_they_do"), "left", None),
    ("Fit Rationale", lambda r: g(r, "firm_fit_rationale"), "left", None),
    ("Outreach Status", lambda r: g(r, "outreach_status") or "Not yet contacted", "left", None),
]
LEADERS = [
    ("Company", lambda r: g(r, "name"), "left", None),
    ("HQ Location", lambda r: g(r, "hq"), "left", None),
    ("Year Founded", lambda r: g(r, "founded_year"), "center", "0"),
    ("Ticker", ticker, "center", None),
    ("Status", lambda r: g(r, "ownership_status"), "center", None),
    ("Mkt Cap / Val.", lambda r: g(r, "market_cap") or money(r.get("total_funding_usd")), "right", None),
    ("Product / Positioning", lambda r: g(r, "what_they_do"), "left", None),
    ("Strategic Relevance", lambda r: g(r, "firm_fit_rationale") or g(r, "strategic_relevance"), "left", None),
]
FUNDED = [
    ("Company", lambda r: g(r, "name"), "left", None),
    ("HQ Location", lambda r: g(r, "hq"), "left", None),
    ("Year Founded", lambda r: g(r, "founded_year"), "center", "0"),
    ("Last Round", lambda r: g(r, "last_round"), "center", None),
    ("Total Raised", lambda r: funding_disp(r.get("total_funding_usd")), "right", None),
    ("Lead Investors", lambda r: g(r, "lead_investors"), "left", None),
    ("Product / Positioning", lambda r: g(r, "what_they_do"), "left", None),
    ("Observations", lambda r: g(r, "firm_fit_rationale") or g(r, "source_note"), "left", None),
]
def classify(rec):
    s = rec.get("map_section")
    if s in ("prospects", "leaders", "funded"):
        return s
    own = str(rec.get("ownership_status", "")).lower()
    ff = str(rec.get("firm_fit", "")).lower()
    tags = str(rec.get("tags", "")).lower()
    if "public" in own or "overfunded" in ff or "landmark" in tags:
        return "leaders"
    # Explicit non-targets never belong in Firm Fit Prospects: anything marked Pass-*,
    # the portfolio platform itself, or an already-acquired/subsidiary company.
    if ff.startswith("pass") or "portfolio" in ff:
        return "leaders"
    if "acquired" in own or "subsidiary" in own:
        return "leaders"
    if rec.get("last_round") or (rec.get("total_funding_usd") or 0) >= 20_000_000:
        return "funded"
    return "prospects"


def load(path):
    meta, companies = {}, []
    for line in Path(path).read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        rec = json.loads(line)
        if any(k.startswith("_template") for k in rec):
            continue
        if rec.get("_theme_meta"):
            meta = rec
        elif rec.get("type", "company") in ("company", ""):
            companies.append(rec)
    return meta, companies


def section(ws, row, title, subtitle, cols, records):
    """Write one section starting at `row`; return the next free row."""
    last = get_column_letter(1 + len(cols))  # data spans B..(B+ncols-1)
    put(ws, f"B{row}", title, f=font(11, True, WHITE), bg=NAVY,
        border=Border(bottom=MED))
    for col in range(2, 2 + len(cols)):
        ws.cell(row=row, column=col).fill = fill(NAVY)
        ws.cell(row=row, column=col).border = Border(bottom=MED)
    if subtitle:
        put(ws, f"{last}{row}", subtitle, f=font(9, False, WHITE), bg=NAVY, align="right")
    ws.row_dimensions[row].height = 22
    row += 1
    for i, (head, _, align, _) in enumerate(cols):
        put(ws, f"{get_column_letter(2+i)}{row}", head, f=font(10, True, WHITE),
            bg=MIDBLUE, align=align, border=BORDER_ALL)
    ws.row_dimensions[row].height = 24
    row += 1
    for rec in records:
        lines = 1
        for i, (_, extract, align, numfmt) in enumerate(cols):
            val = extract(rec)
            put(ws, f"{get_column_letter(2+i)}{row}", val, f=font(10, False, BLUE_INPUT),
                align=align, border=BORDER_ALL, numfmt=numfmt if isinstance(val, (int, float)) else None)
            if align == "left" and val:  # wrapped column — estimate how many lines it needs
                width = COL_WIDTHS.get(get_column_letter(2 + i), 12)
                cpl = max(8, width * 1.05)  # ~chars per line at Calibri 10
                lines = max(lines, math.ceil(len(str(val)) / cpl))
        # Explicit height sized to the tallest wrapped cell, so it renders correctly in
        # Excel, Numbers, and Quick Look alike (auto-fit can't be relied on everywhere).
        ws.row_dimensions[row].height = min(180, max(20, lines * 13.5 + 6))
        row += 1
    return row + 1  # trailing spacer row


def render(in_path, out_path, firm="[Firm Name]", date=None):
    meta, companies = load(in_path)
    groups = {"prospects": [], "leaders": [], "funded": []}
    for r in companies:
        groups[classify(r)].append(r)

    wb = Workbook()
    ws = wb.active
    # Excel forbids : \ / ? * [ ] in sheet titles and caps length at 31.
    ws.title = (re.sub(r"[:\\/?*\[\]]", " ", meta.get("theme") or "Market Map").strip()[:31]
                or "Market Map")
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # header block
    put(ws, "B2", "INVESTMENT THEME MARKET MAP", f=font(20, True, NAVY), wrap=False)
    put(ws, "B3", meta.get("theme", "[Theme Name]"), f=font(12, False, GREY, italic=True), wrap=False)
    put(ws, "J2", f"Prepared by: {firm}", f=font(9, False, GREY), align="right")
    put(ws, "J3", f"Date: {date or datetime.date.today().strftime('%m/%d/%Y')}",
        f=font(9, False, GREY), align="right")
    ws.row_dimensions[2].height = 28
    row = 5

    # The theme narrative (definition, use case, etc.) lives in the Market Brief and the JSONL
    # _theme_meta record; the xlsx is the company landscape, so it goes straight to the tables.
    row = section(ws, row, "FIRM FIT PROSPECTS", "Targets aligned to firm thesis",
                  PROSPECTS, groups["prospects"])
    row = section(ws, row, "MARKET LEADERS", "Category-defining incumbents",
                  LEADERS, groups["leaders"])
    row = section(ws, row, "FUNDED COMPANIES", "Venture-backed scale-ups",
                  FUNDED, groups["funded"])

    put(ws, f"B{row}", "Sources: company websites, analyst reports, and the cited links in the landscape JSONL.",
        f=font(8, False, "FF808080", italic=True))

    ws.freeze_panes = "A5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(out_path)
    n = {k: len(v) for k, v in groups.items()}
    print(f"Rendered {out_path}  (prospects={n['prospects']}, leaders={n['leaders']}, funded={n['funded']})")
    return out_path


def _cell_desc(c):
    """Compact style snapshot of a single cell, for the template manifest."""
    f, al = c.font, c.alignment
    return {
        "value": c.value,
        "font": f"Calibri {int(f.sz or 10)}" + (" bold" if f.b else "") + (" italic" if f.i else ""),
        "font_color": (f.color.rgb if f.color and f.color.rgb else None),
        "align": al.horizontal or "left",
        "wrap": bool(al.wrap_text),
    }


def make_template(out_xlsx, manifest_path, firm="[Firm Name]"):
    """Render the canonical EMPTY template, then derive the manifest from it. The renderer is
    the single source of truth — regenerate both with this after any layout change, never by hand."""
    meta = {"_theme_meta": True, "theme": "[Theme Name]"}
    with tempfile.TemporaryDirectory() as d:
        jp = Path(d) / "t.jsonl"
        jp.write_text(json.dumps(meta))
        render(jp, out_xlsx, firm=firm, date="[MM/DD/YYYY]")

    wb = load_workbook(out_xlsx)
    ws = wb.active
    widths = {col: round(dim.width, 2) for col, dim in ws.column_dimensions.items() if dim.width}
    header_block = {coord: _cell_desc(ws[coord]) for coord in ("B2", "B3", "J2", "J3")}

    titles = {"FIRM FIT PROSPECTS", "MARKET LEADERS", "FUNDED COMPANIES"}
    max_cols = 1 + max(len(PROSPECTS), len(LEADERS), len(FUNDED))
    sections = []
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value not in titles:
            continue
        hdr_row, cols, aligns = r + 1, {}, {}
        for c in range(2, max_cols + 2):
            v = ws.cell(hdr_row, c).value
            if v in (None, ""):
                continue
            L = get_column_letter(c)
            cols[L], aligns[L] = v, (ws.cell(hdr_row, c).alignment.horizontal or "left")
        subtitle = ws.cell(r, 1 + len(cols)).value
        sections.append({
            "name": ws.cell(r, 2).value,
            "header_cell": f"B{r}",
            "subtitle": subtitle if subtitle != ws.cell(r, 2).value else None,
            "column_header_row": hdr_row,
            "columns": cols,
            "column_alignment": aligns,
        })

    footer = {}
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.startswith("Sources:"):
            footer = {f"B{r}": _cell_desc(ws[f"B{r}"])}
            break

    manifest = {
        "manifest_version": "3.0",
        "describes": Path(out_xlsx).as_posix(),
        "template_sha256": hashlib.sha256(Path(out_xlsx).read_bytes()).hexdigest(),
        "generated_by": "skills/render-market-map/scripts/render_market_map.py --make-template",
        "do_not_edit": "Auto-derived from the renderer (the source of truth). Regenerate with --make-template after any layout change; never hand-edit.",
        "sheet": ws.title,
        "purpose": "Structural snapshot of the canonical empty template, so an agent can know its layout without opening the .xlsx live.",
        "fingerprint_use": "If template_sha256 doesn't match the file, the template changed — regenerate this manifest.",
        "layout_conventions": [
            "No Theme Overview block: the theme narrative lives in the Market Brief and the JSONL _theme_meta record; the xlsx is the company landscape only.",
            "Title rows (B2/B3) never wrap. Body-text columns wrap and each data row auto-sizes to its tallest cell.",
            "Confirmed $0 funding renders as 'Bootstrapped'; band-only headcount shows the band via headcount_display.",
        ],
        "global_style": {"font_name": "Calibri",
                         "note": "Value cells use blue font FF0000FF; data cells carry borders on all four sides."},
        "column_widths": widths,
        "header_block": header_block,
        "sections": sections,
        "footer": footer,
    }
    Path(manifest_path).write_text(json.dumps(manifest, indent=2) + "\n")
    print(f"Wrote template {out_xlsx} + manifest {manifest_path} "
          f"(sha {manifest['template_sha256'][:12]}…, {len(sections)} sections, no Theme Overview)")


def selftest():
    rows = [
        {"_theme_meta": True, "theme": "Example Theme", "definition": "A test definition.",
         "use_case": "Who buys it.", "market_size_growth": "$1.0B, 20% CAGR"},
        {"name": "Acme Target", "type": "company", "hq": "Austin, US", "founded_year": 2021,
         "fte": 40, "total_funding_usd": 0, "what_they_do": "Bootstrapped target.",
         "firm_fit_rationale": "Passes filters."},
        {"name": "Band Co", "type": "company", "founded_year": 2019, "fte": None,
         "headcount_display": "11-50", "what_they_do": "Band-only headcount.",
         "firm_fit_rationale": "Passes filters."},
        {"name": "BigCo", "type": "company", "ownership_status": "public (BIG)",
         "founded_year": 2008, "what_they_do": "Incumbent."},
        {"name": "ScaleUp", "type": "company", "last_round": "Series C 2024",
         "total_funding_usd": 60_000_000, "what_they_do": "VC-backed."},
    ]
    with tempfile.TemporaryDirectory() as d:
        jp, xp = Path(d) / "t.jsonl", Path(d) / "t.xlsx"
        jp.write_text("\n".join(json.dumps(r) for r in rows))
        render(jp, xp)
        wb = load_workbook(xp)
        ws = wb.active
        cells = [v for col in ws.iter_cols(values_only=True) for v in col]
        assert ws["B2"].value == "INVESTMENT THEME MARKET MAP"
        # Theme narrative lives in the brief + JSONL _theme_meta, not the xlsx.
        assert "A test definition." not in cells, "theme overview prose should not render in the xlsx"
        # Invariants: confirmed $0 funding reads 'Bootstrapped'; band-only headcount shows the band.
        assert "Bootstrapped" in cells, "confirmed $0 funding should render as 'Bootstrapped'"
        assert "11-50" in cells, "band-only headcount should show the band via headcount_display"
        for name in ("Acme Target", "Band Co", "BigCo", "ScaleUp"):
            assert name in cells, f"missing {name}"
    print("selftest: PASS")


def main():
    ap = argparse.ArgumentParser(description="Render a Market Map JSONL into a banker-grade .xlsx")
    ap.add_argument("--in", dest="inp", help="path to the landscape JSONL")
    ap.add_argument("--out", dest="out", help="path to write the .xlsx")
    ap.add_argument("--firm", default="[Firm Name]")
    ap.add_argument("--date", default=None, help="MM/DD/YYYY (defaults to today)")
    ap.add_argument("--selftest", action="store_true")
    ap.add_argument("--make-template", action="store_true",
                    help="regenerate the canonical template .xlsx + manifest from the renderer")
    ap.add_argument("--manifest", default="templates/market_map_template.manifest.json",
                    help="manifest path for --make-template")
    a = ap.parse_args()
    if a.selftest:
        selftest()
        return
    if a.make_template:
        out = a.out or "templates/Market_Map_Template.xlsx"
        make_template(out, a.manifest, firm=a.firm)
        return
    if not a.inp or not a.out:
        ap.error("--in and --out are required (or use --selftest / --make-template)")
    render(a.inp, a.out, firm=a.firm, date=a.date)


if __name__ == "__main__":
    main()
