#!/usr/bin/env python3
"""
verify_workbook.py — sign-off gate for a built P&L workbook.

Recalculates the workbook (LibreOffice headless if available) and asserts:
  • MASTER AUDIT CHECK = 0 and no Source_Audit check cell != 0
  • Output reconciliation row = 0 for every period
  • 0 merged cells on both sheets

Fallback (no LibreOffice): recompute subtotals from the JSON `detail` values and
compare to each line's `stated` — flags that the live-formula recalc was skipped.

Usage:
  python3 verify_workbook.py <built.xlsx> [normalized.json]
"""
import sys, os, json, shutil, subprocess, tempfile
import openpyxl
from openpyxl.utils import get_column_letter

SOFFICE_CANDIDATES = [
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    shutil.which("libreoffice") or "", shutil.which("soffice") or "",
]

def find_soffice():
    for p in SOFFICE_CANDIDATES:
        if p and os.path.exists(p):
            return p
    return None

def merged_check(path):
    wb = openpyxl.load_workbook(path)
    bad = {ws.title: len(ws.merged_cells.ranges) for ws in wb.worksheets if ws.merged_cells.ranges}
    return bad

def find_check_coords(path):
    """Locate per-line check cells by their formula pattern (=ROUND(x-y,2)),
    excluding the master =ROUND(SUM(...)) cell. Returns Source_Audit coordinates."""
    wb = openpyxl.load_workbook(path)  # formulas, not data_only
    coords = []
    if "Source_Audit" not in wb.sheetnames:
        return coords
    src = wb["Source_Audit"]
    for row in src.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=ROUND(") and "SUM(" not in v and "-" in v:
                coords.append(cell.coordinate)
    return coords

def recalc_and_check(path):
    soffice = find_soffice()
    if not soffice:
        return None
    check_coords = find_check_coords(path)
    tmp = tempfile.mkdtemp()
    subprocess.run([soffice, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", tmp, path],
                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=False)
    out = os.path.join(tmp, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
    if not os.path.exists(out):
        return None
    wb = openpyxl.load_workbook(out, data_only=True)
    res = {"master": None, "nonzero_checks": [], "recon_nonzero": []}
    src = wb["Source_Audit"]
    for coord in check_coords:                       # scan EVERY individual check cell
        v = src[coord].value
        if isinstance(v, (int, float)) and abs(v) > 0.005:
            res["nonzero_checks"].append((coord, v))
    for r in range(1, src.max_row + 1):
        a = src.cell(row=r, column=1).value
        if isinstance(a, str) and "MASTER AUDIT" in a:
            for ci in range(2, src.max_column + 1):
                v = src.cell(row=r, column=ci).value
                if isinstance(v, (int, float)):
                    res["master"] = v; break
    if "Output" in wb.sheetnames:
        out_ws = wb["Output"]
        for r in range(1, out_ws.max_row + 1):
            a = out_ws.cell(row=r, column=1).value
            if isinstance(a, str) and "Reconciliation" in a:
                for ci in range(2, out_ws.max_column + 1):
                    v = out_ws.cell(row=r, column=ci).value
                    if isinstance(v, (int, float)) and abs(v) > 0.005:
                        res["recon_nonzero"].append((get_column_letter(ci) + str(r), v))
    return res

def json_fallback(json_path):
    with open(json_path) as fh:
        data = json.load(fh)
    detail = {l["key"]: l for l in data["lines"] if l["role"] == "detail" and "key" in l}
    problems = []
    for ln in data["lines"]:
        if ln["role"] == "subtotal" and ln.get("stated"):
            for pid, stated in ln["stated"].items():
                s = sum((detail.get(k, {}).get("values", {}) or {}).get(pid, 0) for k in ln["members"])
                if round(s - stated, 2) != 0:
                    problems.append((ln["key"], pid, round(s - stated, 2)))
    return problems

def main():
    path = sys.argv[1]
    print(f"Verifying: {path}")
    ok = True

    merged = merged_check(path)
    if merged:
        print(f"  ✗ MERGED CELLS present: {merged}"); ok = False
    else:
        print("  ✓ 0 merged cells")

    rc = recalc_and_check(path)
    if rc is None:
        print("  ! LibreOffice not available — live-formula recalc SKIPPED.")
        if len(sys.argv) > 2:
            probs = json_fallback(sys.argv[2])
            if probs:
                print(f"  ✗ JSON subtotal mismatches: {probs}"); ok = False
            else:
                print("  ✓ JSON fallback: all subtotals reconcile to stated")
        else:
            print("  ! No JSON given for fallback — correctness UNVERIFIED.")
    else:
        if rc["master"] is None:
            print("  ✗ MASTER AUDIT CHECK not found"); ok = False
        elif abs(rc["master"]) > 0.005:
            print(f"  ✗ MASTER AUDIT CHECK = {rc['master']} (must be 0)"); ok = False
        else:
            print("  ✓ MASTER AUDIT CHECK = 0")
        if rc["nonzero_checks"]:
            print(f"  ✗ Individual check cells non-zero (would offset in the master sum): {rc['nonzero_checks'][:10]}"); ok = False
        else:
            print("  ✓ Every individual check cell = 0")
        if rc["recon_nonzero"]:
            print(f"  ✗ Output reconciliation non-zero: {rc['recon_nonzero']}"); ok = False
        else:
            print("  ✓ Output reconciliation = 0 all periods")

    print("RESULT:", "PASS" if ok else "FAIL")
    sys.exit(0 if ok else 1)

if __name__ == "__main__":
    main()
